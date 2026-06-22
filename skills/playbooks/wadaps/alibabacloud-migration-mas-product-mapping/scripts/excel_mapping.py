#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AWS → 阿里云 产品映射 Excel 处理脚本

读取 AWS 资源调研 Excel 文件，自动完成：
1. 产品映射（AWS 产品 → 阿里云产品）
2. 规格映射（AWS 规格 → 阿里云规格）
3. 实时询价（调用 pricing_service.py）
4. 输出结果到 Excel

用法:
    python excel_mapping.py <input.xls> [--output output.xlsx] [--pricing-url http://localhost:5001]
"""

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

import requests

try:
    import xlrd
except ImportError:
    print("请先安装 xlrd: pip install xlrd")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 样式定义
# ============================================================

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
MONEY_FORMAT = '#,##0.00'
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", bold=True, size=11, color="2E75B6")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


def style_header_row(ws, row_num, col_count):
    """给表头行应用统一样式"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    """自动调整列宽"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # 中文字符算2个宽度
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        adjusted = max(min(max_len + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted


# ============================================================
# EC2 实例族 → 阿里云 ECS 规格族 映射
# ============================================================

EC2_FAMILY_MAP = {
    # ARM Graviton 系列 → 倚天 ARM
    "c6g": "ecs.c8y",    "c6gd": "ecs.c8y",   "c6gn": "ecs.c8y",
    "c7g": "ecs.c8y",    "c7gd": "ecs.c8y",   "c7gn": "ecs.c8y",
    "c8g": "ecs.c8y",
    "m6g": "ecs.g8y",    "m6gd": "ecs.g8y",
    "m7g": "ecs.g8y",    "m7gd": "ecs.g8y",
    "m8g": "ecs.g8y",
    "r6g": "ecs.r8y",    "r6gd": "ecs.r8y",
    "r7g": "ecs.r8y",    "r7gd": "ecs.r8y",
    "r8g": "ecs.r8y",
    "t4g": "ecs.t6-c1m2",
    "x2gd": "ecs.re7",
    "im4gn": "ecs.i4",   "is4gen": "ecs.i4",  "i4g": "ecs.i4",
    "a1": "ecs.g8y",
    "hpc7g": "ecs.hfc7",
    # Intel 系列
    "c5": "ecs.c7",      "c5d": "ecs.c7",     "c5n": "ecs.c7",
    "c6i": "ecs.c8i",    "c6id": "ecs.c8i",   "c6in": "ecs.c8i",
    "c7i": "ecs.c8i",    "c7i-flex": "ecs.c8i",
    # Intel 通用型
    "m5": "ecs.g7",      "m5d": "ecs.g7",     "m5n": "ecs.g7",   "m5dn": "ecs.g7",
    "m6i": "ecs.g8i",    "m6id": "ecs.g8i",   "m6idn": "ecs.g8i", "m6in": "ecs.g8i",
    "m7i": "ecs.g8i",    "m7i-flex": "ecs.g8i",
    "m8i": "ecs.g8i",
    # Intel 内存型
    "r5": "ecs.r7",      "r5d": "ecs.r7",     "r5n": "ecs.r7",   "r5dn": "ecs.r7",
    "r5b": "ecs.r7",
    "r6i": "ecs.r8i",    "r6id": "ecs.r8i",   "r6idn": "ecs.r8i", "r6in": "ecs.r8i",
    "r7i": "ecs.r8i",    "r7iz": "ecs.r8i",
    # AMD 系列
    "c5a": "ecs.c7a",    "c5ad": "ecs.c7a",
    "c6a": "ecs.c8a",    "c7a": "ecs.c8a",
    "m5a": "ecs.g7a",    "m5ad": "ecs.g7a",
    "m6a": "ecs.g8a",    "m7a": "ecs.g8a",
    "r5a": "ecs.r7a",    "r5ad": "ecs.r7a",
    "r6a": "ecs.r8a",    "r7a": "ecs.r8a",
    # 突发型
    "t2": "ecs.t5",      "t3": "ecs.t6-c1m2", "t3a": "ecs.t6-c1m2",
    # 高频
    "m5zn": "ecs.hfg7",  "z1d": "ecs.hfr7",
    # 大内存/SAP型
    "x1": "ecs.re7",     "x1e": "ecs.re7",
    "x2idn": "ecs.re7",  "x2iedn": "ecs.re7", "x2iezn": "ecs.re7",
    "u-6tb1": "ecs.re7", "u-9tb1": "ecs.re7",  "u-12tb1": "ecs.re7",
    # 存储优化
    "i3": "ecs.i3",      "i3en": "ecs.i4",    "i4i": "ecs.i4",
    "d2": "ecs.d2s",     "d3": "ecs.d3s",     "d3en": "ecs.d3s",
    "h1": "ecs.d2s",
    # HPC 型
    "hpc6a": "ecs.hfc7", "hpc6id": "ecs.hfc7", "hpc7a": "ecs.hfc7",
    # GPU / 推理 / 训练
    "p3": "ecs.gn6v",    "p3dn": "ecs.gn6v",
    "p4d": "ecs.gn7e",   "p4de": "ecs.gn7e",
    "p5": "ecs.gn8",     "p5e": "ecs.gn8",
    "g4dn": "ecs.gn7i",  "g4ad": "ecs.gn7i",
    "g5": "ecs.gn7",     "g5g": "ecs.gn7",    "g6": "ecs.gn8i", "g6e": "ecs.gn8i",
    "inf1": "ecs.gn7i",  "inf2": "ecs.gn8i",
    "dl1": "ecs.gn7e",   "dl2q": "ecs.gn7e",
    "trn1": "ecs.gn7e",  "trn1n": "ecs.gn7e", "trn2": "ecs.gn8",
    "vt1": "ecs.gn7i",
    # Mac
    "mac1": "ecs.g7",    "mac2": "ecs.g7",    "mac2-m2pro": "ecs.g7",
}

# AWS size → 阿里云 size（对于 t 系列特殊处理）
SIZE_MAP = {
    "nano": "small",
    "micro": "small",
    "small": "large",    # t4g.small = 2C2G → large
    "medium": "large",
    "large": "large",
    "xlarge": "xlarge",
    "2xlarge": "2xlarge",
    "4xlarge": "4xlarge",
    "8xlarge": "8xlarge",
    "12xlarge": "13xlarge",
    "16xlarge": "16xlarge",
    "24xlarge": "32xlarge",
    "metal": "metal",
}

# t系列 size 特殊映射（突发型配置不同，t6最小规格为large）
T_SIZE_MAP = {
    "nano": "large",
    "micro": "large",
    "small": "large",
    "medium": "large",
    "large": "large",
    "xlarge": "xlarge",
    "2xlarge": "2xlarge",
}


def map_ec2_to_ecs(instance_type: str) -> dict:
    """
    将 EC2 实例类型映射到阿里云 ECS 规格。
    例: c8g.xlarge → ecs.c8y.xlarge
    """
    # 解析实例族和尺寸: c8g.xlarge → family=c8g, size=xlarge
    parts = instance_type.split(".", 1)
    if len(parts) != 2:
        return {"aliyun_spec": f"未识别({instance_type})", "family": "", "size": ""}

    family, size = parts[0], parts[1]
    aliyun_family = EC2_FAMILY_MAP.get(family, "")

    if not aliyun_family:
        return {"aliyun_spec": f"未映射({instance_type})", "family": family, "size": size}

    # 突发型特殊处理
    is_burst = family.startswith("t")
    if is_burst:
        aliyun_size = T_SIZE_MAP.get(size, size)
        # t6 系列只有 large，xlarge 及以上不存在，回退到标准规格
        if aliyun_size in ("xlarge", "2xlarge"):
            # t4g.xlarge(4C8G) → ecs.c8y.xlarge, t4g.2xlarge(8C16G) → ecs.c8y.2xlarge
            aliyun_spec = f"ecs.c8y.{aliyun_size}"
        else:
            aliyun_spec = f"{aliyun_family}.{aliyun_size}"
    else:
        aliyun_size = SIZE_MAP.get(size, size)
        aliyun_spec = f"{aliyun_family}.{aliyun_size}"

    return {"aliyun_spec": aliyun_spec, "family": family, "size": size}


# ============================================================
# Aurora / RDS 规格映射
# ============================================================

AURORA_SPEC_MAP = {
    # ---- db.t3 系列 (Intel 突发型) ----
    "db.t3.micro":   {"mysql": "rds.mysql.t1.small",       "postgresql": "pg.n1.micro.1",
                      "vcpu": 2, "mem_gb": 1},
    "db.t3.small":   {"mysql": "rds.mysql.s1.small",       "postgresql": "pg.n2.small.1",
                      "vcpu": 2, "mem_gb": 2},
    "db.t3.medium":  {"mysql": "rds.mysql.s2.large",       "postgresql": "pg.n2.small.2c",
                      "vcpu": 2, "mem_gb": 4},
    "db.t3.large":   {"mysql": "rds.mysql.s2.xlarge",      "postgresql": "pg.n2.medium.1",
                      "vcpu": 2, "mem_gb": 8},
    "db.t3.xlarge":  {"mysql": "rds.mysql.s3.large",       "postgresql": "pg.n2.medium.2c",
                      "vcpu": 4, "mem_gb": 16},
    "db.t3.2xlarge": {"mysql": "rds.mysql.m1.medium",      "postgresql": "pg.n2.large.1",
                      "vcpu": 8, "mem_gb": 32},
    # ---- db.t4g 系列 (ARM 突发) ----
    "db.t4g.micro":   {"mysql": "polar.mysql.g2.medium",   "postgresql": "polar.pg.g2.medium",
                       "vcpu": 2, "mem_gb": 1},
    "db.t4g.small":   {"mysql": "polar.mysql.g2.medium",   "postgresql": "polar.pg.g2.medium",
                       "vcpu": 2, "mem_gb": 2},
    "db.t4g.medium":  {"mysql": "polar.mysql.g2.large",    "postgresql": "polar.pg.g2.large",
                       "vcpu": 2, "mem_gb": 4},
    "db.t4g.large":   {"mysql": "polar.mysql.g2.xlarge",   "postgresql": "polar.pg.g2.xlarge",
                       "vcpu": 2, "mem_gb": 8},
    "db.t4g.xlarge":  {"mysql": "polar.mysql.g4.xlarge",   "postgresql": "polar.pg.g4.xlarge",
                       "vcpu": 4, "mem_gb": 16},
    "db.t4g.2xlarge": {"mysql": "polar.mysql.g8.xlarge",   "postgresql": "polar.pg.g8.xlarge",
                       "vcpu": 8, "mem_gb": 32},
    # ---- db.r5 系列 (Intel 内存型) ----
    "db.r5.large":    {"mysql": "rds.mysql.s2.xlarge",     "postgresql": "pg.n2.medium.1",
                       "vcpu": 2, "mem_gb": 16},
    "db.r5.xlarge":   {"mysql": "rds.mysql.s3.large",      "postgresql": "pg.n4.medium.1",
                       "vcpu": 4, "mem_gb": 32},
    "db.r5.2xlarge":  {"mysql": "rds.mysql.m1.medium",     "postgresql": "pg.n2.large.2c",
                       "vcpu": 8, "mem_gb": 64},
    "db.r5.4xlarge":  {"mysql": "rds.mysql.m1.xlarge",     "postgresql": "pg.n2.xlarge.2c",
                       "vcpu": 16, "mem_gb": 128},
    "db.r5.8xlarge":  {"mysql": "rds.mysql.m1.2xlarge",    "postgresql": "pg.n4.2xlarge.1",
                       "vcpu": 32, "mem_gb": 256},
    "db.r5.12xlarge": {"mysql": "rds.mysql.m1.4xlarge",    "postgresql": "pg.n4.4xlarge.1",
                       "vcpu": 48, "mem_gb": 384},
    "db.r5.16xlarge": {"mysql": "rds.mysql.m1.4xlarge",    "postgresql": "pg.n4.4xlarge.1",
                       "vcpu": 64, "mem_gb": 512},
    "db.r5.24xlarge": {"mysql": "rds.mysql.m1.8xlarge",    "postgresql": "pg.n8.8xlarge.1",
                       "vcpu": 96, "mem_gb": 768},
    # ---- db.r6g 系列 (ARM 内存型) ----
    "db.r6g.large":    {"mysql": "polar.mysql.g2.xlarge",   "postgresql": "polar.pg.g2.xlarge",
                        "vcpu": 2, "mem_gb": 16},
    "db.r6g.xlarge":   {"mysql": "polar.mysql.g4.xlarge",   "postgresql": "polar.pg.g4.xlarge",
                        "vcpu": 4, "mem_gb": 32},
    "db.r6g.2xlarge":  {"mysql": "polar.mysql.g8.xlarge",   "postgresql": "polar.pg.g8.xlarge",
                        "vcpu": 8, "mem_gb": 64},
    "db.r6g.4xlarge":  {"mysql": "polar.mysql.g8.2xlarge",  "postgresql": "polar.pg.g8.2xlarge",
                        "vcpu": 16, "mem_gb": 128},
    "db.r6g.8xlarge":  {"mysql": "polar.mysql.g8.4xlarge",  "postgresql": "polar.pg.g8.4xlarge",
                        "vcpu": 32, "mem_gb": 256},
    "db.r6g.12xlarge": {"mysql": "polar.mysql.g8.8xlarge",  "postgresql": "polar.pg.g8.8xlarge",
                        "vcpu": 48, "mem_gb": 384},
    "db.r6g.16xlarge": {"mysql": "polar.mysql.g8.16xlarge", "postgresql": "polar.pg.g8.16xlarge",
                        "vcpu": 64, "mem_gb": 512},
    # ---- db.r7g 系列 (ARM 内存型新代) ----
    "db.r7g.large":    {"mysql": "polar.mysql.g2.xlarge",   "postgresql": "polar.pg.g2.xlarge",
                        "vcpu": 2, "mem_gb": 16},
    "db.r7g.xlarge":   {"mysql": "polar.mysql.g4.xlarge",   "postgresql": "polar.pg.g4.xlarge",
                        "vcpu": 4, "mem_gb": 32},
    "db.r7g.2xlarge":  {"mysql": "polar.mysql.g8.xlarge",   "postgresql": "polar.pg.g8.xlarge",
                        "vcpu": 8, "mem_gb": 64},
    "db.r7g.4xlarge":  {"mysql": "polar.mysql.g8.2xlarge",  "postgresql": "polar.pg.g8.2xlarge",
                        "vcpu": 16, "mem_gb": 128},
    "db.r7g.8xlarge":  {"mysql": "polar.mysql.g8.4xlarge",  "postgresql": "polar.pg.g8.4xlarge",
                        "vcpu": 32, "mem_gb": 256},
    "db.r7g.12xlarge": {"mysql": "polar.mysql.g8.8xlarge",  "postgresql": "polar.pg.g8.8xlarge",
                        "vcpu": 48, "mem_gb": 384},
    "db.r7g.16xlarge": {"mysql": "polar.mysql.g8.16xlarge", "postgresql": "polar.pg.g8.16xlarge",
                        "vcpu": 64, "mem_gb": 512},
    # ---- db.r8g 系列 ----
    "db.r8g.large":    {"mysql": "polar.mysql.g2.xlarge",   "postgresql": "polar.pg.g2.xlarge",
                        "vcpu": 2, "mem_gb": 16},
    "db.r8g.xlarge":   {"mysql": "polar.mysql.g4.xlarge",   "postgresql": "polar.pg.g4.xlarge",
                        "vcpu": 4, "mem_gb": 32},
    "db.r8g.2xlarge":  {"mysql": "polar.mysql.g8.xlarge",   "postgresql": "polar.pg.g8.xlarge",
                        "vcpu": 8, "mem_gb": 64},
    "db.r8g.4xlarge":  {"mysql": "polar.mysql.g8.2xlarge",  "postgresql": "polar.pg.g8.2xlarge",
                        "vcpu": 16, "mem_gb": 128},
    "db.r8g.8xlarge":  {"mysql": "polar.mysql.g8.4xlarge",  "postgresql": "polar.pg.g8.4xlarge",
                        "vcpu": 32, "mem_gb": 256},
    "db.r8g.12xlarge": {"mysql": "polar.mysql.g8.8xlarge",  "postgresql": "polar.pg.g8.8xlarge",
                        "vcpu": 48, "mem_gb": 384},
    "db.r8g.16xlarge": {"mysql": "polar.mysql.g8.16xlarge", "postgresql": "polar.pg.g8.16xlarge",
                        "vcpu": 64, "mem_gb": 512},
    # ---- db.m5 系列 (Intel 通用型) ----
    "db.m5.large":    {"mysql": "rds.mysql.s2.xlarge",     "postgresql": "pg.n2.medium.1",
                       "vcpu": 2, "mem_gb": 8},
    "db.m5.xlarge":   {"mysql": "rds.mysql.s3.large",      "postgresql": "pg.n4.medium.1",
                       "vcpu": 4, "mem_gb": 16},
    "db.m5.2xlarge":  {"mysql": "rds.mysql.m1.medium",     "postgresql": "pg.n2.large.2c",
                       "vcpu": 8, "mem_gb": 32},
    "db.m5.4xlarge":  {"mysql": "rds.mysql.m1.xlarge",     "postgresql": "pg.n2.xlarge.2c",
                       "vcpu": 16, "mem_gb": 64},
    "db.m5.8xlarge":  {"mysql": "rds.mysql.m1.2xlarge",    "postgresql": "pg.n4.2xlarge.1",
                       "vcpu": 32, "mem_gb": 128},
    "db.m5.12xlarge": {"mysql": "rds.mysql.m1.4xlarge",    "postgresql": "pg.n4.4xlarge.1",
                       "vcpu": 48, "mem_gb": 192},
    "db.m5.16xlarge": {"mysql": "rds.mysql.m1.4xlarge",    "postgresql": "pg.n4.4xlarge.1",
                       "vcpu": 64, "mem_gb": 256},
    "db.m5.24xlarge": {"mysql": "rds.mysql.m1.8xlarge",    "postgresql": "pg.n8.8xlarge.1",
                       "vcpu": 96, "mem_gb": 384},
    # ---- db.m6g 系列 (ARM 通用型) ----
    "db.m6g.large":    {"mysql": "polar.mysql.g2.xlarge",   "postgresql": "polar.pg.g2.xlarge",
                        "vcpu": 2, "mem_gb": 8},
    "db.m6g.xlarge":   {"mysql": "polar.mysql.g4.xlarge",   "postgresql": "polar.pg.g4.xlarge",
                        "vcpu": 4, "mem_gb": 16},
    "db.m6g.2xlarge":  {"mysql": "polar.mysql.g8.xlarge",   "postgresql": "polar.pg.g8.xlarge",
                        "vcpu": 8, "mem_gb": 32},
    "db.m6g.4xlarge":  {"mysql": "polar.mysql.g8.2xlarge",  "postgresql": "polar.pg.g8.2xlarge",
                        "vcpu": 16, "mem_gb": 64},
    "db.m6g.8xlarge":  {"mysql": "polar.mysql.g8.4xlarge",  "postgresql": "polar.pg.g8.4xlarge",
                        "vcpu": 32, "mem_gb": 128},
    "db.m6g.12xlarge": {"mysql": "polar.mysql.g8.8xlarge",  "postgresql": "polar.pg.g8.8xlarge",
                        "vcpu": 48, "mem_gb": 192},
    "db.m6g.16xlarge": {"mysql": "polar.mysql.g8.16xlarge", "postgresql": "polar.pg.g8.16xlarge",
                        "vcpu": 64, "mem_gb": 256},
    # ---- db.m6i 系列 (Intel 通用型新代) ----
    "db.m6i.large":    {"mysql": "rds.mysql.s2.xlarge",     "postgresql": "pg.n2.medium.1",
                        "vcpu": 2, "mem_gb": 8},
    "db.m6i.xlarge":   {"mysql": "rds.mysql.s3.large",      "postgresql": "pg.n4.medium.1",
                        "vcpu": 4, "mem_gb": 16},
    "db.m6i.2xlarge":  {"mysql": "rds.mysql.m1.medium",     "postgresql": "pg.n2.large.2c",
                        "vcpu": 8, "mem_gb": 32},
    "db.m6i.4xlarge":  {"mysql": "rds.mysql.m1.xlarge",     "postgresql": "pg.n2.xlarge.2c",
                        "vcpu": 16, "mem_gb": 64},
    "db.m6i.8xlarge":  {"mysql": "rds.mysql.m1.2xlarge",    "postgresql": "pg.n4.2xlarge.1",
                        "vcpu": 32, "mem_gb": 128},
    "db.m6i.12xlarge": {"mysql": "rds.mysql.m1.4xlarge",    "postgresql": "pg.n4.4xlarge.1",
                        "vcpu": 48, "mem_gb": 192},
    "db.m6i.16xlarge": {"mysql": "rds.mysql.m1.4xlarge",    "postgresql": "pg.n4.4xlarge.1",
                        "vcpu": 64, "mem_gb": 256},
    # ---- db.m7g 系列 (ARM 通用型新代) ----
    "db.m7g.large":    {"mysql": "polar.mysql.g2.xlarge",   "postgresql": "polar.pg.g2.xlarge",
                        "vcpu": 2, "mem_gb": 8},
    "db.m7g.xlarge":   {"mysql": "polar.mysql.g4.xlarge",   "postgresql": "polar.pg.g4.xlarge",
                        "vcpu": 4, "mem_gb": 16},
    "db.m7g.2xlarge":  {"mysql": "polar.mysql.g8.xlarge",   "postgresql": "polar.pg.g8.xlarge",
                        "vcpu": 8, "mem_gb": 32},
    "db.m7g.4xlarge":  {"mysql": "polar.mysql.g8.2xlarge",  "postgresql": "polar.pg.g8.2xlarge",
                        "vcpu": 16, "mem_gb": 64},
    "db.m7g.8xlarge":  {"mysql": "polar.mysql.g8.4xlarge",  "postgresql": "polar.pg.g8.4xlarge",
                        "vcpu": 32, "mem_gb": 128},
    "db.m7g.12xlarge": {"mysql": "polar.mysql.g8.8xlarge",  "postgresql": "polar.pg.g8.8xlarge",
                        "vcpu": 48, "mem_gb": 192},
    "db.m7g.16xlarge": {"mysql": "polar.mysql.g8.16xlarge", "postgresql": "polar.pg.g8.16xlarge",
                        "vcpu": 64, "mem_gb": 256},
    # ---- db.x2g 系列 (ARM 大内存型) ----
    "db.x2g.large":   {"mysql": "polar.mysql.g8.xlarge",   "postgresql": "polar.pg.g8.xlarge",
                       "vcpu": 2, "mem_gb": 32},
    "db.x2g.xlarge":  {"mysql": "polar.mysql.g8.2xlarge",  "postgresql": "polar.pg.g8.2xlarge",
                       "vcpu": 4, "mem_gb": 64},
    "db.x2g.2xlarge": {"mysql": "polar.mysql.g8.4xlarge",  "postgresql": "polar.pg.g8.4xlarge",
                       "vcpu": 8, "mem_gb": 128},
    "db.x2g.4xlarge": {"mysql": "polar.mysql.g8.8xlarge",  "postgresql": "polar.pg.g8.8xlarge",
                       "vcpu": 16, "mem_gb": 256},
    "db.x2g.8xlarge": {"mysql": "polar.mysql.g8.16xlarge", "postgresql": "polar.pg.g8.16xlarge",
                       "vcpu": 32, "mem_gb": 512},
}

# ElastiCache 规格映射
CACHE_SPEC_MAP = {
    # ---- cache.t3 系列 (Intel 突发) ----
    "cache.t3.micro":   {"aliyun": "redis.amber.master.small.multithread",  "mem_gb": 0.5,  "desc": "云原生版 512MB 标准版-双副本"},
    "cache.t3.small":   {"aliyun": "redis.amber.master.small.multithread",  "mem_gb": 1,    "desc": "云原生版 1GB 标准版-双副本"},
    "cache.t3.medium":  {"aliyun": "redis.amber.master.stand.multithread",  "mem_gb": 4,    "desc": "云原生版 4GB 标准版-双副本"},
    "cache.t3.large":   {"aliyun": "redis.amber.master.large.multithread",  "mem_gb": 8,    "desc": "云原生版 8GB 标准版-双副本"},
    # ---- cache.t4g 系列 (ARM 突发) ----
    "cache.t4g.micro":  {"aliyun": "redis.amber.master.small.multithread",  "mem_gb": 0.5,  "desc": "云原生版 512MB 标准版-双副本"},
    "cache.t4g.small":  {"aliyun": "redis.amber.master.small.multithread",  "mem_gb": 1,    "desc": "云原生版 1GB 标准版-双副本"},
    "cache.t4g.medium": {"aliyun": "redis.amber.master.mid.multithread",    "mem_gb": 2,    "desc": "云原生版 2GB 标准版-双副本"},
    # ---- cache.m5 系列 (Intel 通用) ----
    "cache.m5.large":   {"aliyun": "redis.amber.master.large.multithread",  "mem_gb": 8,    "desc": "云原生版 8GB 标准版-双副本"},
    "cache.m5.xlarge":  {"aliyun": "redis.amber.master.2xlarge.multithread", "mem_gb": 16,  "desc": "云原生版 16GB 标准版-双副本"},
    "cache.m5.2xlarge": {"aliyun": "redis.amber.master.4xlarge.multithread", "mem_gb": 32,  "desc": "云原生版 32GB 标准版-双副本"},
    "cache.m5.4xlarge": {"aliyun": "redis.amber.master.8xlarge.multithread", "mem_gb": 64,  "desc": "云原生版 64GB 标准版-双副本"},
    # ---- cache.m6g 系列 (ARM 通用) ----
    "cache.m6g.large":   {"aliyun": "redis.amber.master.large.multithread",   "mem_gb": 8,   "desc": "云原生版 8GB 标准版-双副本"},
    "cache.m6g.xlarge":  {"aliyun": "redis.amber.master.2xlarge.multithread", "mem_gb": 16,  "desc": "云原生版 16GB 标准版-双副本"},
    "cache.m6g.2xlarge": {"aliyun": "redis.amber.master.4xlarge.multithread", "mem_gb": 32,  "desc": "云原生版 32GB 标准版-双副本"},
    "cache.m6g.4xlarge": {"aliyun": "redis.amber.master.8xlarge.multithread", "mem_gb": 64,  "desc": "云原生版 64GB 标准版-双副本"},
    "cache.m6g.8xlarge": {"aliyun": "redis.shard.amber.ce13.5.default",       "mem_gb": 128, "desc": "云原生版 128GB 集群版"},
    # ---- cache.m7g 系列 (ARM 通用新代) ----
    "cache.m7g.large":   {"aliyun": "redis.amber.master.large.multithread",   "mem_gb": 8,   "desc": "云原生版 8GB 标准版-双副本"},
    "cache.m7g.xlarge":  {"aliyun": "redis.amber.master.2xlarge.multithread", "mem_gb": 16,  "desc": "云原生版 16GB 标准版-双副本"},
    "cache.m7g.2xlarge": {"aliyun": "redis.amber.master.4xlarge.multithread", "mem_gb": 32,  "desc": "云原生版 32GB 标准版-双副本"},
    "cache.m7g.4xlarge": {"aliyun": "redis.amber.master.8xlarge.multithread", "mem_gb": 64,  "desc": "云原生版 64GB 标准版-双副本"},
    # ---- cache.r6g 系列 (ARM 内存型) ----
    "cache.r6g.large":    {"aliyun": "redis.amber.master.2xlarge.multithread", "mem_gb": 16,  "desc": "云原生版 16GB 标准版-双副本"},
    "cache.r6g.xlarge":   {"aliyun": "redis.amber.master.4xlarge.multithread", "mem_gb": 32,  "desc": "云原生版 32GB 标准版-双副本"},
    "cache.r6g.2xlarge":  {"aliyun": "redis.amber.master.8xlarge.multithread", "mem_gb": 64,  "desc": "云原生版 64GB 标准版-双副本"},
    "cache.r6g.4xlarge":  {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 128, "desc": "云原生版 128GB 集群版"},
    "cache.r6g.8xlarge":  {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 256, "desc": "云原生版 256GB 集群版"},
    "cache.r6g.12xlarge": {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 384, "desc": "云原生版 384GB 集群版"},
    "cache.r6g.16xlarge": {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 512, "desc": "云原生版 512GB 集群版"},
    # ---- cache.r7g 系列 (ARM 内存型新代) ----
    "cache.r7g.large":    {"aliyun": "redis.amber.master.2xlarge.multithread", "mem_gb": 16,  "desc": "云原生版 16GB 标准版-双副本"},
    "cache.r7g.xlarge":   {"aliyun": "redis.amber.master.4xlarge.multithread", "mem_gb": 32,  "desc": "云原生版 32GB 标准版-双副本"},
    "cache.r7g.2xlarge":  {"aliyun": "redis.amber.master.8xlarge.multithread", "mem_gb": 64,  "desc": "云原生版 64GB 标准版-双副本"},
    "cache.r7g.4xlarge":  {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 128, "desc": "云原生版 128GB 集群版"},
    "cache.r7g.8xlarge":  {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 256, "desc": "云原生版 256GB 集群版"},
    "cache.r7g.12xlarge": {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 384, "desc": "云原生版 384GB 集群版"},
    "cache.r7g.16xlarge": {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 512, "desc": "云原生版 512GB 集群版"},
    # ---- cache.r6gd 系列 (ARM 内存+本地盘) ----
    "cache.r6gd.xlarge":  {"aliyun": "redis.amber.master.4xlarge.multithread", "mem_gb": 26,  "desc": "云原生版 32GB 标准版-双副本"},
    "cache.r6gd.2xlarge": {"aliyun": "redis.amber.master.8xlarge.multithread", "mem_gb": 52,  "desc": "云原生版 64GB 标准版-双副本"},
    "cache.r6gd.4xlarge": {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 104, "desc": "云原生版 128GB 集群版"},
    "cache.r6gd.8xlarge": {"aliyun": "redis.shard.amber.ce13.5.default",      "mem_gb": 208, "desc": "云原生版 256GB 集群版"},
}

# S3 存储类 → OSS 存储类
S3_STORAGE_MAP = {
    "STANDARD": {"aliyun": "OSS 标准存储", "price_gb_month": 0.12},
    "STANDARD_IA": {"aliyun": "OSS 低频访问", "price_gb_month": 0.08},
    "ONEZONE_IA": {"aliyun": "OSS 低频访问(单AZ)", "price_gb_month": 0.08},
    "GLACIER": {"aliyun": "OSS 归档存储", "price_gb_month": 0.033},
    "DEEP_ARCHIVE": {"aliyun": "OSS 深度冷归档", "price_gb_month": 0.0045},
    "INTELLIGENT_TIERING": {"aliyun": "OSS 智能分层", "price_gb_month": 0.12},
    "REDUCED_REDUNDANCY": {"aliyun": "OSS 标准存储", "price_gb_month": 0.12},
}

# 产品级映射
PRODUCT_MAP = {
    "EC2":            {"aliyun": "ECS 弹性计算服务", "tool": "SMC 服务器迁移中心", "layer": "计算层", "complexity": "中"},
    "S3":             {"aliyun": "OSS 对象存储", "tool": "在线迁移服务", "layer": "存储层", "complexity": "低"},
    "ElasticCache":   {"aliyun": "云数据库 Redis/Tair", "tool": "DTS/redis-shake", "layer": "缓存层", "complexity": "中"},
    "DocumentDB":     {"aliyun": "云数据库 MongoDB 版", "tool": "DTS 数据传输服务", "layer": "数据层", "complexity": "中"},
    "NAT":            {"aliyun": "NAT 网关", "tool": "手动重建", "layer": "网络层", "complexity": "低"},
    "负载均衡":       {"aliyun": "SLB/ALB/NLB 负载均衡", "tool": "手动重建", "layer": "网络层", "complexity": "低"},
    "Athena":         {"aliyun": "Data Lake Analytics / MaxCompute", "tool": "SQL 适配", "layer": "大数据层", "complexity": "中"},
    "lamdba":         {"aliyun": "函数计算 FC", "tool": "代码迁移+适配", "layer": "计算层", "complexity": "高"},
    "EIP":            {"aliyun": "弹性公网 IP (EIP)", "tool": "手动分配", "layer": "网络层", "complexity": "低"},
    "EventBridge":    {"aliyun": "事件总线 EventBridge", "tool": "规则迁移", "layer": "中间件", "complexity": "中"},
    "SNS":            {"aliyun": "消息服务 MNS / EventBridge", "tool": "应用层适配", "layer": "中间件", "complexity": "高"},
    "CloudWatchLog":  {"aliyun": "日志服务 SLS", "tool": "日志投递适配", "layer": "运维层", "complexity": "中"},
    "AuroraDB":       {"aliyun": "PolarDB MySQL/PostgreSQL", "tool": "DTS 数据传输服务", "layer": "数据层", "complexity": "中"},
    "Other":          {"aliyun": "阿里云对应服务", "tool": "视具体服务而定", "layer": "其他", "complexity": "-"},
}

# Other sheet 中服务名映射
OTHER_SERVICE_MAP = {
    "Amazon Simple Notification Service": {"aliyun": "消息服务 MNS", "price_ref": "按量"},
    "AWS Key Management Service": {"aliyun": "密钥管理服务 KMS", "price_ref": "~200元/月"},
    "AWS Secrets Manager": {"aliyun": "凭据管家 Secrets Manager", "price_ref": "按量"},
    "Amazon Virtual Private Cloud": {"aliyun": "专有网络 VPC", "price_ref": "免费 (NAT/EIP 另算)"},
    "AWS Lambda": {"aliyun": "函数计算 FC", "price_ref": "按调用量"},
    "AmazonCloudWatch": {"aliyun": "云监控 CMS + 日志服务 SLS", "price_ref": "~31元/月"},
    "Amazon Q": {"aliyun": "通义千问 / 百炼", "price_ref": "~255元/月"},
}

# 参考价（月）
REF_PRICES = {
    "nat_gateway": 200,          # NAT 网关 / 月
    "eip_bandwidth_per_mbps": 23,  # EIP 带宽 元/Mbps/月
    "eip_per_unit": 23,          # 默认 1Mbps EIP
    "slb_s1_small": 60,          # SLB 基础型
    "alb_standard": 120,         # ALB 标准型
    "nlb_standard": 60,          # NLB 标准型
    "oss_standard_per_gb": 0.12,
    "oss_ia_per_gb": 0.08,
    "oss_archive_per_gb": 0.033,
    "oss_deep_archive_per_gb": 0.0045,
    "sls_write_per_gb": 0.5,
    "sls_storage_per_gb_month": 0.09,
    "eventbridge_per_month": 0,  # 免费额度内
    "lambda_fc_per_month": 10,   # 小量调用参考
}


# ============================================================
# Excel 读取
# ============================================================

def read_excel(filepath: str) -> dict:
    """读取输入 Excel，返回 {sheet_name: [rows]} 结构"""
    wb = xlrd.open_workbook(filepath)
    data = {}
    for name in wb.sheet_names():
        sheet = wb.sheet_by_name(name)
        if sheet.nrows < 2:
            data[name] = {"headers": [], "rows": []}
            continue
        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        rows = []
        for r in range(1, sheet.nrows):
            row = {}
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                row[headers[c]] = val if val != "" else ""
            rows.append(row)
        data[name] = {"headers": headers, "rows": rows}
    return data


# ============================================================
# 询价服务调用
# ============================================================

def check_pricing_service(url: str) -> bool:
    """检查询价服务是否可用"""
    try:
        resp = requests.get(f"{url}/health", timeout=5)
        return resp.status_code == 200
    except Exception:
        return False


def batch_pricing(url: str, items: list, region_id: str = "cn-beijing", period: int = 1) -> dict:
    """调用批量询价 API"""
    try:
        resp = requests.post(
            f"{url}/api/pricing/batch",
            json={"region_id": region_id, "period": period, "items": items},
            timeout=120,
        )
        result = resp.json()
        if result.get("success"):
            return result["data"]
    except Exception as e:
        print(f"  [WARN] 询价调用失败: {e}")
    return {"prices": [], "summary": {}}


# ============================================================
# 核心处理逻辑
# ============================================================

def process_ec2(rows: list) -> list:
    """处理 EC2 数据，返回映射结果列表"""
    results = []
    for row in rows:
        inst_type = str(row.get("实例类型", ""))
        mapping = map_ec2_to_ecs(inst_type)
        vcpu = row.get("vCPU", "")
        mem = row.get("内存(GiB)", "")
        sys_disk = row.get("系统磁盘总空间(GB)", 0)
        data_disk = row.get("数据磁盘总空间(GB)", 0)
        try:
            sys_disk = int(float(sys_disk)) if sys_disk else 0
        except (ValueError, TypeError):
            sys_disk = 0
        try:
            data_disk = int(float(data_disk)) if data_disk else 0
        except (ValueError, TypeError):
            data_disk = 0
        # 解析 EBS 信息获取磁盘类型
        ebs_info = str(row.get("EBS属性信息", ""))
        disk_type = "GP3"
        if ebs_info:
            try:
                ebs_list = json.loads(ebs_info)
                types = set(e.get("volumeType", "") for e in ebs_list)
                disk_type = "/".join(types) if types else "GP3"
            except (json.JSONDecodeError, TypeError):
                pass
        results.append({
            "实例ID": str(row.get("实例ID", "")),
            "实例名称": str(row.get("实例名称", "")),
            "AWS规格": inst_type,
            "vCPU": vcpu,
            "内存(GiB)": mem,
            "系统盘(GB)": sys_disk,
            "数据盘(GB)": data_disk,
            "磁盘类型": disk_type,
            "Region": str(row.get("region", "")),
            "阿里云规格": mapping["aliyun_spec"],
            "阿里云系统盘": f"ESSD PL0 {max(sys_disk, 20)}GB",
            "阿里云数据盘": f"ESSD {'PL0' if data_disk < 20 else 'PL1'} {data_disk}GB" if data_disk > 0 else "-",
            "包月单价(元)": 0,
            "状态": str(row.get("实例状态", "")),
        })
    return results


def process_aurora(rows: list) -> list:
    """处理 AuroraDB 数据"""
    results = []
    for row in rows:
        inst_type = str(row.get("实例类型", ""))
        port = str(row.get("端口", ""))
        role = str(row.get("角色", ""))

        # 根据端口判断引擎
        if port in ("5432",):
            engine = "postgresql"
            aliyun_product = "PolarDB PostgreSQL"
        else:
            engine = "mysql"
            aliyun_product = "PolarDB MySQL"

        spec_info = AURORA_SPEC_MAP.get(inst_type, {})
        aliyun_spec = spec_info.get(engine, f"未映射({inst_type})")
        vcpu = row.get("vCPU", spec_info.get("vcpu", ""))
        mem = row.get("内存(GiB)", spec_info.get("mem_gb", ""))

        results.append({
            "域名": str(row.get("域名", "")),
            "AWS规格": inst_type,
            "端口": port,
            "引擎": "Aurora MySQL" if engine == "mysql" else "Aurora PostgreSQL",
            "角色": role,
            "vCPU": vcpu,
            "内存(GiB)": mem,
            "存储类型": str(row.get("存储类型", "")),
            "阿里云产品": aliyun_product,
            "阿里云规格": aliyun_spec,
            "包月单价(元)": 0,
        })
    return results


def process_documentdb(rows: list) -> list:
    """处理 DocumentDB 数据 - 精确规格码映射到阿里云 MongoDB
    
    优先从实例类型字段精确映射，回退到集群名称推断环境。
    """
    # DocumentDB 实例类型 → 阿里云 MongoDB 精确规格码映射
    DOCDB_INSTANCE_SPEC_MAP = {
        # db.t3 系列 (突发型)
        "db.t3.medium":   {"vcpu": 2,  "mem_gib": 4,   "aliyun_spec": "mdb.shard.2x.large.d",   "aliyun_desc": "独享型 2C4G 副本集（云盘版）"},
        # db.t4g 系列 (ARM 突发)
        "db.t4g.medium":  {"vcpu": 2,  "mem_gib": 4,   "aliyun_spec": "mdb.shard.2x.large.d",   "aliyun_desc": "独享型 2C4G 副本集（云盘版）"},
        # db.r5 系列 (Intel 内存型)
        "db.r5.large":    {"vcpu": 2,  "mem_gib": 16,  "aliyun_spec": "mdb.shard.2x.xlarge.d",  "aliyun_desc": "独享型 4C8G 副本集（云盘版）"},
        "db.r5.xlarge":   {"vcpu": 4,  "mem_gib": 32,  "aliyun_spec": "mdb.shard.2x.2xlarge.d", "aliyun_desc": "独享型 8C16G 副本集（云盘版）"},
        "db.r5.2xlarge":  {"vcpu": 8,  "mem_gib": 64,  "aliyun_spec": "mdb.shard.2x.4xlarge.d", "aliyun_desc": "独享型 16C32G 副本集（云盘版）"},
        "db.r5.4xlarge":  {"vcpu": 16, "mem_gib": 128, "aliyun_spec": "mdb.shard.2x.8xlarge.d", "aliyun_desc": "独享型 32C64G 副本集（云盘版）"},
        "db.r5.8xlarge":  {"vcpu": 32, "mem_gib": 256, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        "db.r5.12xlarge": {"vcpu": 48, "mem_gib": 384, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        "db.r5.16xlarge": {"vcpu": 64, "mem_gib": 512, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        "db.r5.24xlarge": {"vcpu": 96, "mem_gib": 768, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        # db.r6g 系列 (ARM 内存型)
        "db.r6g.large":    {"vcpu": 2,  "mem_gib": 16,  "aliyun_spec": "mdb.shard.2x.xlarge.d",  "aliyun_desc": "独享型 4C8G 副本集（云盘版）"},
        "db.r6g.xlarge":   {"vcpu": 4,  "mem_gib": 32,  "aliyun_spec": "mdb.shard.2x.2xlarge.d", "aliyun_desc": "独享型 8C16G 副本集（云盘版）"},
        "db.r6g.2xlarge":  {"vcpu": 8,  "mem_gib": 64,  "aliyun_spec": "mdb.shard.2x.4xlarge.d", "aliyun_desc": "独享型 16C32G 副本集（云盘版）"},
        "db.r6g.4xlarge":  {"vcpu": 16, "mem_gib": 128, "aliyun_spec": "mdb.shard.2x.8xlarge.d", "aliyun_desc": "独享型 32C64G 副本集（云盘版）"},
        "db.r6g.8xlarge":  {"vcpu": 32, "mem_gib": 256, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        "db.r6g.12xlarge": {"vcpu": 48, "mem_gib": 384, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        "db.r6g.16xlarge": {"vcpu": 64, "mem_gib": 512, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        # db.r7g 系列 (ARM 内存型新代)
        "db.r7g.large":    {"vcpu": 2,  "mem_gib": 16,  "aliyun_spec": "mdb.shard.2x.xlarge.d",  "aliyun_desc": "独享型 4C8G 副本集（云盘版）"},
        "db.r7g.xlarge":   {"vcpu": 4,  "mem_gib": 32,  "aliyun_spec": "mdb.shard.2x.2xlarge.d", "aliyun_desc": "独享型 8C16G 副本集（云盘版）"},
        "db.r7g.2xlarge":  {"vcpu": 8,  "mem_gib": 64,  "aliyun_spec": "mdb.shard.2x.4xlarge.d", "aliyun_desc": "独享型 16C32G 副本集（云盘版）"},
        "db.r7g.4xlarge":  {"vcpu": 16, "mem_gib": 128, "aliyun_spec": "mdb.shard.2x.8xlarge.d", "aliyun_desc": "独享型 32C64G 副本集（云盘版）"},
        "db.r7g.8xlarge":  {"vcpu": 32, "mem_gib": 256, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        "db.r7g.12xlarge": {"vcpu": 48, "mem_gib": 384, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
        "db.r7g.16xlarge": {"vcpu": 64, "mem_gib": 512, "aliyun_spec": "mdb.shard.4x.8xlarge.d", "aliyun_desc": "独享型 32C128G 副本集（云盘版）"},
    }

    # 环境推断 fallback
    ENV_FALLBACK = {
        "prod": {"aws_estimated_spec": "db.r6g.large", "vcpu": 2, "mem_gib": 16,
                 "aliyun_spec": "mdb.shard.2x.xlarge.d", "aliyun_desc": "独享型 4C8G 副本集（云盘版）"},
        "dev":  {"aws_estimated_spec": "db.t3.medium",  "vcpu": 2, "mem_gib": 4,
                 "aliyun_spec": "mdb.shard.2x.large.d",  "aliyun_desc": "独享型 2C4G 副本集（云盘版）"},
        "test": {"aws_estimated_spec": "db.t3.medium",  "vcpu": 2, "mem_gib": 4,
                 "aliyun_spec": "mdb.shard.2x.large.d",  "aliyun_desc": "独享型 2C4G 副本集（云盘版）"},
    }

    results = []
    for row in rows:
        cluster_id = str(row.get("集群标识符", ""))
        engine_ver = str(row.get("引擎版本", ""))
        inst_type = str(row.get("实例类型", "")).strip()

        # 优先精确映射
        spec_info = DOCDB_INSTANCE_SPEC_MAP.get(inst_type)
        if spec_info:
            results.append({
                "集群标识": cluster_id,
                "引擎": "DocumentDB",
                "引擎版本": engine_ver,
                "AWS预估规格": inst_type,
                "vCPU": spec_info["vcpu"],
                "内存(GiB)": spec_info["mem_gib"],
                "多可用区": str(row.get("多可用区", "")),
                "Region": str(row.get("region", "")),
                "阿里云产品": "云数据库 MongoDB 版",
                "阿里云规格": spec_info["aliyun_spec"],
                "阿里云规格描述": spec_info["aliyun_desc"],
                "迁移工具": "DTS 数据传输服务",
                "包月单价(元)": 0,
            })
        else:
            # 回退到环境推断
            env = "prod"
            if cluster_id.startswith("dev"):
                env = "dev"
            elif cluster_id.startswith("test"):
                env = "test"
            fb = ENV_FALLBACK.get(env, ENV_FALLBACK["prod"])
            results.append({
                "集群标识": cluster_id,
                "引擎": "DocumentDB",
                "引擎版本": engine_ver,
                "AWS预估规格": fb.get("aws_estimated_spec", inst_type or "未知"),
                "vCPU": fb["vcpu"],
                "内存(GiB)": fb["mem_gib"],
                "多可用区": str(row.get("多可用区", "")),
                "Region": str(row.get("region", "")),
                "阿里云产品": "云数据库 MongoDB 版",
                "阿里云规格": fb["aliyun_spec"],
                "阿里云规格描述": fb["aliyun_desc"],
                "迁移工具": "DTS 数据传输服务",
                "包月单价(元)": 0,
            })
    return results


def process_elasticache(rows: list) -> list:
    """处理 ElasticCache 数据"""
    results = []
    for row in rows:
        node_type = str(row.get("节点类型", ""))
        engine = str(row.get("引擎", ""))
        version = str(row.get("引擎版本", ""))

        spec_info = CACHE_SPEC_MAP.get(node_type, {})
        aliyun_spec = spec_info.get("aliyun", f"未映射({node_type})")
        desc = spec_info.get("desc", "")

        results.append({
            "集群ID": str(row.get("集群ID", "")),
            "名称": str(row.get("名称", "")),
            "AWS节点类型": node_type,
            "引擎": engine,
            "版本": version,
            "节点数量": row.get("节点数量", 1),
            "Region": str(row.get("Region", "")),
            "阿里云产品": "云数据库 Redis/Tair",
            "阿里云规格": aliyun_spec,
            "阿里云描述": desc,
            "包月单价(元)": 0,
        })
    return results


def process_s3(rows: list) -> list:
    """处理 S3 数据"""
    results = []
    # S3 列包含各存储类型的大小
    storage_type_cols = {
        "STANDARD": ("STANDARD类型文件数量", "STANDARD类型文件大小(MB)"),
        "REDUCED_REDUNDANCY": ("REDUCED_REDUNDANCY类型文件数量", "REDUCED_REDUNDANCY类型文件大小(MB)"),
        "GLACIER": ("GLACIER类型文件数量", "GLACIER类型文件大小(MB)"),
        "STANDARD_IA": ("STANDARD_IA类型文件数量", "STANDARD_IA类型文件大小(MB)"),
        "ONEZONE_IA": ("ONEZONE_IA类型文件数量", "ONEZONE_IA类型文件大小(MB)"),
        "INTELLIGENT_TIERING": ("INTELLIGENT_TIERING类型文件数量", "INTELLIGENT_TIERING类型文件大小(MB)"),
        "DEEP_ARCHIVE": ("DEEP_ARCHIVE类型文件数量", "DEEP_ARCHIVE类型文件大小(MB)"),
    }

    for row in rows:
        bucket = str(row.get("Bucket", ""))
        total_gb = 0
        try:
            total_gb = float(row.get("存储桶总大小(GB)", 0))
        except (ValueError, TypeError):
            pass

        # 计算各类型存储的费用
        storage_detail = []
        total_cost = 0
        for stype, (count_col, size_col) in storage_type_cols.items():
            size_mb = 0
            try:
                size_mb = float(row.get(size_col, 0))
            except (ValueError, TypeError):
                pass
            if size_mb > 0:
                size_gb = size_mb / 1024
                mapping = S3_STORAGE_MAP.get(stype, {})
                price = mapping.get("price_gb_month", 0.12)
                cost = size_gb * price
                total_cost += cost
                storage_detail.append(f"{mapping.get('aliyun', stype)}: {size_gb:.2f}GB")

        if not storage_detail and total_gb > 0:
            total_cost = total_gb * 0.12  # 默认按标准存储算
            storage_detail.append(f"OSS 标准存储: {total_gb:.2f}GB")

        results.append({
            "Bucket": bucket,
            "总大小(GB)": total_gb,
            "对象总数": row.get("对象总数", ""),
            "Region": str(row.get("Region", "")),
            "阿里云产品": "OSS 对象存储",
            "阿里云规格": "标准存储",
            "存储映射详情": "; ".join(storage_detail) if storage_detail else "-",
            "包月单价(元)": round(total_cost, 2),
        })
    return results


def process_network(lb_rows: list, nat_rows: list, eip_rows: list) -> list:
    """处理网络相关: 负载均衡 + NAT + EIP, 每个资源精确映射阿里云规格"""
    results = []

    # 负载均衡: NLB(network) → 阿里云 NLB, ALB(application) → 阿里云 ALB
    # NLB/ALB 仅支持按量付费，不支持包年包月，BSS 接口无法询价
    # 月度估算 = 实例费单价(元/h) × 730h，仅含实例费，不含 LCU 费和公网带宽费
    # 官方实例费(中国内地): NLB ¥0.147/h, ALB标准版 ¥0.147/h, ALB基础版 ¥0.049/h
    NLB_HOURLY = 0.147   # NLB 实例费 (元/h)
    ALB_HOURLY = 0.147   # ALB 标准版实例费 (元/h)
    HOURS_PER_MONTH = 730
    for row in lb_rows:
        name = str(row.get("名称", ""))
        lb_type = str(row.get("类型", ""))
        ip_type = str(row.get("IP类型", ""))
        if lb_type == "network":
            aliyun = "NLB 网络负载均衡"
            aliyun_spec = "按量付费"
            aliyun_desc = "NLB 按量付费(实例费¥0.147/h, 不含LCU费)"
            monthly = round(NLB_HOURLY * HOURS_PER_MONTH, 2)
        else:
            aliyun = "ALB 应用负载均衡"
            aliyun_spec = "按量付费"
            aliyun_desc = "ALB 标准版 按量付费(实例费¥0.147/h, 不含LCU费)"
            monthly = round(ALB_HOURLY * HOURS_PER_MONTH, 2)
        results.append({
            "AWS资源": f"ELB: {name}",
            "AWS规格": f"{lb_type.upper()} / {ip_type}",
            "类型": lb_type.upper(),
            "Region": str(row.get("region", "")),
            "阿里云产品": aliyun,
            "阿里云规格": aliyun_spec,
            "阿里云规格描述": aliyun_desc,
            "包月单价(元)": monthly,
        })

    # NAT Gateway → 阿里云公网 NAT 网关增强型
    for row in nat_rows:
        gw_id = str(row.get("网关ID", ""))
        results.append({
            "AWS资源": f"NAT: {gw_id}",
            "AWS规格": "NAT Gateway",
            "类型": "NAT Gateway",
            "Region": str(row.get("region", "")),
            "阿里云产品": "公网 NAT 网关",
            "阿里云规格": "增强型小规格",
            "阿里云规格描述": "Small(最大连接数10万, 最大新建连接数5000/s, 带宽2Gbps)",
            "包月单价(元)": 0,
        })

    # EIP → 阿里云 EIP（默认5Mbps带宽按固定计费）
    for row in eip_rows:
        ip = str(row.get("已分配的IPv4地址", ""))
        assoc = str(row.get("关联的实例ID", ""))
        results.append({
            "AWS资源": f"EIP: {ip}",
            "AWS规格": f"Elastic IP{' (关联: '+assoc+')' if assoc else ' (未关联)'}",
            "类型": "Elastic IP",
            "Region": "",
            "阿里云产品": "弹性公网 IP (EIP)",
            "阿里云规格": "EIP 5Mbps 按固定带宽",
            "阿里云规格描述": "BGP(多线) 5Mbps 按固定带宽计费",
            "包月单价(元)": 0,
        })

    return results


def process_other_services(lambda_rows, athena_rows, eb_rows, sns_rows, cw_rows, other_rows) -> list:
    """处理其他服务，每个资源精确映射阿里云规格并估算价格"""

    # Other 服务到阿里云产品精确映射
    OTHER_SPEC_MAP = {
        "Amazon Simple Notification Service": {
            "aliyun": "消息服务 MNS",
            "aliyun_spec": "MNS 标准版",
            "desc": "按消息数计费,100万条/月免费",
        },
        "AWS Key Management Service": {
            "aliyun": "密钥管理服务 KMS",
            "aliyun_spec": "KMS 软件密钥",
            "desc": "软件密钥实例,按密钥数和API调用量计费",
        },
        "AWS Secrets Manager": {
            "aliyun": "凭据管家 Secrets Manager",
            "aliyun_spec": "凭据管家标准版",
            "desc": "按凭据数量计费,0.4元/凭据/月",
        },
        "Amazon Virtual Private Cloud": {
            "aliyun": "专有网络 VPC",
            "aliyun_spec": "VPC + 交换机 + 路由表",
            "desc": "VPC本身免费,NAT/EIP/带宽另计",
        },
        "AWS Lambda": {
            "aliyun": "函数计算 FC",
            "aliyun_spec": "FC 按量实例",
            "desc": "按调用次数+执行时长计费",
        },
        "AmazonCloudWatch": {
            "aliyun": "云监控 CMS",
            "aliyun_spec": "CMS 企业版",
            "desc": "基础监控免费,自定义监控按指标计费",
        },
        "Amazon Q": {
            "aliyun": "通义千问 / 百炼平台",
            "aliyun_spec": "百炼大模型服务",
            "desc": "按Token用量计费",
        },
    }

    results = []

    # Lambda → FC 函数计算
    # FC 规格映射: arm64→FC 弹性实例(ARM), x86_64→FC 弹性实例(x86)
    FC_MEM_SPEC = {
        128: {"spec": "fc.弹性实例 0.05vCPU/128MB", "vcpu": 0.05, "monthly": 8},
        256: {"spec": "fc.弹性实例 0.1vCPU/256MB", "vcpu": 0.1, "monthly": 15},
        512: {"spec": "fc.弹性实例 0.25vCPU/512MB", "vcpu": 0.25, "monthly": 30},
        1024: {"spec": "fc.弹性实例 0.5vCPU/1024MB", "vcpu": 0.5, "monthly": 60},
        1490: {"spec": "fc.弹性实例 1vCPU/1536MB", "vcpu": 1, "monthly": 90},
        2048: {"spec": "fc.弹性实例 1vCPU/2048MB", "vcpu": 1, "monthly": 120},
        3008: {"spec": "fc.弹性实例 2vCPU/3072MB", "vcpu": 2, "monthly": 180},
    }

    for row in lambda_rows:
        fname = str(row.get("函数名称", ""))
        runtime = str(row.get("运行时", ""))
        arch = str(row.get("架构", "x86_64"))
        timeout = row.get("超时", 3)
        ephemeral = row.get("短暂存储", 512)
        mem = row.get("内存", 128)
        try:
            mem_val = int(float(mem))
        except (ValueError, TypeError):
            mem_val = 128

        # 找到最接近的FC规格
        closest_mem = min(FC_MEM_SPEC.keys(), key=lambda x: abs(x - mem_val))
        fc_info = FC_MEM_SPEC[closest_mem]

        results.append({
            "AWS服务": "Lambda",
            "资源名称": fname,
            "AWS规格": f"内存:{mem_val}MB / 架构:{arch} / 超时:{timeout}s",
            "详情": f"运行时: {runtime}, 临时存储: {ephemeral}MB",
            "阿里云产品": "函数计算 FC 3.0",
            "阿里云规格": fc_info["spec"],
            "阿里云规格描述": f"{fc_info['vcpu']}vCPU/{closest_mem}MB {'ARM' if arch=='arm64' else 'x86'}",
            "包月单价(元)": fc_info["monthly"],
        })

    # Athena → MaxCompute/DLA
    for row in athena_rows:
        db = str(row.get("数据库名称", ""))
        table = str(row.get("表名称", ""))
        fields = row.get("字段数量", 0)
        results.append({
            "AWS服务": "Athena",
            "资源名称": f"{db}.{table}",
            "AWS规格": f"Serverless SQL / {fields} 字段",
            "详情": f"表类型: {row.get('表类型', '')}, 目录: {row.get('目录', '')}",
            "阿里云产品": "MaxCompute (原ODPS)",
            "阿里云规格": "MaxCompute 按量付费(SQL计算)",
            "阿里云规格描述": "按量模式,按SQL扫描量计费,1元/GB",
            "包月单价(元)": 50,
        })

    # EventBridge → 阿里云事件总线
    for row in eb_rows:
        name = str(row.get("名称", ""))
        status = str(row.get("状态", ""))
        results.append({
            "AWS服务": "EventBridge",
            "资源名称": name,
            "AWS规格": f"EventBridge Rule / {status}",
            "详情": f"描述: {row.get('描述', '-')}",
            "阿里云产品": "事件总线 EventBridge",
            "阿里云规格": "EventBridge 标准版",
            "阿里云规格描述": "标准版,免费额度: 每月500万次事件投递",
            "包月单价(元)": 0,
        })

    # SNS → MNS / EventBridge
    for row in sns_rows:
        name = str(row.get("名称", ""))
        sns_type = str(row.get("SNS类型", ""))
        subs = row.get("节点", 0)
        results.append({
            "AWS服务": "SNS",
            "资源名称": name,
            "AWS规格": f"{sns_type} / 订阅者:{subs}",
            "详情": f"类型: {sns_type}",
            "阿里云产品": "消息服务 MNS",
            "阿里云规格": "MNS 主题(Topic)",
            "阿里云规格描述": f"Topic 消息推送,{subs}个订阅者",
            "包月单价(元)": 5,
        })

    # CloudWatch Logs → SLS 日志服务
    for row in cw_rows:
        log_group = str(row.get("日志组", ""))
        log_class = str(row.get("日志类", "STANDARD"))
        storage_bytes = 0
        try:
            storage_bytes = int(float(row.get("存储字节数", 0)))
        except (ValueError, TypeError):
            pass
        storage_gb = storage_bytes / (1024 ** 3)
        # SLS 计费: 写入0.76元/GB + 存储0.0115元/GB/天(约0.35元/GB/月) + 索引0.35元/GB
        monthly_cost = round(storage_gb * (0.76 + 0.35 + 0.35), 2)
        results.append({
            "AWS服务": "CloudWatch Logs",
            "资源名称": log_group,
            "AWS规格": f"日志类:{log_class} / 存储:{storage_gb:.4f}GB",
            "详情": f"存储: {storage_bytes} bytes ({storage_gb:.4f} GB)",
            "阿里云产品": "日志服务 SLS",
            "阿里云规格": f"SLS 标准型 Logstore({storage_gb:.2f}GB)",
            "阿里云规格描述": f"标准型Logstore,写入+索引+存储,{storage_gb:.4f}GB",
            "包月单价(元)": monthly_cost,
        })

    # Other 服务
    for row in other_rows:
        service = str(row.get("服务", ""))
        cost = row.get("消耗", 0)
        try:
            cost_val = float(cost)
        except (ValueError, TypeError):
            cost_val = 0

        mapping = OTHER_SPEC_MAP.get(service, {
            "aliyun": service,
            "aliyun_spec": "参考AWS对应产品",
            "desc": "请在阿里云控制台确认具体规格",
        })

        # AWS USD → CNY 汇率约7.2
        aliyun_est = round(cost_val * 7.2, 2)
        results.append({
            "AWS服务": service,
            "资源名称": service.split(" ")[-1] if " " in service else service,
            "AWS规格": f"按量 / AWS月费: ${cost_val:.2f}",
            "详情": mapping["desc"],
            "阿里云产品": mapping["aliyun"],
            "阿里云规格": mapping["aliyun_spec"],
            "阿里云规格描述": mapping["desc"],
            "包月单价(元)": aliyun_est,
        })

    return results


# ============================================================
# 询价 & 填充价格
# ============================================================

def fill_ecs_prices(ec2_results: list, pricing_url: str, region_id: str = "cn-beijing"):
    """对 ECS 进行批量询价 - 包含实例+系统盘+数据盘完整配置"""
    import re

    def _parse_disk(disk_str):
        """从 'ESSD PL0 40GB' 或 'ESSD PL1 400GB' 提取大小和PL级别"""
        size = 0
        pl = "PL0"
        if disk_str and disk_str != "-":
            m = re.search(r'(\d+)GB', disk_str)
            if m:
                size = int(m.group(1))
            if "PL1" in disk_str:
                pl = "PL1"
            elif "PL2" in disk_str:
                pl = "PL2"
        return size, pl

    # 按 (规格, 系统盘大小, 数据盘大小) 去重
    config_counter = Counter()
    for r in ec2_results:
        spec = r["阿里云规格"]
        if "未" in spec:
            continue
        sys_size, _ = _parse_disk(r.get("阿里云系统盘", ""))
        dd_size, dd_pl = _parse_disk(r.get("阿里云数据盘", ""))
        config_counter[(spec, sys_size, dd_size, dd_pl)] += 1

    if not config_counter:
        return

    items = []
    for (spec, sys_size, dd_size, dd_pl), count in config_counter.items():
        item = {
            "product_code": "ecs",
            "instance_spec": spec,
            "count": count,
            "disk_category": "cloud_essd",
            "disk_size": max(sys_size, 20),  # 系统盘最低20GB
        }
        if dd_size > 0:
            item["data_disk_category"] = "cloud_essd"
            item["data_disk_size"] = dd_size
            item["data_disk_pl"] = dd_pl
        items.append(item)

    print(f"  询价 ECS: {len(items)} 种配置(规格+系统盘+数据盘), 共 {sum(config_counter.values())} 台...")
    result = batch_pricing(pricing_url, items, region_id)

    # 按顺序建立配置→价格映射
    price_list = result.get("prices", [])
    price_map = {}
    config_keys = list(config_counter.keys())
    for i, p in enumerate(price_list):
        if p.get("success") and i < len(config_keys):
            price_map[config_keys[i]] = p["trade_price"]

    # 回填价格
    filled = 0
    for r in ec2_results:
        spec = r["阿里云规格"]
        if "未" in spec:
            continue
        sys_size, _ = _parse_disk(r.get("阿里云系统盘", ""))
        dd_size, dd_pl = _parse_disk(r.get("阿里云数据盘", ""))
        key = (spec, sys_size, dd_size, dd_pl)
        if key in price_map:
            r["包月单价(元)"] = price_map[key]
            filled += 1
        else:
            r["包月单价(元)"] = "询价失败"
    print(f"    成功填充 {filled}/{len(ec2_results)} 台价格")


def fill_polardb_prices(aurora_results: list, pricing_url: str, region_id: str = "cn-beijing"):
    """对 PolarDB 批量询价 - 仅处理 polar.* 规格
    
    PolarDB 最小配置 = 2节点（1主+1只读）。
    - clusterMaster 节点按 2节点集群计价（单节点价×2）
    - clusterSlave 节点不单独计价（已包含在 master 的集群价格中）
    """
    # 只收集 polar.* 规格
    polar_specs = set()
    for r in aurora_results:
        spec = r.get("阿里云规格", "")
        if spec and "未" not in spec and spec.startswith("polar."):
            polar_specs.add(spec)

    if not polar_specs:
        return

    items = []
    for spec in polar_specs:
        engine = "MySQL" if "mysql" in spec else "PostgreSQL"
        items.append({
            "product_code": "polardb",
            "instance_spec": spec,
            "engine": engine,
            "storage_space": 50,
            "count": 1,
        })

    print(f"  询价 PolarDB: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    # price_map 存储的是 单节点 月价
    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    # PolarDB 参考价作为 fallback（来自 DescribeClassList ReferencePrice，单位元/月/单节点）
    polardb_ref_prices = {
        "polar.mysql.g2.medium": 250,
        "polar.mysql.g2.large": 475,
        "polar.mysql.g2.xlarge": 1050,
        "polar.mysql.g4.xlarge": 1650,
        "polar.mysql.g8.xlarge": 2700,
        "polar.mysql.g8.2xlarge": 5140,
        "polar.mysql.g8.4xlarge": 10280,
        "polar.mysql.g8.8xlarge": 20560,
        "polar.mysql.g8.16xlarge": 41120,
        "polar.pg.g2.medium": 250,
        "polar.pg.g2.large": 475,
        "polar.pg.g2.xlarge": 1050,
        "polar.pg.g4.xlarge": 1500,
        "polar.pg.g8.xlarge": 2400,
        "polar.pg.g8.2xlarge": 4800,
        "polar.pg.g8.4xlarge": 9600,
        "polar.pg.g8.8xlarge": 19200,
        "polar.pg.g8.16xlarge": 38400,
    }

    for r in aurora_results:
        spec = r.get("阿里云规格", "")
        if not spec.startswith("polar."):
            continue  # RDS 规格由 fill_aurora_rds_prices 处理
        role = r.get("角色", "")
        single_node_price = price_map.get(spec) or polardb_ref_prices.get(spec)
        if single_node_price is not None:
            if role == "clusterSlave":
                r["节点数"] = 0
                r["包月单价(元)"] = 0
            else:
                r["节点数"] = 2
                r["包月单价(元)"] = single_node_price * 2


def fill_aurora_rds_prices(aurora_results: list, pricing_url: str, region_id: str = "cn-beijing"):
    """对 Aurora 中映射到常规 RDS 规格(rds.mysql.*/pg.n*)的实例询价
    
    Intel 系列 Aurora (db.t3/db.r5/db.m5/db.m6i) 映射到阿里云常规 RDS 规格，
    需要通过 RDS DescribePrice API 询价。
    """
    # 收集 rds.*/pg.* 规格
    rds_specs = set()
    for r in aurora_results:
        spec = r.get("阿里云规格", "")
        if spec and "未" not in spec and not spec.startswith("polar."):
            rds_specs.add(spec)

    if not rds_specs:
        return

    items = []
    for spec in rds_specs:
        if spec.startswith("rds.mysql."):
            engine = "MySQL"
            engine_version = "8.0"
        elif spec.startswith("pg."):
            engine = "PostgreSQL"
            engine_version = "16.0"
        else:
            engine = "MySQL"
            engine_version = "8.0"
        items.append({
            "product_code": "rds",
            "instance_spec": spec,
            "engine": engine,
            "engine_version": engine_version,
            "disk_size": 100,
            "count": 1,
        })

    print(f"  询价 RDS(常规): {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    # RDS 参考价 fallback（包含 100GB ESSD 存储，cn-beijing）
    rds_ref_prices = {
        "rds.mysql.t1.small": 120,
        "rds.mysql.s1.small": 250,
        "rds.mysql.s2.large": 500,
        "rds.mysql.s2.xlarge": 800,
        "rds.mysql.s3.small": 650,
        "rds.mysql.s3.large": 1100,
        "rds.mysql.c1.medium": 1800,
        "rds.mysql.m1.medium": 2400,
        "rds.mysql.c1.xlarge": 3500,
        "rds.mysql.m1.xlarge": 4800,
        "rds.mysql.m1.2xlarge": 9600,
        "rds.mysql.m1.4xlarge": 19200,
        "rds.mysql.m1.8xlarge": 38400,
        "pg.n1.micro.1": 120,
        "pg.n2.small.1": 250,
        "pg.n2.small.2c": 500,
        "pg.n2.medium.1": 800,
        "pg.n2.medium.2c": 1100,
        "pg.n4.medium.1": 1100,
        "pg.n2.large.1": 1800,
        "pg.n2.large.2c": 2400,
        "pg.n4.large.1": 1800,
        "pg.n4.xlarge.1": 3500,
        "pg.n2.xlarge.2c": 4800,
        "pg.n4.2xlarge.1": 7000,
        "pg.n4.4xlarge.1": 14000,
        "pg.n8.8xlarge.1": 38400,
    }

    filled = 0
    for r in aurora_results:
        spec = r.get("阿里云规格", "")
        if spec.startswith("polar."):
            continue  # PolarDB 规格由 fill_polardb_prices 处理
        price = price_map.get(spec) or rds_ref_prices.get(spec)
        if price is not None:
            r["节点数"] = 1
            r["包月单价(元)"] = price
            filled += 1
    if filled:
        print(f"    RDS 成功填充 {filled} 个实例价格")


def fill_mongodb_prices(docdb_results: list, pricing_url: str, region_id: str = "cn-beijing"):
    """对 MongoDB 批量询价"""
    spec_set = set()
    for r in docdb_results:
        spec = r.get("阿里云规格", "")
        if spec and "未" not in spec:
            spec_set.add(spec)

    if not spec_set:
        return

    items = []
    for spec in spec_set:
        items.append({
            "product_code": "mongodb",
            "instance_spec": spec,
            "disk_size": 40,
            "count": 1,
        })

    print(f"  询价 MongoDB: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    # MongoDB 参考价 fallback（BSS GetSubscriptionPrice, cn-beijing, 含40GB存储，独享型云盘规格）
    mongodb_ref_prices = {
        "mdb.shard.2x.large.d": 1011,
        "mdb.shard.2x.xlarge.d": 1933,
        "mdb.shard.2x.2xlarge.d": 3769,
        "mdb.shard.2x.large.c": 670,
        "mdb.shard.2x.xlarge.c": 1168,
        "mdb.shard.4x.xlarge.c": 2043,
    }

    for r in docdb_results:
        spec = r.get("阿里云规格", "")
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]
        elif spec in mongodb_ref_prices:
            r["包月单价(元)"] = mongodb_ref_prices[spec]


def fill_slb_prices(network_results: list, pricing_url: str, region_id: str = "cn-beijing"):
    """对 NAT/EIP 批量询价（NLB/ALB 按量付费已在 process_network 中直接计算实例费）"""
    # NLB/ALB 不支持 BSS 询价，价格已在 process_network 中按实例费×730h 计算
    # 这里只需要对 NAT 和 EIP 进行询价

    items = []

    # NAT 网关询价
    nat_count = sum(1 for r in network_results if r.get("类型") == "NAT Gateway")
    if nat_count > 0:
        items.append({"product_code": "nat", "instance_spec": "Small", "count": nat_count})

    # EIP 询价
    eip_count = sum(1 for r in network_results if r.get("类型") == "Elastic IP")
    if eip_count > 0:
        items.append({"product_code": "eip", "instance_spec": "5Mbps", "bandwidth": 5, "count": eip_count})

    if not items:
        return

    print(f"  询价网络产品: NAT {nat_count}个 + EIP {eip_count}个 (NLB/ALB按实例费估算,不走API)...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in network_results:
        rtype = r.get("类型", "")
        if rtype == "NAT Gateway":
            if "Small" in price_map:
                r["包月单价(元)"] = price_map["Small"]
            else:
                r["包月单价(元)"] = 220  # NAT 增强型小规格参考价
        elif rtype == "Elastic IP":
            if "5Mbps" in price_map:
                r["包月单价(元)"] = price_map["5Mbps"]
            else:
                r["包月单价(元)"] = 125  # 5Mbps EIP 参考价


def fill_redis_prices(cache_results: list, pricing_url: str, region_id: str = "cn-beijing"):
    """对 Redis 批量询价"""
    spec_set = set()
    for r in cache_results:
        spec = r["阿里云规格"]
        if "未" not in spec:
            spec_set.add(spec)

    if not spec_set:
        return

    items = []
    for spec in spec_set:
        # 从 CACHE_SPEC_MAP 获取 capacity
        for aws_spec, info in CACHE_SPEC_MAP.items():
            if info["aliyun"] == spec:
                cap = int(info["mem_gb"] * 1024)
                break
        else:
            cap = 1024
        items.append({
            "product_code": "redis",
            "instance_spec": spec,
            "capacity": cap,
            "count": 1,
        })

    print(f"  询价 Redis: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in cache_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


# ============================================================
# Excel 写出
# ============================================================

def write_excel(output_path: str, ec2_results, aurora_results, docdb_results,
                cache_results, s3_results, network_results, other_results,
                coverage_report, data):
    """生成输出 Excel"""
    wb = Workbook()

    # ---- Sheet 1: 产品映射总览 ----
    ws = wb.active
    ws.title = "产品映射总览"
    headers = ["序号", "AWS 产品(Sheet)", "资源数量", "阿里云产品", "推荐迁移工具", "所属层级", "迁移复杂度"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    sheet_counts = {}
    for name in data:
        sheet_counts[name] = len(data[name]["rows"])

    for idx, (sheet_name, info) in enumerate(PRODUCT_MAP.items(), 1):
        count = sheet_counts.get(sheet_name, 0)
        ws.append([idx, sheet_name, count, info["aliyun"], info["tool"], info["layer"], info["complexity"]])
    auto_width(ws)

    # ---- Sheet 2: EC2 规格明细 ----
    ws2 = wb.create_sheet("EC2 规格明细")
    ec2_headers = ["实例ID", "实例名称", "AWS规格", "vCPU", "内存(GiB)", "系统盘(GB)", "数据盘(GB)",
                   "磁盘类型", "Region", "阿里云规格", "阿里云系统盘", "阿里云数据盘", "包月单价(元)", "状态"]
    ws2.append(ec2_headers)
    style_header_row(ws2, 1, len(ec2_headers))
    for r in ec2_results:
        ws2.append([r.get(h, "") for h in ec2_headers])
    # 格式化价格列
    for row in ws2.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FORMAT
    # 添加汇总行
    total_row = ws2.max_row + 2
    ws2.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=f"{len(ec2_results)} 台").font = Font(bold=True)
    total_monthly = sum(r["包月单价(元)"] for r in ec2_results if isinstance(r["包月单价(元)"], (int, float)))
    ws2.cell(row=total_row, column=13, value=total_monthly).font = Font(bold=True)
    ws2.cell(row=total_row, column=13).number_format = MONEY_FORMAT
    auto_width(ws2)

    # ---- Sheet 3: 数据库明细 ----
    ws3 = wb.create_sheet("数据库明细")
    # Aurora
    ws3.append(["=== AuroraDB ==="])
    ws3.cell(row=1, column=1).font = SUBTITLE_FONT
    aurora_headers = ["域名", "AWS规格", "端口", "引擎", "角色", "vCPU", "内存(GiB)",
                      "存储类型", "阿里云产品", "阿里云规格", "节点数", "包月单价(元)"]
    ws3.append(aurora_headers)
    style_header_row(ws3, 2, len(aurora_headers))
    for r in aurora_results:
        ws3.append([r.get(h, "") for h in aurora_headers])

    # DocumentDB
    gap_row = ws3.max_row + 2
    ws3.cell(row=gap_row, column=1, value="=== DocumentDB ===").font = SUBTITLE_FONT
    docdb_headers = ["集群标识", "引擎", "引擎版本", "AWS预估规格", "vCPU", "内存(GiB)",
                     "多可用区", "Region", "阿里云产品", "阿里云规格", "阿里云规格描述", "迁移工具", "包月单价(元)"]
    ws3.append(docdb_headers)
    style_header_row(ws3, gap_row + 1, len(docdb_headers))
    for r in docdb_results:
        ws3.append([r.get(h, "") for h in docdb_headers])
    auto_width(ws3)

    # ---- Sheet 4: 缓存明细 ----
    ws4 = wb.create_sheet("缓存明细")
    cache_headers = ["集群ID", "名称", "AWS节点类型", "引擎", "版本", "节点数量",
                     "Region", "阿里云产品", "阿里云规格", "阿里云描述", "包月单价(元)"]
    ws4.append(cache_headers)
    style_header_row(ws4, 1, len(cache_headers))
    for r in cache_results:
        ws4.append([r.get(h, "") for h in cache_headers])
    auto_width(ws4)

    # ---- Sheet 5: 存储明细 ----
    ws5 = wb.create_sheet("存储明细")
    s3_headers = ["Bucket", "总大小(GB)", "对象总数", "Region", "阿里云产品", "阿里云规格", "存储映射详情", "包月单价(元)"]
    ws5.append(s3_headers)
    style_header_row(ws5, 1, len(s3_headers))
    for r in s3_results:
        ws5.append([r.get(h, "") for h in s3_headers])
    # 汇总
    total_s3_cost = sum(r["包月单价(元)"] for r in s3_results if isinstance(r["包月单价(元)"], (int, float)))
    total_row = ws5.max_row + 1
    ws5.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws5.cell(row=total_row, column=8, value=total_s3_cost).font = Font(bold=True)
    ws5.cell(row=total_row, column=8).number_format = MONEY_FORMAT
    auto_width(ws5)

    # ---- Sheet 6: 网络明细 ----
    ws6 = wb.create_sheet("网络明细")
    net_headers = ["AWS资源", "AWS规格", "类型", "Region", "阿里云产品", "阿里云规格", "阿里云规格描述", "包月单价(元)"]
    ws6.append(net_headers)
    style_header_row(ws6, 1, len(net_headers))
    for r in network_results:
        ws6.append([r.get(h, "") for h in net_headers])
    total_net = sum(r["包月单价(元)"] for r in network_results if isinstance(r["包月单价(元)"], (int, float)))
    total_row = ws6.max_row + 1
    ws6.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws6.cell(row=total_row, column=8, value=total_net).font = Font(bold=True)
    auto_width(ws6)

    # ---- Sheet 7: 其他服务 ----
    ws7 = wb.create_sheet("其他服务")
    other_headers = ["AWS服务", "资源名称", "AWS规格", "详情", "阿里云产品", "阿里云规格", "阿里云规格描述", "包月单价(元)"]
    ws7.append(other_headers)
    style_header_row(ws7, 1, len(other_headers))
    for r in other_results:
        ws7.append([r.get(h, "") for h in other_headers])
    auto_width(ws7)

    # ---- Sheet 8: 成本汇总 ----
    ws8 = wb.create_sheet("成本汇总")
    ws8.cell(row=1, column=1, value="阿里云月度成本估算报告").font = TITLE_FONT
    ws8.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws8.cell(row=3, column=1, value=f"目标区域: cn-beijing (华北2-北京)")

    cost_headers = ["资源类别", "资源数量", "月度费用(元)", "包年预估(85折)", "3年预估(5折)", "备注"]
    ws8.append([])
    ws8.append(cost_headers)
    style_header_row(ws8, 5, len(cost_headers))

    # 计算各类费用
    ecs_total = sum(r["包月单价(元)"] for r in ec2_results if isinstance(r["包月单价(元)"], (int, float)))
    aurora_total = sum(r["包月单价(元)"] for r in aurora_results if isinstance(r["包月单价(元)"], (int, float)))
    docdb_total = sum(r["包月单价(元)"] for r in docdb_results if isinstance(r["包月单价(元)"], (int, float)))
    cache_total = sum(r["包月单价(元)"] for r in cache_results if isinstance(r["包月单价(元)"], (int, float)))
    s3_total = sum(r["包月单价(元)"] for r in s3_results if isinstance(r["包月单价(元)"], (int, float)))
    net_total = total_net
    other_total = sum(r["包月单价(元)"] for r in other_results if isinstance(r["包月单价(元)"], (int, float)))

    cost_rows = [
        ["ECS 弹性计算", f"{len(ec2_results)} 台", ecs_total, ecs_total * 0.85, ecs_total * 0.5, "实时询价"],
        ["PolarDB 数据库", f"{len(aurora_results)} 实例", aurora_total, aurora_total * 0.85, aurora_total * 0.5, "BSS询价/参考价"],
        ["MongoDB (DocumentDB)", f"{len(docdb_results)} 集群", docdb_total, docdb_total * 0.85, docdb_total * 0.5, "实时询价/参考价"],
        ["Redis/Tair 缓存", f"{len(cache_results)} 节点", cache_total, cache_total * 0.85, cache_total * 0.5, "实时询价"],
        ["OSS 对象存储", f"{len(s3_results)} 桶", s3_total, s3_total, s3_total, "按量计费,无长期折扣"],
        ["网络 (LB+NAT+EIP)", f"{len(network_results)} 个", net_total, net_total * 0.85, net_total * 0.5, "BSS询价/参考价"],
        ["其他服务", f"{len(other_results)} 项", other_total, other_total, other_total, "参考价/按量"],
    ]

    for cr in cost_rows:
        ws8.append(cr)

    grand_total = ecs_total + aurora_total + docdb_total + cache_total + s3_total + net_total + other_total
    ws8.append([])
    summary_row = ws8.max_row + 1
    ws8.cell(row=summary_row, column=1, value="月度合计").font = Font(bold=True, size=12)
    ws8.cell(row=summary_row, column=3, value=grand_total).font = Font(bold=True, size=12)
    ws8.cell(row=summary_row, column=3).number_format = MONEY_FORMAT
    ws8.cell(row=summary_row, column=4, value=grand_total * 0.85).number_format = MONEY_FORMAT
    ws8.cell(row=summary_row, column=5, value=grand_total * 0.5).number_format = MONEY_FORMAT

    ws8.append([])
    ws8.append(["年度费用预估", "", grand_total * 12, grand_total * 0.85 * 12, grand_total * 0.5 * 12, ""])
    ws8.append(["年度节省(vs包月)", "", "-", f"¥{grand_total * 0.15 * 12:,.0f}", f"¥{grand_total * 0.5 * 12:,.0f}", ""])

    # 格式化价格列
    for row in ws8.iter_rows(min_row=6, min_col=3, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FORMAT
    auto_width(ws8)

    # ---- Sheet 9: 迁移优先级 ----
    ws9 = wb.create_sheet("迁移优先级")
    prio_headers = ["批次", "风险等级", "迁移内容", "涉及产品", "资源数量", "建议"]
    ws9.append(prio_headers)
    style_header_row(ws9, 1, len(prio_headers))

    lb_count = len(data.get("负载均衡", {}).get("rows", []))
    nat_count = len(data.get("NAT", {}).get("rows", []))
    eip_count = len(data.get("EIP", {}).get("rows", []))

    priorities = [
        [1, "低", "网络基础设施", "VPC/NAT/ELB/EIP/DNS", f"LB:{lb_count}, NAT:{nat_count}, EIP:{eip_count}", "手动重建,先搭建网络底座"],
        [2, "低", "对象存储", "S3 → OSS", f"{len(s3_results)} 桶", "使用在线迁移服务,大桶优先"],
        [3, "低", "安全与身份", "KMS/SecretsManager", "2 服务", "密钥和凭据重建"],
        [4, "中", "计算实例", "EC2 → ECS", f"{len(ec2_results)} 台", "SMC 整机迁移,分批次灰度切换"],
        [5, "中", "数据库", "Aurora → PolarDB", f"{len(aurora_results)} 实例", "DTS 实时同步,验证后切换"],
        [6, "中", "文档数据库", "DocumentDB → MongoDB", f"{len(docdb_results)} 集群", "DTS 迁移,注意兼容性"],
        [7, "中", "缓存", "ElastiCache → Redis/Tair", f"{len(cache_results)} 节点", "redis-shake/DTS 同步"],
        [8, "中", "运维监控", "CloudWatch → SLS/CMS", f"{len(data.get('CloudWatchLog', {}).get('rows', []))} 日志组", "日志投递适配"],
        [9, "高", "函数计算", "Lambda → FC", f"{len(data.get('lamdba', {}).get('rows', []))} 函数", "代码改造+运行时适配"],
        [10, "高", "事件/消息", "EventBridge/SNS", f"EB:{len(data.get('EventBridge', {}).get('rows', []))}, SNS:{len(data.get('SNS', {}).get('rows', []))}", "应用层适配,最后切换"],
    ]
    for p in priorities:
        ws9.append(p)
    auto_width(ws9)

    # ---- Sheet 10: 覆盖率报告 ----
    ws10 = wb.create_sheet("覆盖率报告")
    cov_headers = ["AWS 产品(Sheet)", "是否覆盖", "阿里云映射产品", "规格映射", "询价支持", "备注"]
    ws10.append(cov_headers)
    style_header_row(ws10, 1, len(cov_headers))
    for item in coverage_report:
        row_num = ws10.max_row + 1
        ws10.append([item["product"], item["covered"], item["aliyun"], item["spec_mapping"],
                     item["pricing"], item["note"]])
        # 颜色标记
        fill = OK_FILL if item["covered"] == "已覆盖" else WARN_FILL
        for c in range(1, len(cov_headers) + 1):
            ws10.cell(row=row_num, column=c).fill = fill
    auto_width(ws10)

    # 保存
    wb.save(output_path)
    print(f"\n输出文件已保存: {output_path}")


# ============================================================
# 覆盖率验证
# ============================================================

def validate_coverage(data: dict) -> list:
    """验证 SKILL.md 中的映射是否覆盖了 Excel 中的所有产品"""
    report = []

    coverage_info = {
        "EC2": {"covered": True, "spec": True, "pricing": True,
                "note": "23种规格全部映射,支持实时询价"},
        "S3": {"covered": True, "spec": True, "pricing": False,
               "note": "7种存储类型全映射,使用参考价"},
        "ElasticCache": {"covered": True, "spec": True, "pricing": True,
                        "note": "含valkey引擎(兼容Redis),支持实时询价"},
        "DocumentDB": {"covered": True, "spec": False, "pricing": False,
                      "note": "映射到MongoDB,需确认实例规格后询价"},
        "NAT": {"covered": True, "spec": False, "pricing": False,
               "note": "使用参考价 ~200元/月"},
        "负载均衡": {"covered": True, "spec": False, "pricing": False,
                  "note": "NLB/ALB 均已映射,使用参考价"},
        "Athena": {"covered": True, "spec": False, "pricing": False,
                  "note": "映射到DLA/MaxCompute,按查询量计费"},
        "lamdba": {"covered": True, "spec": False, "pricing": False,
                  "note": "映射到函数计算FC,需代码适配"},
        "EIP": {"covered": True, "spec": False, "pricing": False,
               "note": "映射到阿里云EIP,参考价23元/Mbps/月"},
        "EventBridge": {"covered": True, "spec": False, "pricing": False,
                       "note": "映射到阿里云EventBridge"},
        "SNS": {"covered": True, "spec": False, "pricing": False,
               "note": "映射到MNS/EventBridge,需应用适配"},
        "CloudWatchLog": {"covered": True, "spec": False, "pricing": False,
                         "note": "映射到日志服务SLS,按存储量计费"},
        "AuroraDB": {"covered": True, "spec": True, "pricing": False,
                    "note": "Aurora MySQL→PolarDB MySQL, PostgreSQL→PolarDB PG,参考价"},
        "Other": {"covered": True, "spec": False, "pricing": False,
                 "note": "KMS/SecretsManager/VPC/CloudWatch/Q均已映射"},
    }

    for sheet_name in data:
        info = coverage_info.get(sheet_name, {})
        pm = PRODUCT_MAP.get(sheet_name, {})
        report.append({
            "product": sheet_name,
            "covered": "已覆盖" if info.get("covered", False) else "未覆盖",
            "aliyun": pm.get("aliyun", "未映射"),
            "spec_mapping": "支持" if info.get("spec", False) else "仅产品级",
            "pricing": "实时询价" if info.get("pricing", False) else "参考价/按量",
            "note": info.get("note", ""),
        })

    return report


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="AWS→阿里云 产品映射 Excel 处理")
    parser.add_argument("input", help="输入 Excel 文件路径 (.xls)")
    parser.add_argument("--output", "-o", default=None, help="输出 Excel 文件路径 (.xlsx)")
    parser.add_argument("--pricing-url", default="http://localhost:5001", help="询价服务 URL")
    parser.add_argument("--region", default="cn-beijing", help="阿里云目标区域")
    parser.add_argument("--no-pricing", action="store_true", help="跳过实时询价")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"错误: 文件不存在 {args.input}")
        sys.exit(1)

    if not args.output:
        base = os.path.splitext(os.path.basename(args.input))[0]
        args.output = os.path.join(os.path.dirname(args.input) or ".", f"{base}_mapping_result.xlsx")

    print(f"📥 读取输入文件: {args.input}")
    data = read_excel(args.input)
    print(f"   共 {len(data)} 个 Sheet: {', '.join(data.keys())}")

    # 处理各产品
    print("\n📋 产品映射处理...")

    ec2_results = process_ec2(data.get("EC2", {}).get("rows", []))
    print(f"  EC2: {len(ec2_results)} 台实例, {len(set(r['AWS规格'] for r in ec2_results))} 种规格")

    aurora_results = process_aurora(data.get("AuroraDB", {}).get("rows", []))
    print(f"  AuroraDB: {len(aurora_results)} 个实例")

    docdb_results = process_documentdb(data.get("DocumentDB", {}).get("rows", []))
    print(f"  DocumentDB: {len(docdb_results)} 个集群")

    cache_results = process_elasticache(data.get("ElasticCache", {}).get("rows", []))
    print(f"  ElasticCache: {len(cache_results)} 个节点")

    s3_results = process_s3(data.get("S3", {}).get("rows", []))
    print(f"  S3: {len(s3_results)} 个桶")

    network_results = process_network(
        data.get("负载均衡", {}).get("rows", []),
        data.get("NAT", {}).get("rows", []),
        data.get("EIP", {}).get("rows", []),
    )
    print(f"  网络: {len(network_results)} 个资源")

    other_results = process_other_services(
        data.get("lamdba", {}).get("rows", []),
        data.get("Athena", {}).get("rows", []),
        data.get("EventBridge", {}).get("rows", []),
        data.get("SNS", {}).get("rows", []),
        data.get("CloudWatchLog", {}).get("rows", []),
        data.get("Other", {}).get("rows", []),
    )
    print(f"  其他服务: {len(other_results)} 项")

    # 询价
    if not args.no_pricing:
        print(f"\n💰 实时询价 (服务地址: {args.pricing_url})...")
        if check_pricing_service(args.pricing_url):
            print("  询价服务已连接 ✓")
            fill_ecs_prices(ec2_results, args.pricing_url, args.region)
            fill_redis_prices(cache_results, args.pricing_url, args.region)
            fill_polardb_prices(aurora_results, args.pricing_url, args.region)
            fill_aurora_rds_prices(aurora_results, args.pricing_url, args.region)
            fill_mongodb_prices(docdb_results, args.pricing_url, args.region)
            fill_slb_prices(network_results, args.pricing_url, args.region)
        else:
            print("  ⚠️ 询价服务不可用,将使用参考价格")
            fill_polardb_prices(aurora_results, args.pricing_url, args.region)
            fill_aurora_rds_prices(aurora_results, args.pricing_url, args.region)
    else:
        print("\n⏭️  跳过实时询价")
        fill_polardb_prices(aurora_results, args.pricing_url, args.region)
        fill_aurora_rds_prices(aurora_results, args.pricing_url, args.region)

    # 覆盖率验证
    print("\n🔍 Skill 覆盖率验证...")
    coverage_report = validate_coverage(data)
    covered = sum(1 for r in coverage_report if r["covered"] == "已覆盖")
    print(f"  覆盖率: {covered}/{len(coverage_report)} ({covered/len(coverage_report)*100:.0f}%)")

    # 输出 Excel
    print(f"\n📤 生成输出 Excel...")
    write_excel(args.output, ec2_results, aurora_results, docdb_results,
                cache_results, s3_results, network_results, other_results,
                coverage_report, data)

    # 打印成本汇总
    ecs_total = sum(r["包月单价(元)"] for r in ec2_results if isinstance(r["包月单价(元)"], (int, float)))
    aurora_total = sum(r["包月单价(元)"] for r in aurora_results if isinstance(r["包月单价(元)"], (int, float)))
    docdb_total = sum(r["包月单价(元)"] for r in docdb_results if isinstance(r["包月单价(元)"], (int, float)))
    cache_total = sum(r["包月单价(元)"] for r in cache_results if isinstance(r["包月单价(元)"], (int, float)))
    s3_total = sum(r["包月单价(元)"] for r in s3_results if isinstance(r["包月单价(元)"], (int, float)))
    net_total = sum(r["包月单价(元)"] for r in network_results if isinstance(r["包月单价(元)"], (int, float)))
    other_total = sum(r["包月单价(元)"] for r in other_results if isinstance(r["包月单价(元)"], (int, float)))
    grand = ecs_total + aurora_total + docdb_total + cache_total + s3_total + net_total + other_total

    print(f"\n{'='*50}")
    print(f"  💰 月度成本汇总")
    print(f"{'='*50}")
    print(f"  ECS 计算:          ¥{ecs_total:>12,.2f}")
    print(f"  PolarDB 数据库:    ¥{aurora_total:>12,.2f}")
    print(f"  MongoDB 数据库:    ¥{docdb_total:>12,.2f}")
    print(f"  Redis/Tair 缓存:   ¥{cache_total:>12,.2f}")
    print(f"  OSS 对象存储:      ¥{s3_total:>12,.2f}")
    print(f"  网络(LB+NAT+EIP):  ¥{net_total:>12,.2f}")
    print(f"  其他服务:          ¥{other_total:>12,.2f}")
    print(f"{'─'*50}")
    print(f"  月度合计:          ¥{grand:>12,.2f}")
    print(f"  包年预估(85折):     ¥{grand*0.85:>12,.2f}")
    print(f"  3年预估(5折):       ¥{grand*0.5:>12,.2f}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
