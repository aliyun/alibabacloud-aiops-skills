#!/usr/bin/env python3
"""Azure 资源列表 → 阿里云映射与报价工具

读取 Azure 资源调研 Excel 文件，自动完成：
1. 产品映射（Azure → 阿里云）
2. 规格映射（Azure VM/数据库/缓存 → 阿里云 ECS/RDS/Redis/MongoDB）
3. 实时询价（调用 pricing_service.py）
4. 输出结果到 Excel

用法:
    python azure_mapping.py <input.xlsx> [--output result.xlsx] [--pricing-url http://localhost:5001] [--region cn-beijing]
"""

import argparse
import math
import re
import sys
from collections import Counter

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Azure VM → 阿里云 ECS 规格映射
# ============================================================

# Azure VM 系列 → 阿里云规格族前缀
# 命名规则: Standard_{Series}{vCPU}{options}_v{version}
AZURE_TO_ALI_FAMILY = {
    # 突发型
    "bsv2": "ecs.t6", "basv2": "ecs.t6",
    "b1ls": "ecs.t6", "b1s": "ecs.t6", "b2s": "ecs.t6", "b2ms": "ecs.t6",
    "b4ms": "ecs.t6", "b8ms": "ecs.t6", "b12ms": "ecs.t6", "b16ms": "ecs.t6",
    "b20ms": "ecs.t6",
    # 通用型 Intel (1:4)
    "dv4": "ecs.g7", "dsv4": "ecs.g7", "dv5": "ecs.g8i", "dsv5": "ecs.g8i",
    "ddsv5": "ecs.g8i", "ddv5": "ecs.g8i", "ddsv6": "ecs.g8i",
    # 通用型 AMD (1:4)
    "dasv5": "ecs.g8a", "dadsv5": "ecs.g8a", "dasv6": "ecs.g8a",
    # 通用型 ARM (1:4)
    "dpsv5": "ecs.g8y", "dplsv5": "ecs.g8y", "dpdsv6": "ecs.g8y", "dpldsv6": "ecs.g8y",
    # 计算型 Intel (1:2)
    "fsv2": "ecs.c7", "fv6": "ecs.c8i", "falsv6": "ecs.c8i",
    # 计算型 AMD (1:2)
    "fasv6": "ecs.c8a",
    # 内存型 Intel (1:8)
    "ev4": "ecs.r7", "esv4": "ecs.r7", "ev5": "ecs.r8i", "esv5": "ecs.r8i",
    "edsv5": "ecs.r8i", "ebdsv5": "ecs.r8i", "ebsv5": "ecs.r8i",
    "edsv6": "ecs.r8i",
    # 内存型 AMD (1:8)
    "easv5": "ecs.r8a", "eadsv5": "ecs.r8a", "easv6": "ecs.r8a",
    # 内存型 ARM (1:8)
    "epsv5": "ecs.r8y", "epdsv5": "ecs.r8y",
    # 超大内存型 / SAP
    "mv2": "ecs.re7", "msv2": "ecs.re7", "mdsv2": "ecs.re7",
    # 存储优化型
    "lsv3": "ecs.i3", "lasv3": "ecs.i4", "lsv2": "ecs.i3",
    # GPU 型
    "ncast4v3": "ecs.gn7i", "ncadsv4": "ecs.gn7e", "ncv3": "ecs.gn7e",
    "ncadsa100v4": "ecs.gn7e", "nca100v4": "ecs.gn7e",
    "ndv2": "ecs.gn7e", "ndsh100v5": "ecs.gn8",
    "nvv4": "ecs.vgn7i", "nvadsa10v5": "ecs.vgn7i",
    # HPC 型
    "hbv3": "ecs.hfc7", "hbv4": "ecs.hfc7", "hxv3": "ecs.hfc7", "hxv4": "ecs.hfc7",
}

# 阿里云 ECS 规格尺寸映射 (vCPU → 尺寸名)
ALI_SIZE_MAP = {
    1: "large",
    2: "large",
    4: "xlarge",
    8: "2xlarge",
    12: "3xlarge",
    16: "4xlarge",
    24: "6xlarge",
    32: "8xlarge",
    48: "13xlarge",
    60: "16xlarge",
    64: "16xlarge",
    96: "24xlarge",
    104: "26xlarge",
    128: "32xlarge",
    208: "52xlarge",
}


def parse_azure_vm_spec(spec_str: str) -> dict:
    """解析 Azure VM 规格字符串, 如 'Standard_D8s_v5' 或 '8vCPUs | 32GB | Standard_D8s_v5'

    Azure 命名规则: Standard_{Series}{vCPU}{options}_v{version}
    Examples:
      Standard_D8s_v5 → series=D, vcpu=8, options=s, version=5
      Standard_E16ds_v5 → series=E, vcpu=16, options=ds, version=5
      Standard_F8s_v2 → series=F, vcpu=8, options=s, version=2
      Standard_NC6s_v3 → series=NC, vcpu=6, options=s, version=3
      Standard_B2ms → series=B, vcpu=2, options=ms, version=0
    """
    result = {"raw": spec_str, "vcpu": 0, "mem_gb": 0, "azure_series": "", "azure_spec": ""}

    # 尝试从 pipe 分隔格式提取 vCPU/Mem (如华为云格式)
    if "|" in spec_str:
        parts = [p.strip() for p in spec_str.split("|")]
        if len(parts) >= 3:
            m = re.search(r"(\d+)\s*vCPU", parts[0], re.IGNORECASE)
            if m:
                result["vcpu"] = int(m.group(1))
            m = re.search(r"(\d+)\s*GB", parts[1], re.IGNORECASE)
            if m:
                result["mem_gb"] = int(m.group(1))
            spec_str = parts[2].strip()

    result["azure_spec"] = spec_str

    # 提取 Standard_{...} 部分
    m = re.search(r'Standard_(.+)', spec_str, re.IGNORECASE)
    if not m:
        return result

    body = m.group(1)  # e.g. "D8s_v5", "E16ds_v5", "B2ms", "NC6s_v3"

    # 解析系列名、vCPU 数、选项、版本
    # Azure 命名: Standard_{Series}{vCPU}{options}_v{version}
    # 特殊: NC4as_T4_v3 (含 GPU 型号), NCas_T4_v3 (无 vCPU 数)
    # 先尝试标准格式
    pm = re.match(r'([A-Za-z]+?)(\d+)([A-Za-z]*)(?:_v(\d+))?$', body)
    if not pm:
        # 尝试含 GPU 型号格式: NC4as_T4_v3 → series=NC, vCPU=4, options=as, version=3
        pm = re.match(r'([A-Za-z]+?)(\d+)([A-Za-z]*?)_(?:[A-Za-z]+\d*)_v(\d+)$', body)
    if not pm:
        # 尝试无 vCPU 数格式: NCas_T4_v3 (仅系列名)
        pm = re.match(r'([A-Za-z]+[a-z]*)_.*_v(\d+)$', body)
        if pm:
            # 无法从名和推断 vCPU, 使用默认
            series_letters = pm.group(1)
            vcpu = result.get("vcpu", 0) or 4
            options = ""
            version = int(pm.group(2))
            result["azure_series"] = series_letters.lower()
            if not result["vcpu"]:
                result["vcpu"] = vcpu
            result["_family_key"] = f"{series_letters.lower()}v{version}"
            result["_version"] = version
            result["_options"] = options.lower()
            # 跳过标准流程的设置
            pm = None  # 标记已手动处理
    if pm:
        series_letters = pm.group(1)
        vcpu = int(pm.group(2))
        options = pm.group(3) or ""
        version = int(pm.group(4)) if pm.group(4) else 0

        result["azure_series"] = series_letters.lower()
        if not result["vcpu"]:
            result["vcpu"] = vcpu

        # 构建 family key 用于查表
        # 尝试多种组合: 完整key → series+options+version → series+version → series
        family_key = f"{series_letters.lower()}{options.lower()}v{version}" if version else f"{series_letters.lower()}{options.lower()}"
        result["_family_key"] = family_key
        result["_version"] = version
        result["_options"] = options.lower()

    # 如果没有从 pipe 格式拿到 mem_gb, 按照系列+规格名精确推断
    if not result["mem_gb"] and result["vcpu"]:
        series = result["azure_series"]
        vcpu = result["vcpu"]
        # B 系列内存不规律, 需精确映射
        B_MEM_MAP = {1: 1, 2: 4, 4: 16, 8: 32, 12: 24, 16: 64, 20: 80}
        if series in ("b",):
            result["mem_gb"] = B_MEM_MAP.get(vcpu, vcpu * 4)
        elif series in ("d",):
            result["mem_gb"] = vcpu * 4  # D 通用型 1:4
        elif series in ("f",):
            result["mem_gb"] = vcpu * 2  # F 计算型 1:2
        elif series in ("e",):
            result["mem_gb"] = vcpu * 8  # E 内存型 1:8
        elif series in ("m",):
            result["mem_gb"] = vcpu * 14  # M 超大内存
        elif series in ("l",):
            result["mem_gb"] = result["vcpu"] * 8  # L 存储优化
        else:
            result["mem_gb"] = result["vcpu"] * 4  # 默认 1:4

    return result


def _resolve_ali_family(parsed: dict) -> str:
    """根据解析后的 Azure 规格信息查找阿里云规格族"""
    family_key = parsed.get("_family_key", "")
    version = parsed.get("_version", 0)
    options = parsed.get("_options", "")
    series = parsed.get("azure_series", "")

    # 直接精确匹配
    if family_key in AZURE_TO_ALI_FAMILY:
        return AZURE_TO_ALI_FAMILY[family_key]

    # 去掉 version 后缀匹配: e.g. "dsv5" → "dsv5"
    key_no_v = f"{series}{options}"
    for k, v in AZURE_TO_ALI_FAMILY.items():
        if k.startswith(key_no_v):
            return v

    # 按系列字母推断
    if series in ("b",):
        return "ecs.t6"
    elif series in ("d",):
        return "ecs.g8i" if version >= 5 else "ecs.g7"
    elif series in ("f",):
        return "ecs.c8i" if version >= 6 else "ecs.c7"
    elif series in ("e",):
        return "ecs.r8i" if version >= 5 else "ecs.r7"
    elif series in ("m",):
        return "ecs.re7"
    elif series in ("l",):
        return "ecs.i3"
    elif series in ("n",) or series.startswith("nc") or series.startswith("nd") or series.startswith("nv"):
        return "ecs.gn7i"
    elif series in ("h",) or series.startswith("hb") or series.startswith("hx"):
        return "ecs.hfc7"
    else:
        # 按内存比推断
        vcpu = parsed.get("vcpu", 0)
        mem = parsed.get("mem_gb", 0)
        if vcpu > 0 and mem > 0:
            ratio = mem / vcpu
            if ratio <= 2:
                return "ecs.c7"
            elif ratio <= 4:
                return "ecs.g7"
            elif ratio <= 8:
                return "ecs.r7"
            else:
                return "ecs.re7"
        return "ecs.g7"


def map_azure_vm(parsed: dict) -> dict:
    """将 Azure VM 规格映射到阿里云 ECS 规格"""
    vcpu = parsed["vcpu"]
    mem_gb = parsed["mem_gb"]

    ali_family = _resolve_ali_family(parsed)

    # 突发型限制: t6 仅支持 large (2vCPU)
    # 回退策略: Azure B 系列为 x86 架构, 回退到 Intel 规格
    if ali_family == "ecs.t6" and vcpu > 2:
        ratio = mem_gb / vcpu if vcpu > 0 else 4
        if ratio <= 2:
            ali_family = "ecs.c8i"   # Intel 计算型 1:2
        elif ratio <= 4:
            ali_family = "ecs.g8i"   # Intel 通用型 1:4
        else:
            ali_family = "ecs.r8i"   # Intel 内存型 1:8

    # 确定尺寸
    size = ALI_SIZE_MAP.get(vcpu)
    if not size:
        for k in sorted(ALI_SIZE_MAP.keys()):
            if k >= vcpu:
                size = ALI_SIZE_MAP[k]
                vcpu = k
                break
        if not size:
            size = "32xlarge"
            vcpu = 128

    ali_spec = f"{ali_family}.{size}"

    note = ""
    if vcpu >= 104:
        note = " [超大规格,需确认可用性]"

    ratio = mem_gb / vcpu if vcpu > 0 else 2
    if ratio <= 2:
        type_desc = "计算优化型"
    elif ratio <= 4:
        type_desc = "通用型"
    elif ratio <= 8:
        type_desc = "内存优化型"
    else:
        type_desc = "大内存型"

    return {
        "aliyun_spec": ali_spec,
        "aliyun_desc": f"{type_desc} {vcpu}C{mem_gb}G{note}",
        "vcpu": vcpu,
        "mem_gb": mem_gb,
    }


# ============================================================
# 环境推断 (prod/beta/test)
# ============================================================

def infer_env(name: str, project: str = "") -> str:
    """从名称和项目推断环境: prod/beta/test"""
    text = f"{name} {project}".lower()
    if "prod" in text or "prd" in text:
        return "prod"
    elif "beta" in text or "demo" in text or "staging" in text or "gray" in text or "stg" in text:
        return "beta"
    elif "test" in text or "dev" in text or "sandbox" in text:
        return "test"
    return "prod"  # 默认 prod


# ============================================================
# Azure 数据库 → 阿里云 RDS 规格映射
# ============================================================

# Azure Database for MySQL/PostgreSQL 规格映射
# 规格命名: {SKU类型}_Standard_{Series}{vCPU}{options}
#   B = 突发型, D = 通用型, MO = 内存优化型
AZURE_DB_SPEC_MAP = {
    # B 突发型
    "b_standard_b1s":   {"vcpu": 1,  "mem_gb": 2,   "mysql": "rds.mysql.s1.small",     "pg": "pg.n2.small.1"},
    "b_standard_b1ms":  {"vcpu": 1,  "mem_gb": 2,   "mysql": "rds.mysql.s1.small",     "pg": "pg.n2.small.1"},
    "b_standard_b2s":   {"vcpu": 2,  "mem_gb": 4,   "mysql": "rds.mysql.s2.large",     "pg": "pg.n2.small.2c"},
    "b_standard_b2ms":  {"vcpu": 2,  "mem_gb": 4,   "mysql": "rds.mysql.s2.large",     "pg": "pg.n2.small.2c"},
    "b_standard_b4ms":  {"vcpu": 4,  "mem_gb": 8,   "mysql": "rds.mysql.s3.small",     "pg": "pg.n4.medium.1"},
    "b_standard_b8ms":  {"vcpu": 8,  "mem_gb": 16,  "mysql": "rds.mysql.c1.medium",    "pg": "pg.n4.large.1"},
    "b_standard_b12ms": {"vcpu": 12, "mem_gb": 24,  "mysql": "rds.mysql.c1.xlarge",    "pg": "pg.n4.xlarge.1"},
    "b_standard_b16ms": {"vcpu": 16, "mem_gb": 32,  "mysql": "rds.mysql.c1.xlarge",    "pg": "pg.n4.xlarge.1"},
    "b_standard_b20ms": {"vcpu": 20, "mem_gb": 80,  "mysql": "rds.mysql.m1.xlarge",    "pg": "pg.n2.xlarge.2c"},
    # D 通用型
    "d_standard_d2ds":  {"vcpu": 2,  "mem_gb": 8,   "mysql": "rds.mysql.s2.xlarge",    "pg": "pg.n2.medium.1"},
    "d_standard_d2s":   {"vcpu": 2,  "mem_gb": 8,   "mysql": "rds.mysql.s2.xlarge",    "pg": "pg.n2.medium.1"},
    "d_standard_d4ds":  {"vcpu": 4,  "mem_gb": 16,  "mysql": "rds.mysql.s3.large",     "pg": "pg.n4.medium.1"},
    "d_standard_d4s":   {"vcpu": 4,  "mem_gb": 16,  "mysql": "rds.mysql.s3.large",     "pg": "pg.n4.medium.1"},
    "d_standard_d8ds":  {"vcpu": 8,  "mem_gb": 32,  "mysql": "rds.mysql.m1.medium",    "pg": "pg.n2.large.2c"},
    "d_standard_d8s":   {"vcpu": 8,  "mem_gb": 32,  "mysql": "rds.mysql.m1.medium",    "pg": "pg.n2.large.2c"},
    "d_standard_d16ds": {"vcpu": 16, "mem_gb": 64,  "mysql": "rds.mysql.m1.xlarge",    "pg": "pg.n2.xlarge.2c"},
    "d_standard_d16s":  {"vcpu": 16, "mem_gb": 64,  "mysql": "rds.mysql.m1.xlarge",    "pg": "pg.n2.xlarge.2c"},
    "d_standard_d32ds": {"vcpu": 32, "mem_gb": 128, "mysql": "rds.mysql.m1.2xlarge",   "pg": "pg.n4.2xlarge.1"},
    "d_standard_d32s":  {"vcpu": 32, "mem_gb": 128, "mysql": "rds.mysql.m1.2xlarge",   "pg": "pg.n4.2xlarge.1"},
    "d_standard_d48ds": {"vcpu": 48, "mem_gb": 192, "mysql": "rds.mysql.m1.4xlarge",   "pg": "pg.n4.4xlarge.1"},
    "d_standard_d64ds": {"vcpu": 64, "mem_gb": 256, "mysql": "rds.mysql.m1.4xlarge",   "pg": "pg.n4.4xlarge.1"},
    # MO 内存优化型
    "mo_standard_e2ds":  {"vcpu": 2,  "mem_gb": 16,  "mysql": "rds.mysql.s2.xlarge",   "pg": "pg.n2.medium.1"},
    "mo_standard_e4ds":  {"vcpu": 4,  "mem_gb": 32,  "mysql": "rds.mysql.s3.large",    "pg": "pg.n4.medium.1"},
    "mo_standard_e8ds":  {"vcpu": 8,  "mem_gb": 64,  "mysql": "rds.mysql.m1.medium",   "pg": "pg.n2.large.2c"},
    "mo_standard_e16ds": {"vcpu": 16, "mem_gb": 128, "mysql": "rds.mysql.m1.xlarge",   "pg": "pg.n2.xlarge.2c"},
    "mo_standard_e32ds": {"vcpu": 32, "mem_gb": 256, "mysql": "rds.mysql.m1.2xlarge",  "pg": "pg.n4.2xlarge.1"},
    "mo_standard_e48ds": {"vcpu": 48, "mem_gb": 384, "mysql": "rds.mysql.m1.4xlarge",  "pg": "pg.n4.4xlarge.1"},
    "mo_standard_e64ds": {"vcpu": 64, "mem_gb": 512, "mysql": "rds.mysql.m1.4xlarge",  "pg": "pg.n4.4xlarge.1"},
    "mo_standard_e96ds": {"vcpu": 96, "mem_gb": 672, "mysql": "rds.mysql.m1.8xlarge",  "pg": "pg.n8.8xlarge.1"},
}

# vCPU/内存 → 阿里云 RDS 规格 (回退用)
AZURE_RDS_VCPU_MAP = {
    (1, 2):    ("rds.mysql.s1.small",    "pg.n2.small.1"),
    (2, 4):    ("rds.mysql.s2.large",    "pg.n2.small.2c"),
    (2, 8):    ("rds.mysql.s2.xlarge",   "pg.n2.medium.1"),
    (4, 8):    ("rds.mysql.s3.small",    "pg.n4.medium.1"),
    (4, 16):   ("rds.mysql.s3.large",    "pg.n4.medium.1"),
    (8, 16):   ("rds.mysql.c1.medium",   "pg.n4.large.1"),
    (8, 32):   ("rds.mysql.m1.medium",   "pg.n2.large.2c"),
    (16, 32):  ("rds.mysql.c1.xlarge",   "pg.n4.xlarge.1"),
    (16, 64):  ("rds.mysql.m1.xlarge",   "pg.n2.xlarge.2c"),
    (32, 128): ("rds.mysql.m1.2xlarge",  "pg.n4.2xlarge.1"),
    (64, 256): ("rds.mysql.m1.4xlarge",  "pg.n4.4xlarge.1"),
}

# Azure Cache for Redis → 阿里云 Redis 规格映射
AZURE_REDIS_SPEC_MAP = {
    # (tier, level): (ali_spec, capacity_mb)
    ("basic", "c0"):    ("redis.amber.master.small.multithread", 256),
    ("basic", "c1"):    ("redis.amber.master.small.multithread", 1024),
    ("basic", "c2"):    ("redis.amber.master.mid.multithread", 2048),
    ("basic", "c3"):    ("redis.amber.master.large.multithread", 8192),
    ("standard", "c0"): ("redis.amber.master.small.multithread", 256),
    ("standard", "c1"): ("redis.amber.master.small.multithread", 1024),
    ("standard", "c2"): ("redis.amber.master.mid.multithread", 2048),
    ("standard", "c3"): ("redis.amber.master.large.multithread", 8192),
    ("standard", "c4"): ("redis.amber.master.2xlarge.multithread", 16384),
    ("standard", "c5"): ("redis.amber.master.4xlarge.multithread", 32768),
    ("standard", "c6"): ("redis.amber.master.8xlarge.multithread", 65536),
    ("premium", "p1"):  ("redis.amber.master.large.multithread", 8192),
    ("premium", "p2"):  ("redis.amber.master.2xlarge.multithread", 16384),
    ("premium", "p3"):  ("redis.amber.master.4xlarge.multithread", 32768),
    ("premium", "p4"):  ("redis.amber.master.8xlarge.multithread", 65536),
    ("premium", "p5"):  ("redis.shard.amber.ce13.5.default", 131072),
}

# Redis 容量 → 阿里云 Redis 规格 (按 MB 回退)
REDIS_CAP_MAP = {
    256:   ("redis.amber.master.small.multithread", 256),
    1024:  ("redis.amber.master.small.multithread", 1024),
    2048:  ("redis.amber.master.mid.multithread", 2048),
    4096:  ("redis.amber.master.stand.multithread", 4096),
    8192:  ("redis.amber.master.large.multithread", 8192),
    16384: ("redis.amber.master.2xlarge.multithread", 16384),
    32768: ("redis.amber.master.4xlarge.multithread", 32768),
    65536: ("redis.amber.master.8xlarge.multithread", 65536),
}

# Cosmos DB (MongoDB API) RU/s → 阿里云 MongoDB 规格映射
COSMOS_RU_MAP = [
    (1000,  "mdb.shard.2x.large.d",    "独享型 2C4G"),
    (4000,  "mdb.shard.2x.xlarge.d",   "独享型 4C8G"),
    (10000, "mdb.shard.2x.2xlarge.d",  "独享型 8C16G"),
    (50000, "mdb.shard.2x.4xlarge.d",  "独享型 16C32G"),
    (999999999, "mdb.shard.2x.8xlarge.d", "独享型 32C64G"),
]

# Azure MongoDB (vCosmosDB) vCPU/内存 → 阿里云 MongoDB
AZURE_MONGODB_SPEC_MAP = {
    (2, 4):   "mdb.shard.2x.large.d",
    (4, 8):   "mdb.shard.2x.xlarge.d",
    (8, 16):  "mdb.shard.2x.2xlarge.d",
    (16, 32): "mdb.shard.2x.4xlarge.d",
    (32, 64): "mdb.shard.2x.8xlarge.d",
}


# ============================================================
# 主处理函数
# ============================================================

def _parse_vcpu_mem(spec_str: str) -> tuple:
    """尝试从规格字符串解析 vCPU 和内存, 返回 (vcpu, mem_gb) 或 (0, 0)"""
    vcpu, mem = 0, 0
    m = re.search(r'(\d+)\s*vCPU', spec_str, re.IGNORECASE)
    if m:
        vcpu = int(m.group(1))
    m = re.search(r'(\d+)\s*GB', spec_str, re.IGNORECASE)
    if m:
        mem = int(m.group(1))
    if not vcpu:
        m = re.search(r'(\d+)C(\d+)G', spec_str)
        if m:
            vcpu, mem = int(m.group(1)), int(m.group(2))
    return vcpu, mem


def process_vm(wb) -> list:
    """处理 Azure VM 虚拟机"""
    results = []
    # 尝试多个可能的 Sheet 名
    sheet_names = [
        "Virtual Machines", "虚拟机", "VM", "VirtualMachines",
        "Compute", "Azure VM",
    ]
    ws = None
    for sn in sheet_names:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        # 尝试模糊匹配
        for sn in wb.sheetnames:
            if "vm" in sn.lower() or "virtual" in sn.lower() or "compute" in sn.lower():
                ws = wb[sn]
                break
    if ws is None:
        return results

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
        region = str(row[2] or "") if len(row) > 2 else ""
        project = str(row[3] or "") if len(row) > 3 else ""
        spec_str = str(row[7] or "") if len(row) > 7 else str(row[4] or "") if len(row) > 4 else ""
        image = str(row[8] or "") if len(row) > 8 else ""

        parsed = parse_azure_vm_spec(spec_str if spec_str else name)
        ali = map_azure_vm(parsed)
        env = infer_env(name, project)

        sys_disk_desc = "ESSD PL0 80GB"

        results.append({
            "Azure资源": name,
            "Azure规格": spec_str if spec_str else parsed["azure_spec"],
            "环境": env,
            "区域": region,
            "操作系统": image[:30] if image else "",
            "阿里云产品": "ECS 云服务器",
            "阿里云规格": ali["aliyun_spec"],
            "阿里云规格描述": ali["aliyun_desc"],
            "阿里云系统盘": sys_disk_desc,
            "阿里云数据盘": "",
            "包月单价(元)": 0,
        })
    return results


def process_db(wb) -> list:
    """处理 Azure Database for MySQL/PostgreSQL"""
    results = []
    sheet_names = [
        "Azure Database", "Database", "MySQL", "PostgreSQL",
        "Azure SQL", "SQL Database", "DB",
    ]
    ws = None
    for sn in sheet_names:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        for sn in wb.sheetnames:
            if "database" in sn.lower() or "mysql" in sn.lower() or "postgres" in sn.lower() or "sql" in sn.lower():
                ws = wb[sn]
                break
    if ws is None:
        return results

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
        project = str(row[3] or "") if len(row) > 3 else ""
        spec_str = str(row[7] or "") if len(row) > 7 else str(row[4] or "") if len(row) > 4 else ""
        env = infer_env(name, project)

        # 推断引擎
        engine = "MySQL"
        engine_ver = "8.0"
        name_lower = name.lower()
        if "pg" in name_lower or "postgres" in name_lower:
            engine = "PostgreSQL"
            engine_ver = "16.0"
        elif "sqlserver" in name_lower or "mssql" in name_lower:
            engine = "SQLServer"
            engine_ver = "2019_ent"

        is_pg = engine == "PostgreSQL"

        # 尝试从规格名精确匹配
        spec_key = spec_str.lower().replace(" ", "_").replace("-", "_")
        ali_spec = None
        source = "精确"

        if spec_key in AZURE_DB_SPEC_MAP:
            info = AZURE_DB_SPEC_MAP[spec_key]
            ali_spec = info["pg"] if is_pg else info["mysql"]
            vcpu, mem_gb = info["vcpu"], info["mem_gb"]
            storage = max(100, mem_gb * 10)
        else:
            # 尝试解析 vCPU/内存
            vcpu, mem_gb = _parse_vcpu_mem(spec_str)
            if not vcpu:
                vcpu, mem_gb = _parse_vcpu_mem(name)

            if vcpu and mem_gb:
                key = (vcpu, mem_gb)
                if key in AZURE_RDS_VCPU_MAP:
                    mysql_spec, pg_spec = AZURE_RDS_VCPU_MAP[key]
                    ali_spec = pg_spec if is_pg else mysql_spec
                else:
                    best_key = min(AZURE_RDS_VCPU_MAP.keys(),
                                   key=lambda k: abs(k[0] - vcpu) + abs(k[1] - mem_gb) * 0.5)
                    mysql_spec, pg_spec = AZURE_RDS_VCPU_MAP[best_key]
                    ali_spec = pg_spec if is_pg else mysql_spec
                storage = max(100, mem_gb * 10)
            else:
                source = "推断"
                if env == "prod":
                    ali_spec = "pg.n2.large.2c" if is_pg else "rds.mysql.m1.medium"
                    vcpu, mem_gb = 8, 32
                else:
                    ali_spec = "pg.n2.small.2c" if is_pg else "rds.mysql.s2.large"
                    vcpu, mem_gb = 2, 4
                storage = 200 if env == "prod" else 100

        ali_desc = f"{engine} {vcpu}C{mem_gb}G 高可用版"
        confirm_tag = "" if source == "精确" else " [需确认]"

        results.append({
            "Azure资源": name,
            "Azure规格": spec_str if spec_str else f"{vcpu}C{mem_gb}G ({source})",
            "环境": env,
            "阿里云产品": f"RDS {engine}",
            "阿里云规格": ali_spec,
            "阿里云规格描述": f"{ali_desc} (存储{storage}GB){confirm_tag}",
            "引擎": engine,
            "引擎版本": engine_ver,
            "存储(GB)": storage,
            "包月单价(元)": 0,
        })
    return results


def process_redis(wb) -> list:
    """处理 Azure Cache for Redis"""
    results = []
    sheet_names = [
        "Azure Cache for Redis", "Redis", "Cache", "Azure Redis",
    ]
    ws = None
    for sn in sheet_names:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        for sn in wb.sheetnames:
            if "redis" in sn.lower() or "cache" in sn.lower():
                ws = wb[sn]
                break
    if ws is None:
        return results

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
        project = str(row[3] or "") if len(row) > 3 else ""
        spec_str = str(row[7] or "") if len(row) > 7 else str(row[4] or "") if len(row) > 4 else ""
        env = infer_env(name, project)

        source = "精确"
        ali_spec = None
        ali_cap = 0

        # 尝试从 Azure 规格层级匹配 (如 "Standard C3", "Premium P2")
        text = f"{name} {spec_str}".lower()
        for (tier, level), (spec, cap) in AZURE_REDIS_SPEC_MAP.items():
            if tier in text and level in text:
                ali_spec = spec
                ali_cap = cap
                break

        if not ali_spec:
            # 尝试从内存容量匹配
            m = re.search(r'(\d+)\s*GB', f"{name} {spec_str}", re.IGNORECASE)
            if m:
                cap_mb = int(m.group(1)) * 1024
                for cap in sorted(REDIS_CAP_MAP.keys()):
                    if cap >= cap_mb:
                        ali_spec, ali_cap = REDIS_CAP_MAP[cap]
                        break
                if not ali_spec:
                    ali_spec, ali_cap = REDIS_CAP_MAP[65536]
            else:
                m = re.search(r'(\d+)\s*MB', f"{name} {spec_str}", re.IGNORECASE)
                if m:
                    cap_mb = int(m.group(1))
                    for cap in sorted(REDIS_CAP_MAP.keys()):
                        if cap >= cap_mb:
                            ali_spec, ali_cap = REDIS_CAP_MAP[cap]
                            break
                    if not ali_spec:
                        ali_spec, ali_cap = REDIS_CAP_MAP[65536]

        if not ali_spec:
            source = "推断"
            if env == "prod":
                ali_spec = "redis.amber.master.2xlarge.multithread"
                ali_cap = 16384
            else:
                ali_spec = "redis.amber.master.stand.multithread"
                ali_cap = 4096

        cap_desc = f"{ali_cap // 1024}GB" if ali_cap >= 1024 else f"{ali_cap}MB"
        confirm_tag = "" if source == "精确" else " [需确认]"

        results.append({
            "Azure资源": name,
            "Azure规格": spec_str,
            "环境": env,
            "阿里云产品": "Redis/Tair 云原生版",
            "阿里云规格": ali_spec,
            "阿里云规格描述": f"云原生版 {cap_desc} 标准版-双副本{confirm_tag}",
            "容量(MB)": ali_cap,
            "包月单价(元)": 0,
        })
    return results


def process_cosmosdb(wb) -> list:
    """处理 Azure Cosmos DB (MongoDB API) → 阿里云 MongoDB"""
    results = []
    sheet_names = [
        "Cosmos DB", "CosmosDB", "MongoDB", "DocumentDB",
    ]
    ws = None
    for sn in sheet_names:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        for sn in wb.sheetnames:
            if "cosmos" in sn.lower() or "mongo" in sn.lower():
                ws = wb[sn]
                break
    if ws is None:
        return results

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
        project = str(row[3] or "") if len(row) > 3 else ""
        spec_str = str(row[7] or "") if len(row) > 7 else str(row[4] or "") if len(row) > 4 else ""
        env = infer_env(name, project)

        source = "精确"
        ali_spec = None
        ali_desc = ""

        # 尝试解析 RU/s
        m = re.search(r'(\d+)\s*RU', f"{name} {spec_str}", re.IGNORECASE)
        if m:
            ru = int(m.group(1))
            for threshold, spec, desc in COSMOS_RU_MAP:
                if ru <= threshold:
                    ali_spec = spec
                    ali_desc = f"MongoDB {desc} ({ru} RU/s 等效)"
                    break
        else:
            # 尝试解析 vCPU/内存
            vcpu, mem_gb = _parse_vcpu_mem(spec_str)
            if not vcpu:
                vcpu, mem_gb = _parse_vcpu_mem(name)
            if vcpu and mem_gb:
                key = (vcpu, mem_gb)
                if key in AZURE_MONGODB_SPEC_MAP:
                    ali_spec = AZURE_MONGODB_SPEC_MAP[key]
                else:
                    best_key = min(AZURE_MONGODB_SPEC_MAP.keys(),
                                   key=lambda k: abs(k[0] - vcpu) + abs(k[1] - mem_gb) * 0.5)
                    ali_spec = AZURE_MONGODB_SPEC_MAP[best_key]
                ali_desc = f"MongoDB 独享型 {vcpu}C{mem_gb}G"

        if not ali_spec:
            source = "推断"
            if env == "prod":
                ali_spec = "mdb.shard.2x.xlarge.d"
                ali_desc = "MongoDB 独享型 4C8G"
            else:
                ali_spec = "mdb.shard.2x.large.d"
                ali_desc = "MongoDB 独享型 2C4G"

        confirm_tag = "" if source == "精确" else " [需确认]"

        results.append({
            "Azure资源": name,
            "Azure规格": spec_str,
            "环境": env,
            "阿里云产品": "云数据库 MongoDB 版",
            "阿里云规格": ali_spec,
            "阿里云规格描述": f"{ali_desc}{confirm_tag}",
            "磁盘(GB)": 40,
            "包月单价(元)": 0,
        })
    return results


def process_network(wb) -> list:
    """处理网络资源 (Load Balancer, Application Gateway, Public IP, NAT Gateway)"""
    NLB_HOURLY = 0.147
    HOURS_PER_MONTH = 730
    monthly_lb = round(NLB_HOURLY * HOURS_PER_MONTH, 2)

    results = []

    # Load Balancer / Application Gateway
    for sn in wb.sheetnames:
        sn_lower = sn.lower()
        if "load" in sn_lower or "balancer" in sn_lower or "gateway" in sn_lower or "lb" == sn_lower:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not row[0]:
                    continue
                name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
                project = str(row[3] or "") if len(row) > 3 else ""
                results.append({
                    "Azure资源": f"LB: {name}",
                    "Azure规格": sn,
                    "环境": infer_env(name, project),
                    "阿里云产品": "ALB 应用负载均衡",
                    "阿里云规格": "按量付费",
                    "阿里云规格描述": f"ALB 标准版 按量付费(实例费¥{NLB_HOURLY}/h, 不含LCU费)",
                    "包月单价(元)": monthly_lb,
                })

    # Public IP
    for sn in wb.sheetnames:
        sn_lower = sn.lower()
        if "public ip" in sn_lower or "eip" in sn_lower or "publicip" in sn_lower:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not row[0]:
                    continue
                name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
                project = str(row[3] or "") if len(row) > 3 else ""
                # 尝试解析带宽
                text = f"{name} {row[4] if len(row) > 4 else ''} {row[7] if len(row) > 7 else ''}"
                m = re.search(r'(\d+)\s*Mbps', text, re.IGNORECASE)
                mbps = int(m.group(1)) if m else 5
                results.append({
                    "Azure资源": f"EIP: {name}",
                    "Azure规格": f"{mbps} Mbps",
                    "环境": infer_env(name, project),
                    "阿里云产品": "弹性公网IP (EIP)",
                    "阿里云规格": f"EIP {mbps}Mbps 按固定带宽",
                    "阿里云规格描述": f"BGP多线 {mbps}Mbps 按固定带宽计费",
                    "带宽(Mbps)": mbps,
                    "包月单价(元)": 0,
                })

    # NAT Gateway
    for sn in wb.sheetnames:
        if "nat" in sn.lower():
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not row[0]:
                    continue
                name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
                project = str(row[3] or "") if len(row) > 3 else ""
                results.append({
                    "Azure资源": f"NAT: {name}",
                    "Azure规格": "NAT Gateway",
                    "环境": infer_env(name, project),
                    "阿里云产品": "NAT 网关",
                    "阿里云规格": "增强型小规格",
                    "阿里云规格描述": "公网NAT网关 增强型小规格(连接数10万, 新建5000/s)",
                    "包月单价(元)": 0,
                })
    return results


def process_other_services(wb) -> list:
    """处理其他服务 (仅产品映射, 不询价)"""
    MAPPING = {
        "storage account": ("OSS 对象存储", "对象存储"),
        "blob": ("OSS 对象存储", "对象存储"),
        "file share": ("NAS 文件存储", "文件存储"),
        "azure files": ("NAS 文件存储", "文件存储"),
        "container": ("ACK 容器服务", "容器"),
        "aks": ("ACK 容器服务", "容器"),
        "kubernetes": ("ACK 容器服务", "容器"),
        "function": ("函数计算 FC", "Serverless"),
        "cdn": ("CDN", "CDN"),
        "front door": ("CDN + WAF", "CDN"),
        "vpn": ("VPN 网关", "网络"),
        "expressroute": ("高速通道", "网络"),
        "dns": ("云解析 PrivateZone", "DNS"),
        "key vault": ("KMS 密钥管理", "安全"),
        "security center": ("云安全中心", "安全"),
        "monitor": ("SLS 日志服务", "日志"),
        "log analytics": ("SLS 日志服务", "日志"),
        "event grid": ("事件总线 EventBridge", "事件"),
        "event hub": ("消息队列 Kafka 版", "消息队列"),
        "service bus": ("消息服务 MNS", "消息队列"),
        "search": ("Elasticsearch", "搜索"),
        "backup": ("云备份 HBR", "备份"),
        "synapse": ("MaxCompute", "大数据"),
        "data factory": ("DataWorks", "大数据"),
        "vnet": ("VPC 专有网络", "网络(免费)"),
        "nsg": ("安全组", "网络(免费)"),
    }

    results = []
    processed_sheets = set()

    # 已被其他函数处理的 Sheet 关键词
    skip_keywords = ["vm", "virtual", "compute", "database", "mysql", "postgres",
                     "sql", "redis", "cache", "cosmos", "mongo",
                     "load", "balancer", "gateway", "public ip", "eip", "nat"]

    for sn in wb.sheetnames:
        sn_lower = sn.lower()
        if any(kw in sn_lower for kw in skip_keywords):
            continue
        if sn in processed_sheets:
            continue
        processed_sheets.add(sn)

        # 匹配已知产品
        ali_name, category = None, None
        for keyword, (ali, cat) in MAPPING.items():
            if keyword in sn_lower:
                ali_name, category = ali, cat
                break

        if not ali_name:
            ali_name = f"对标产品(待确认): {sn}"
            category = "其他"

        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            name = str(row[1] or "") if len(row) > 1 else str(row[0] or "")
            results.append({
                "Azure产品": sn,
                "Azure资源": name,
                "阿里云产品": ali_name,
                "分类": category,
                "备注": "仅映射, 需根据实际用量报价" if "免费" not in category else "免费",
            })
    return results


# ============================================================
# 询价函数
# ============================================================

def batch_pricing(pricing_url: str, items: list, region_id: str) -> dict:
    """调用批量询价 API"""
    try:
        resp = requests.post(
            f"{pricing_url}/api/pricing/batch",
            json={"region_id": region_id, "period": 1, "items": items},
            timeout=120,
        )
        data = resp.json()
        return data.get("data", {})
    except Exception as e:
        print(f"  询价失败: {e}")
        return {}


def fill_ecs_prices(ecs_results: list, pricing_url: str, region_id: str):
    """ECS 批量询价"""
    spec_set = {}
    for r in ecs_results:
        spec = r["阿里云规格"]
        data_disk_desc = r.get("阿里云数据盘", "")
        key = (spec, data_disk_desc)
        if key not in spec_set:
            data_disk_size = 0
            data_disk_pl = "PL1"
            if data_disk_desc:
                m = re.search(r"(\d+)\s*GB", data_disk_desc)
                if m:
                    data_disk_size = int(m.group(1))
                if "PL0" in data_disk_desc:
                    data_disk_pl = "PL0"
            item = {
                "product_code": "ecs",
                "instance_spec": spec,
                "disk_size": 80,
                "count": 1,
            }
            if data_disk_size > 0:
                item["data_disk_size"] = data_disk_size
                item["data_disk_pl"] = data_disk_pl
            spec_set[key] = item

    items = list(spec_set.values())
    if not items:
        return

    print(f"  询价 ECS: {len(items)} 种配置...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    filled = 0
    for r in ecs_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]
            filled += 1

    unfilled = [r for r in ecs_results if r["包月单价(元)"] == 0]
    for r in unfilled:
        desc = r.get("阿里云规格描述", "")
        m = re.search(r"(\d+)C", desc)
        if m:
            vcpu = int(m.group(1))
            r["包月单价(元)"] = round(vcpu * 50, 2)
            r["阿里云规格描述"] += " [参考估价]"
            filled += 1

    print(f"    成功填充 {filled}/{len(ecs_results)} 台价格")


def fill_rds_prices(rds_results: list, pricing_url: str, region_id: str):
    """RDS 批量询价"""
    items = []
    spec_set = set()
    for r in rds_results:
        spec = r["阿里云规格"]
        if spec not in spec_set:
            spec_set.add(spec)
            items.append({
                "product_code": "rds",
                "instance_spec": spec,
                "engine": r.get("引擎", "MySQL"),
                "engine_version": r.get("引擎版本", "8.0"),
                "disk_size": r.get("存储(GB)", 200),
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 RDS: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in rds_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


def fill_redis_prices(redis_results: list, pricing_url: str, region_id: str):
    """Redis 批量询价"""
    items = []
    spec_set = set()
    for r in redis_results:
        spec = r["阿里云规格"]
        if spec not in spec_set:
            spec_set.add(spec)
            items.append({
                "product_code": "redis",
                "instance_spec": spec,
                "capacity": r.get("容量(MB)", 4096),
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 Redis: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in redis_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


def fill_mongodb_prices(mongodb_results: list, pricing_url: str, region_id: str):
    """MongoDB 批量询价"""
    items = []
    spec_set = set()
    for r in mongodb_results:
        spec = r["阿里云规格"]
        if spec not in spec_set:
            spec_set.add(spec)
            items.append({
                "product_code": "mongodb",
                "instance_spec": spec,
                "disk_size": r.get("磁盘(GB)", 40),
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 MongoDB: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in mongodb_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


def fill_eip_prices(network_results: list, pricing_url: str, region_id: str):
    """EIP 批量询价"""
    eip_results = [r for r in network_results if "EIP" in r.get("阿里云产品", "")]
    bw_set = set()
    items = []
    for r in eip_results:
        bw = r.get("带宽(Mbps)", 5)
        if bw not in bw_set:
            bw_set.add(bw)
            items.append({
                "product_code": "eip",
                "instance_spec": f"{bw}Mbps",
                "bandwidth": bw,
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 EIP: {len(items)} 种带宽...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in eip_results:
        bw = r.get("带宽(Mbps)", 5)
        key = f"{bw}Mbps"
        if key in price_map:
            r["包月单价(元)"] = price_map[key]
        else:
            if bw <= 5:
                r["包月单价(元)"] = bw * 23
            else:
                r["包月单价(元)"] = 5 * 23 + (bw - 5) * 80


def fill_nat_prices(network_results: list, pricing_url: str, region_id: str):
    """NAT 网关询价"""
    nat_results = [r for r in network_results if "NAT" in r.get("阿里云产品", "")]
    if not nat_results:
        return
    items = [{"product_code": "nat", "instance_spec": "Small", "count": len(nat_results)}]
    print(f"  询价 NAT: {len(nat_results)} 个...")
    result = batch_pricing(pricing_url, items, region_id)

    price = 220
    for p in result.get("prices", []):
        if p.get("success"):
            price = p["trade_price"]

    for r in nat_results:
        r["包月单价(元)"] = price


# ============================================================
# Excel 输出
# ============================================================

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
MONEY_FMT = '#,##0.00'


def write_sheet(ws, title: str, headers: list, rows: list):
    """写入一个工作表"""
    ws.title = title
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for row_idx, row_data in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER
            if h in ("包月单价(元)", "月度小计(元)", "月度费用(元)"):
                cell.number_format = MONEY_FMT

    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row_idx in range(2, min(len(rows) + 2, 20)):
            val = str(ws.cell(row=row_idx, column=col).value or "")
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 4


def generate_output(output_path: str, vm_results, db_results, redis_results,
                    cosmosdb_results, network_results, other_results, region_id):
    """生成输出 Excel"""
    wb_out = openpyxl.Workbook()

    # Sheet 1: 资源映射总览
    overview = []
    lb_results = [r for r in network_results if "ALB" in r.get("阿里云产品", "")]
    eip_results = [r for r in network_results if "EIP" in r.get("阿里云产品", "")]
    nat_results = [r for r in network_results if "NAT" in r.get("阿里云产品", "")]

    categories = [
        ("Azure VM", len(vm_results), "ECS 云服务器", "SMC 服务器迁移"),
        ("Azure Database", len(db_results), "RDS MySQL/PostgreSQL", "DTS 数据迁移"),
        ("Azure Redis", len(redis_results), "Redis/Tair 云原生版", "DTS/Redis-shake"),
        ("Cosmos DB/MongoDB", len(cosmosdb_results), "云数据库 MongoDB 版", "DTS 数据迁移"),
        ("Load Balancer", len(lb_results), "ALB 应用负载均衡", "手动重建"),
        ("Public IP", len(eip_results), "弹性公网IP (EIP)", "手动分配"),
        ("NAT Gateway", len(nat_results), "NAT 网关", "手动重建"),
    ]
    other_cats = Counter()
    for o in other_results:
        key = (o.get("Azure产品", ""), o.get("阿里云产品", ""))
        other_cats[key] += 1
    for (az, ali), cnt in other_cats.most_common():
        categories.append((az, cnt, ali, "手动重建"))

    for i, (az, cnt, ali, tool) in enumerate(categories, 1):
        overview.append({
            "序号": i,
            "Azure产品": az,
            "数量": cnt,
            "阿里云产品": ali,
            "推荐迁移工具": tool,
        })

    ws1 = wb_out.active
    write_sheet(ws1, "资源映射总览",
                ["序号", "Azure产品", "数量", "阿里云产品", "推荐迁移工具"],
                overview)

    # Sheet 2: ECS 计算明细
    ws2 = wb_out.create_sheet()
    ecs_headers = ["Azure资源", "Azure规格", "环境", "操作系统",
                   "阿里云规格", "阿里云规格描述", "阿里云系统盘", "包月单价(元)"]
    write_sheet(ws2, "ECS计算明细", ecs_headers, vm_results)
    total_row = len(vm_results) + 2
    ws2.cell(row=total_row, column=7, value="ECS 合计:").font = Font(bold=True)
    ws2.cell(row=total_row, column=8).value = f"=SUM(H2:H{total_row - 1})"
    ws2.cell(row=total_row, column=8).number_format = MONEY_FMT
    ws2.cell(row=total_row, column=8).font = Font(bold=True)

    # Sheet 3: 数据库明细
    ws3 = wb_out.create_sheet()
    db_all = []
    for r in db_results:
        db_all.append({
            "Azure资源": r["Azure资源"], "Azure规格": r["Azure规格"], "环境": r["环境"],
            "阿里云产品": r["阿里云产品"], "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"], "包月单价(元)": r["包月单价(元)"],
        })
    for r in redis_results:
        db_all.append({
            "Azure资源": r["Azure资源"], "Azure规格": r["Azure规格"], "环境": r["环境"],
            "阿里云产品": r["阿里云产品"], "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"], "包月单价(元)": r["包月单价(元)"],
        })
    for r in cosmosdb_results:
        db_all.append({
            "Azure资源": r["Azure资源"], "Azure规格": r["Azure规格"], "环境": r["环境"],
            "阿里云产品": r["阿里云产品"], "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"], "包月单价(元)": r["包月单价(元)"],
        })
    db_headers = ["Azure资源", "Azure规格", "环境", "阿里云产品",
                  "阿里云规格", "阿里云规格描述", "包月单价(元)"]
    write_sheet(ws3, "数据库明细", db_headers, db_all)
    total_row = len(db_all) + 2
    ws3.cell(row=total_row, column=6, value="数据库合计:").font = Font(bold=True)
    ws3.cell(row=total_row, column=7).value = f"=SUM(G2:G{total_row - 1})"
    ws3.cell(row=total_row, column=7).number_format = MONEY_FMT
    ws3.cell(row=total_row, column=7).font = Font(bold=True)

    # Sheet 4: 网络明细
    ws4 = wb_out.create_sheet()
    net_headers = ["Azure资源", "Azure规格", "阿里云产品",
                   "阿里云规格", "阿里云规格描述", "包月单价(元)"]
    write_sheet(ws4, "网络明细", net_headers, network_results)
    total_row = len(network_results) + 2
    ws4.cell(row=total_row, column=5, value="网络合计:").font = Font(bold=True)
    ws4.cell(row=total_row, column=6).value = f"=SUM(F2:F{total_row - 1})"
    ws4.cell(row=total_row, column=6).number_format = MONEY_FMT
    ws4.cell(row=total_row, column=6).font = Font(bold=True)

    # Sheet 5: 其他服务
    ws5 = wb_out.create_sheet()
    other_headers = ["Azure产品", "Azure资源", "阿里云产品", "分类", "备注"]
    write_sheet(ws5, "其他服务", other_headers, other_results)

    # Sheet 6: 成本汇总
    ws6 = wb_out.create_sheet()
    ecs_total = sum(r.get("包月单价(元)", 0) for r in vm_results if isinstance(r.get("包月单价(元)", 0), (int, float)))
    db_total = sum(r.get("包月单价(元)", 0) for r in db_results if isinstance(r.get("包月单价(元)", 0), (int, float)))
    redis_total = sum(r.get("包月单价(元)", 0) for r in redis_results if isinstance(r.get("包月单价(元)", 0), (int, float)))
    mongo_total = sum(r.get("包月单价(元)", 0) for r in cosmosdb_results if isinstance(r.get("包月单价(元)", 0), (int, float)))
    net_total = sum(r.get("包月单价(元)", 0) for r in network_results if isinstance(r.get("包月单价(元)", 0), (int, float)))
    database_total = db_total + redis_total + mongo_total
    monthly_total = ecs_total + database_total + net_total

    summary = [
        {"分类": "ECS 计算", "数量": f"{len(vm_results)} 台", "月度费用(元)": round(ecs_total, 2)},
        {"分类": "RDS 数据库", "数量": f"{len(db_results)} 个", "月度费用(元)": round(db_total, 2)},
        {"分类": "Redis/Tair 缓存", "数量": f"{len(redis_results)} 个", "月度费用(元)": round(redis_total, 2)},
        {"分类": "MongoDB", "数量": f"{len(cosmosdb_results)} 个", "月度费用(元)": round(mongo_total, 2)},
        {"分类": "网络(LB+EIP+NAT)", "数量": f"{len(network_results)} 个", "月度费用(元)": round(net_total, 2)},
        {"分类": "", "数量": "", "月度费用(元)": ""},
        {"分类": "月度合计", "数量": "", "月度费用(元)": round(monthly_total, 2)},
        {"分类": "包年预估(85折)", "数量": "", "月度费用(元)": round(monthly_total * 12 * 0.85, 2)},
        {"分类": "3年预估(5折)", "数量": "", "月度费用(元)": round(monthly_total * 36 * 0.5, 2)},
    ]
    summary_headers = ["分类", "数量", "月度费用(元)"]
    write_sheet(ws6, "成本汇总", summary_headers, summary)
    for row_idx in range(8, 11):
        for col in range(1, 4):
            ws6.cell(row=row_idx, column=col).font = Font(bold=True)

    wb_out.save(output_path)
    return monthly_total, summary


# ============================================================
# main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Azure 资源列表 → 阿里云映射与报价")
    parser.add_argument("input", help="Azure 资源列表 Excel 文件路径")
    parser.add_argument("--output", "-o", help="输出文件路径")
    parser.add_argument("--region", default="cn-shanghai", help="阿里云目标区域 (默认 cn-shanghai)")
    parser.add_argument("--pricing-url", default="http://localhost:5001", help="询价服务地址")
    parser.add_argument("--no-pricing", action="store_true", help="跳过实时询价")
    args = parser.parse_args()

    input_path = args.input
    output_path = args.output or input_path.replace(".xlsx", "_阿里云映射报价.xlsx")
    region_id = args.region
    pricing_url = args.pricing_url

    print(f"📂 输入文件: {input_path}")
    print(f"🎯 目标区域: {region_id}")
    print()

    # 读取输入
    wb = openpyxl.load_workbook(input_path, data_only=True)
    print(f"📋 Azure → 阿里云 产品映射处理...")
    print(f"   共 {len(wb.sheetnames)} 个 Sheet: {', '.join(wb.sheetnames)}")

    # 处理各类资源
    vm_results = process_vm(wb)
    print(f"  VM: {len(vm_results)} 台实例")

    db_results = process_db(wb)
    print(f"  Database: {len(db_results)} 个实例")

    redis_results = process_redis(wb)
    print(f"  Redis: {len(redis_results)} 个实例")

    cosmosdb_results = process_cosmosdb(wb)
    print(f"  Cosmos DB/MongoDB: {len(cosmosdb_results)} 个实例")

    network_results = process_network(wb)
    print(f"  网络: {len(network_results)} 个资源")

    other_results = process_other_services(wb)
    print(f"  其他服务: {len(other_results)} 项")

    # 询价
    if not args.no_pricing:
        print()
        print(f"💰 实时询价 (服务地址: {pricing_url})...")
        try:
            r = requests.get(f"{pricing_url}/health", timeout=5)
            print("  询价服务已连接 ✓")
        except Exception:
            print("  ⚠️ 询价服务未运行, 使用参考价格")

        fill_ecs_prices(vm_results, pricing_url, region_id)
        fill_rds_prices(db_results, pricing_url, region_id)
        fill_redis_prices(redis_results, pricing_url, region_id)
        fill_mongodb_prices(cosmosdb_results, pricing_url, region_id)
        fill_eip_prices(network_results, pricing_url, region_id)
        fill_nat_prices(network_results, pricing_url, region_id)
    else:
        print("\n⏭️  跳过实时询价")

    # 生成输出
    print()
    print(f"📤 生成输出 Excel...")
    monthly_total, summary = generate_output(
        output_path, vm_results, db_results, redis_results,
        cosmosdb_results, network_results, other_results, region_id
    )

    print(f"\n输出文件已保存: {output_path}")
    print("=" * 50)
    print("  💰 月度成本汇总")
    print("=" * 50)
    for s in summary:
        cat = s["分类"]
        cost = s["月度费用(元)"]
        if not cat:
            print("─" * 50)
            continue
        if isinstance(cost, (int, float)):
            cnt = s.get("数量", "")
            print(f"  {cat:20s} ¥ {cost:>12,.2f}  {cnt}")
        else:
            print(f"  {cat:20s}")
    print("=" * 50)


if __name__ == "__main__":
    main()
