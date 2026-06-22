#!/usr/bin/env python3
"""华为云资源列表 → 阿里云映射与报价工具"""

import argparse
import math
import re
import sys
from collections import Counter

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 华为云 ECS → 阿里云 ECS 规格映射
# ============================================================

# 华为云规格族 → 阿里云规格族前缀
HW_TO_ALI_FAMILY = {
    # 通用型 (1:4)
    "s3": "ecs.g6", "s6": "ecs.g7", "s6e": "ecs.g7",
    "s7": "ecs.g7", "s7n": "ecs.g8i",
    "g6": "ecs.g7", "g7": "ecs.g7", "g7ne": "ecs.g7ne",
    "sn3": "ecs.g6",
    # 计算型 (1:2)
    "c3": "ecs.c7", "c3ne": "ecs.c7", "c6": "ecs.c7", "c6s": "ecs.c7",
    "c7": "ecs.c7", "c7n": "ecs.c8i", "c7t": "ecs.c7",
    "c9": "ecs.c8y", "ac7": "ecs.c8y", "hc2": "ecs.hfc7",
    # 内存型 (1:8)
    "m3": "ecs.r7", "m3ne": "ecs.r7", "m6": "ecs.r7", "m6e": "ecs.r7",
    "m7": "ecs.r7", "m7n": "ecs.r8i", "m7ne": "ecs.r7",
    # 超大内存型
    "e3": "ecs.r7", "e3ne": "ecs.r7",
    "x1": "ecs.re7", "x1e": "ecs.re7",
    # ARM 鲲鹏/倚天型
    "kc1": "ecs.c8y", "km1": "ecs.g8y", "kr1": "ecs.r8y", "ka1": "ecs.g8y",
    # 本地盘/存储型
    "d6": "ecs.d3s", "d7": "ecs.d3s", "i3": "ecs.i3", "i3l": "ecs.i4", "ir3": "ecs.i3",
    # GPU/异构计算型
    "p2s": "ecs.gn6v", "p2vs": "ecs.gn6v", "p3": "ecs.gn7e",
    "pi2": "ecs.gn7i", "g6v": "ecs.vgn7i",
    # 突发型/共享型
    "t6": "ecs.t6", "t6e": "ecs.t6",
}

# 阿里云 ECS 规格尺寸映射 (vCPU → 尺寸名)
ALI_SIZE_MAP = {
    1: "large",       # 部分型号 1C 用 large
    2: "large",
    4: "xlarge",
    8: "2xlarge",
    12: "3xlarge",
    16: "4xlarge",
    24: "6xlarge",
    32: "8xlarge",
    48: "12xlarge",
    60: "16xlarge",
    64: "16xlarge",
    96: "24xlarge",
    104: "26xlarge",
    128: "32xlarge",
    208: "52xlarge",
}


def parse_hw_ecs_spec(spec_str: str) -> dict:
    """解析华为云 ECS 规格字符串, 如 '4vCPUs | 8GB | c7.xlarge.2'"""
    parts = [p.strip() for p in spec_str.split("|")]
    result = {"raw": spec_str, "vcpu": 0, "mem_gb": 0, "hw_family": "", "hw_spec": ""}
    if len(parts) >= 3:
        # vCPU
        m = re.search(r"(\d+)\s*vCPU", parts[0])
        if m:
            result["vcpu"] = int(m.group(1))
        # Memory
        m = re.search(r"(\d+)\s*GB", parts[1])
        if m:
            result["mem_gb"] = int(m.group(1))
        # 规格名
        result["hw_spec"] = parts[2].strip()
        # 提取华为规格族 (e.g., c7t, c6s, e3)
        m = re.match(r"([a-z]+\d+[a-z]*)", result["hw_spec"])
        if m:
            result["hw_family"] = m.group(1)
    return result


def map_ecs_spec(hw: dict) -> dict:
    """将华为云 ECS 规格映射到阿里云 ECS 规格"""
    vcpu = hw["vcpu"]
    mem_gb = hw["mem_gb"]
    hw_family = hw["hw_family"]

    # 确定阿里云规格族
    ali_family = HW_TO_ALI_FAMILY.get(hw_family, "")
    if not ali_family:
        # 根据内存比推断
        ratio = mem_gb / vcpu if vcpu > 0 else 2
        if ratio <= 2:
            ali_family = "ecs.c7"
        elif ratio <= 4:
            ali_family = "ecs.g7"
        else:
            ali_family = "ecs.r7"

    # s6 共享型仅支持 1:1 和 1:2 内存比, 超出则升级到 g7/r7
    if ali_family == "ecs.s6" and vcpu > 0:
        ratio = mem_gb / vcpu
        if ratio > 2:
            ali_family = "ecs.g7" if ratio <= 4 else "ecs.r7"

    # 确定尺寸
    size = ALI_SIZE_MAP.get(vcpu)
    if not size:
        # 找最接近的
        for k in sorted(ALI_SIZE_MAP.keys()):
            if k >= vcpu:
                size = ALI_SIZE_MAP[k]
                vcpu = k  # 向上取整
                break
        if not size:
            size = "32xlarge"
            vcpu = 128

    ali_spec = f"{ali_family}.{size}"

    # 超大规格标注 (可能在部分区域不可用)
    note = ""
    if vcpu >= 104:
        note = " [超大规格,需确认可用性]"

    # 根据内存比确定描述
    ratio = mem_gb / vcpu if vcpu > 0 else 2
    if ratio <= 2:
        type_desc = "计算优化型"
    elif ratio <= 4:
        type_desc = "通用型"
    elif ratio <= 8:
        type_desc = "内存优化型"
    else:
        type_desc = "大内存型"

    return {
        "aliyun_spec": ali_spec,
        "aliyun_desc": f"{type_desc} {vcpu}C{mem_gb}G{note}",
        "vcpu": vcpu,
        "mem_gb": mem_gb,
    }


# ============================================================
# 华为云 EVS → 阿里云 ESSD 映射
# ============================================================

def parse_evs_spec(spec_str: str) -> dict:
    """解析 EVS 磁盘规格, 如 '通用型SSD | 300GB'"""
    parts = [p.strip() for p in spec_str.split("|")]
    result = {"type": "", "size_gb": 0, "aliyun_type": "cloud_essd", "aliyun_pl": "PL0"}
    if len(parts) >= 2:
        result["type"] = parts[0].strip()
        m = re.search(r"(\d+)\s*GB", parts[1])
        if m:
            result["size_gb"] = int(m.group(1))
    # 映射磁盘级别
    if "超高IO" in result["type"]:
        result["aliyun_pl"] = "PL1" if result["size_gb"] >= 20 else "PL0"
    else:
        result["aliyun_pl"] = "PL0"
    return result


# ============================================================
# 环境推断 (prod/beta/test)
# ============================================================

def infer_env(name: str, project: str = "") -> str:
    """从名称和项目推断环境: prod/beta/test"""
    text = f"{name} {project}".lower()
    if "prod" in text:
        return "prod"
    elif "beta" in text or "demo" in text or "gray" in text:
        return "beta"
    elif "test" in text or "dev" in text:
        return "test"
    return "prod"  # 默认 prod


# ============================================================
# 主处理函数
# ============================================================

def process_ecs(wb) -> list:
    """处理 ECS 云服务器"""
    ws = wb["弹性云服务器 ECS.云服务器"]
    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        region = str(row[2] or "")
        project = str(row[3] or "")
        spec_str = str(row[7] or "")
        image = str(row[8] or "")

        hw = parse_hw_ecs_spec(spec_str)
        ali = map_ecs_spec(hw)
        env = infer_env(name, project)

        # 确定系统盘 (默认 80GB ESSD PL0)
        sys_disk = 80
        sys_disk_desc = f"ESSD PL0 {max(sys_disk, 20)}GB"

        results.append({
            "华为云资源": name,
            "华为云规格": spec_str,
            "环境": env,
            "区域": region,
            "操作系统": image[:30] if image else "",
            "阿里云产品": "ECS 云服务器",
            "阿里云规格": ali["aliyun_spec"],
            "阿里云规格描述": ali["aliyun_desc"],
            "阿里云系统盘": sys_disk_desc,
            "阿里云数据盘": "",  # 后续从 EVS 关联
            "包月单价(元)": 0,
        })
    return results


def process_evs(wb) -> tuple:
    """处理 EVS 云硬盘, 返回 (系统盘统计, 数据盘列表)"""
    ws = wb["云硬盘 EVS.磁盘"]
    sys_disks = []
    data_disks = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        spec_str = str(row[5] or "")
        disk_attr = str(row[6] or "")  # 系统盘/数据盘
        project = str(row[7] or "")

        evs = parse_evs_spec(spec_str)
        env = infer_env(name, project)

        entry = {
            "name": name,
            "hw_type": evs["type"],
            "size_gb": evs["size_gb"],
            "disk_attr": disk_attr,
            "aliyun_pl": evs["aliyun_pl"],
            "aliyun_desc": f"ESSD {evs['aliyun_pl']} {evs['size_gb']}GB",
            "env": env,
        }
        if "系统盘" in disk_attr:
            sys_disks.append(entry)
        else:
            data_disks.append(entry)
    return sys_disks, data_disks


# 华为云 RDS 规格 → 阿里云 RDS 规格码映射 (按 vCPU/内存)
HW_RDS_SPEC_MAP = {
    # (vCPU, mem_gb): (ali_mysql_spec, ali_pg_spec)
    (1, 2):    ("rds.mysql.s1.small",    "pg.n2.small.1"),
    (2, 4):    ("rds.mysql.s2.large",    "pg.n2.small.2c"),
    (2, 8):    ("rds.mysql.s2.xlarge",   "pg.n2.medium.2c"),
    (4, 8):    ("rds.mysql.s3.small",    "pg.n4.medium.1"),
    (4, 16):   ("rds.mysql.s3.large",    "pg.n2.medium.2c"),
    (8, 16):   ("rds.mysql.c1.medium",   "pg.n4.large.1"),
    (8, 32):   ("rds.mysql.m1.medium",   "pg.n2.large.2c"),
    (16, 32):  ("rds.mysql.c1.xlarge",   "pg.n4.xlarge.1"),
    (16, 64):  ("rds.mysql.m1.xlarge",   "pg.n2.xlarge.2c"),
    (32, 128): ("rds.mysql.m1.2xlarge",  "pg.n4.2xlarge.1"),
    (64, 256): ("rds.mysql.m1.4xlarge",  "pg.n4.4xlarge.1"),
}


def _parse_rds_vcpu_mem(spec_str: str) -> tuple:
    """尝试从规格字符串解析 vCPU 和内存, 返回 (vcpu, mem_gb) 或 (0, 0)"""
    vcpu, mem = 0, 0
    m = re.search(r'(\d+)\s*vCPU', spec_str, re.IGNORECASE)
    if m:
        vcpu = int(m.group(1))
    m = re.search(r'(\d+)\s*GB', spec_str, re.IGNORECASE)
    if m:
        mem = int(m.group(1))
    # 也支持 "4C8G" 格式
    if not vcpu:
        m = re.search(r'(\d+)C(\d+)G', spec_str)
        if m:
            vcpu, mem = int(m.group(1)), int(m.group(2))
    return vcpu, mem


def _lookup_rds_spec(vcpu: int, mem_gb: int, engine: str) -> str:
    """根据 vCPU/内存 查找最佳阿里云 RDS 规格码"""
    is_pg = "pg" in engine.lower() or "postgres" in engine.lower()
    key = (vcpu, mem_gb)
    if key in HW_RDS_SPEC_MAP:
        mysql_spec, pg_spec = HW_RDS_SPEC_MAP[key]
        return pg_spec if is_pg else mysql_spec
    # 找最接近的 (优先内存匹配)
    best_key, best_dist = None, float('inf')
    for k in HW_RDS_SPEC_MAP:
        dist = abs(k[0] - vcpu) + abs(k[1] - mem_gb) * 0.5
        if dist < best_dist:
            best_dist = dist
            best_key = k
    if best_key:
        mysql_spec, pg_spec = HW_RDS_SPEC_MAP[best_key]
        return pg_spec if is_pg else mysql_spec
    return "rds.mysql.m1.medium"  # fallback 8C32G


def process_rds(wb) -> list:
    """处理 RDS 云数据库 (优先从规格字段精确映射，回退到环境推断)"""
    ws = wb["云数据库 RDS.实例"]
    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        project = str(row[3] or "")
        spec_str = str(row[7] or "") if len(row) > 7 else ""
        env = infer_env(name, project)

        # 从名称推断引擎
        engine = "MySQL"
        engine_ver = "5.7"
        name_lower = name.lower()
        if "pg" in name_lower or "postgres" in name_lower:
            engine = "PostgreSQL"
            engine_ver = "15.0"
        elif "sqlserver" in name_lower or "mssql" in name_lower:
            engine = "SQLServer"
            engine_ver = "2019_ent"
        elif "mariadb" in name_lower:
            engine = "MariaDB"
            engine_ver = "10.3"
        elif "mysql80" in name_lower or "mysql8" in name_lower:
            engine_ver = "8.0"

        # 尝试从规格字段或名称解析 vCPU/内存
        vcpu, mem_gb = _parse_rds_vcpu_mem(spec_str)
        if not vcpu:
            vcpu, mem_gb = _parse_rds_vcpu_mem(name)
        source = "精确"

        if vcpu and mem_gb:
            ali_spec = _lookup_rds_spec(vcpu, mem_gb, engine)
            ali_desc = f"{engine} {vcpu}C{mem_gb}G 高可用版"
            storage = max(100, mem_gb * 10)  # 存储估算
        else:
            # 环境推断回退
            source = "推断"
            if env == "prod":
                ali_spec = "rds.mysql.m1.medium" if engine == "MySQL" else "pg.n2.large.2c"
                ali_desc = f"{engine} 8C32G 高可用版"
                storage = 200
            else:
                ali_spec = "rds.mysql.s2.large" if engine == "MySQL" else "pg.n2.small.2c"
                ali_desc = f"{engine} 2C4G 高可用版"
                storage = 100

        # 只读实例用较小规格
        if "只读" in name:
            if vcpu and mem_gb:
                # 只读实例降一档
                lower_vcpu = max(1, vcpu // 2)
                lower_mem = max(2, mem_gb // 2)
                ali_spec = _lookup_rds_spec(lower_vcpu, lower_mem, engine)
                ali_desc = f"{engine} {lower_vcpu}C{lower_mem}G 只读实例"
            else:
                ali_spec = "rds.mysql.s2.large"
                ali_desc = f"{engine} 2C4G 只读实例"

        hw_spec_desc = f"{vcpu}C{mem_gb}G ({source})" if vcpu else f"推断: {env}"
        confirm_tag = "" if source == "精确" else " [需确认]"

        results.append({
            "华为云资源": name,
            "华为云规格": hw_spec_desc,
            "环境": env,
            "阿里云产品": f"RDS {engine}",
            "阿里云规格": ali_spec,
            "阿里云规格描述": f"{ali_desc} (存储{storage}GB){confirm_tag}",
            "引擎": engine,
            "引擎版本": engine_ver,
            "存储(GB)": storage,
            "包月单价(元)": 0,
        })
    return results


# 华为云 DCS Redis 容量 → 阿里云 Redis 规格码映射
HW_REDIS_SPEC_MAP = {
    # mem_mb: (ali_spec, capacity_mb)
    256:   ("redis.amber.master.small.multithread", 256),
    1024:  ("redis.amber.master.small.multithread", 1024),
    2048:  ("redis.amber.master.mid.multithread", 2048),
    4096:  ("redis.amber.master.stand.multithread", 4096),
    8192:  ("redis.amber.master.large.multithread", 8192),
    16384: ("redis.amber.master.2xlarge.multithread", 16384),
    32768: ("redis.amber.master.4xlarge.multithread", 32768),
    65536: ("redis.amber.master.8xlarge.multithread", 65536),
}


def _parse_redis_capacity(spec_str: str, name: str) -> int:
    """从规格字段或名称解析 Redis 容量(MB)，返回 0 表示未解析"""
    text = f"{spec_str} {name}"
    # 匹配 GB
    m = re.search(r'(\d+)\s*GB', text, re.IGNORECASE)
    if m:
        return int(m.group(1)) * 1024
    # 匹配 MB
    m = re.search(r'(\d+)\s*MB', text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return 0


def _lookup_redis_spec(capacity_mb: int) -> tuple:
    """根据容量查找最接近的 Redis 规格, 返回 (ali_spec, capacity_mb)"""
    if capacity_mb in HW_REDIS_SPEC_MAP:
        return HW_REDIS_SPEC_MAP[capacity_mb]
    # 向上取整到最接近的档位
    for cap in sorted(HW_REDIS_SPEC_MAP.keys()):
        if cap >= capacity_mb:
            return HW_REDIS_SPEC_MAP[cap]
    # 超过64GB返回最大
    return HW_REDIS_SPEC_MAP[65536]


def process_redis(wb) -> list:
    """处理 Redis 缓存 (优先从规格字段精确映射，回退到环境推断)"""
    ws = wb["分布式缓存服务 DCS.Redis实例"]
    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        project = str(row[3] or "")
        spec_str = str(row[7] or "") if len(row) > 7 else ""
        env = infer_env(name, project)

        # 尝试解析容量
        capacity_mb = _parse_redis_capacity(spec_str, name)
        source = "精确"

        if capacity_mb:
            ali_spec, ali_cap = _lookup_redis_spec(capacity_mb)
            ali_desc = f"云原生版 {ali_cap // 1024}GB 标准版-双副本" if ali_cap >= 1024 else f"云原生版 {ali_cap}MB 标准版-双副本"
        else:
            # 环境推断回退
            source = "推断"
            if env == "prod":
                ali_spec = "redis.amber.master.2xlarge.multithread"
                ali_desc = "云原生版 16GB 标准版-双副本"
                ali_cap = 16384
            else:
                ali_spec = "redis.amber.master.stand.multithread"
                ali_desc = "云原生版 4GB 标准版-双副本"
                ali_cap = 4096

        hw_spec_desc = f"{capacity_mb // 1024}GB ({source})" if capacity_mb >= 1024 else (f"{capacity_mb}MB ({source})" if capacity_mb else f"推断: {env}")
        confirm_tag = "" if source == "精确" else " [需确认]"

        results.append({
            "华为云资源": name,
            "华为云规格": hw_spec_desc,
            "环境": env,
            "阿里云产品": "Redis/Tair 云原生版",
            "阿里云规格": ali_spec,
            "阿里云规格描述": f"{ali_desc}{confirm_tag}",
            "容量(MB)": ali_cap,
            "包月单价(元)": 0,
        })
    return results


def process_elb(wb) -> list:
    """处理 ELB 负载均衡器 (按量付费, 仅实例费)"""
    ws = wb["弹性负载均衡 ELB.负载均衡器"]
    NLB_HOURLY = 0.147  # NLB/ALB 实例费 (元/h)
    HOURS_PER_MONTH = 730
    monthly = round(NLB_HOURLY * HOURS_PER_MONTH, 2)

    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        project = str(row[3] or "")

        # 推断类型
        name_lower = name.lower()
        if "wan" in name_lower or "公网" in name_lower or "external" in name_lower:
            lb_type = "公网"
            aliyun_prod = "ALB 应用负载均衡"
        else:
            lb_type = "私网"
            aliyun_prod = "ALB 应用负载均衡"

        results.append({
            "华为云资源": f"ELB: {name}",
            "华为云规格": lb_type,
            "环境": infer_env(name, project),
            "阿里云产品": aliyun_prod,
            "阿里云规格": "按量付费",
            "阿里云规格描述": f"ALB 标准版 按量付费(实例费¥{NLB_HOURLY}/h, 不含LCU费)",
            "包月单价(元)": monthly,
        })
    return results


# ============================================================
# 华为云 GaussDB → 阿里云 PolarDB 规格映射
# ============================================================

HW_GAUSSDB_SPEC_MAP = {
    # (vCPU, mem_gb): (polar_mysql_spec, polar_pg_spec)
    (2, 8):    ("polar.mysql.g2.xlarge",    "polar.pg.g2.xlarge"),
    (4, 16):   ("polar.mysql.g4.xlarge",    "polar.pg.g4.xlarge"),
    (8, 32):   ("polar.mysql.g8.xlarge",    "polar.pg.g8.xlarge"),
    (8, 64):   ("polar.mysql.g8.2xlarge",   "polar.pg.g8.2xlarge"),
    (16, 64):  ("polar.mysql.g8.2xlarge",   "polar.pg.g8.2xlarge"),
    (16, 128): ("polar.mysql.g8.4xlarge",   "polar.pg.g8.4xlarge"),
    (32, 256): ("polar.mysql.g8.8xlarge",   "polar.pg.g8.8xlarge"),
    (64, 512): ("polar.mysql.g8.16xlarge",  "polar.pg.g8.16xlarge"),
}


def process_gaussdb(wb) -> list:
    """处理 GaussDB → PolarDB (优先从规格字段精确映射)"""
    results = []
    # 尝试多个可能的 Sheet 名
    sheet_names = [
        "云数据库 GaussDB.实例",
        "云数据库 GaussDB(for MySQL).实例",
        "GaussDB(for MySQL).实例",
        "GaussDB.实例",
    ]
    ws = None
    for sn in sheet_names:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        return results

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        project = str(row[3] or "")
        spec_str = str(row[7] or "") if len(row) > 7 else ""
        env = infer_env(name, project)

        # 推断引擎
        is_pg = "postgres" in name.lower() or "opengauss" in name.lower()
        engine = "PostgreSQL" if is_pg else "MySQL"

        # 解析 vCPU/内存
        vcpu, mem_gb = _parse_rds_vcpu_mem(spec_str)
        if not vcpu:
            vcpu, mem_gb = _parse_rds_vcpu_mem(name)

        if vcpu and mem_gb:
            key = (vcpu, mem_gb)
            if key in HW_GAUSSDB_SPEC_MAP:
                mysql_spec, pg_spec = HW_GAUSSDB_SPEC_MAP[key]
            else:
                # 找最接近的
                best_key = min(HW_GAUSSDB_SPEC_MAP.keys(),
                               key=lambda k: abs(k[0] - vcpu) + abs(k[1] - mem_gb) * 0.5)
                mysql_spec, pg_spec = HW_GAUSSDB_SPEC_MAP[best_key]
            ali_spec = pg_spec if is_pg else mysql_spec
            ali_desc = f"PolarDB {engine} {vcpu}C{mem_gb}G"
            source = "精确"
        else:
            source = "推断"
            if env == "prod":
                ali_spec = "polar.pg.g8.xlarge" if is_pg else "polar.mysql.g8.xlarge"
                ali_desc = f"PolarDB {engine} 8C32G"
            else:
                ali_spec = "polar.pg.g4.xlarge" if is_pg else "polar.mysql.g4.xlarge"
                ali_desc = f"PolarDB {engine} 4C16G"

        hw_spec_desc = f"{vcpu}C{mem_gb}G ({source})" if vcpu else f"推断: {env}"
        confirm_tag = "" if source == "精确" else " [需确认]"

        results.append({
            "华为云资源": name,
            "华为云规格": hw_spec_desc,
            "环境": env,
            "阿里云产品": f"PolarDB {engine}",
            "阿里云规格": ali_spec,
            "阿里云规格描述": f"{ali_desc}{confirm_tag}",
            "引擎": engine,
            "包月单价(元)": 0,
        })
    return results


# ============================================================
# 华为云 DDS MongoDB → 阿里云 MongoDB 规格映射
# ============================================================

HW_MONGODB_SPEC_MAP = {
    # (vCPU, mem_gb): ali_spec
    (2, 4):   "mdb.shard.2x.large.d",
    (4, 8):   "mdb.shard.2x.xlarge.d",
    (8, 16):  "mdb.shard.2x.2xlarge.d",
    (16, 32): "mdb.shard.2x.4xlarge.d",
    (32, 64): "mdb.shard.2x.8xlarge.d",
}


def process_mongodb(wb) -> list:
    """处理 DDS MongoDB → 阿里云 MongoDB (优先从规格字段精确映射)"""
    results = []
    sheet_names = [
        "文档数据库服务 DDS.实例",
        "文档数据库 DDS.实例",
        "DDS.实例",
    ]
    ws = None
    for sn in sheet_names:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        return results

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        project = str(row[3] or "")
        spec_str = str(row[7] or "") if len(row) > 7 else ""
        env = infer_env(name, project)

        vcpu, mem_gb = _parse_rds_vcpu_mem(spec_str)
        if not vcpu:
            vcpu, mem_gb = _parse_rds_vcpu_mem(name)

        if vcpu and mem_gb:
            key = (vcpu, mem_gb)
            if key in HW_MONGODB_SPEC_MAP:
                ali_spec = HW_MONGODB_SPEC_MAP[key]
            else:
                best_key = min(HW_MONGODB_SPEC_MAP.keys(),
                               key=lambda k: abs(k[0] - vcpu) + abs(k[1] - mem_gb) * 0.5)
                ali_spec = HW_MONGODB_SPEC_MAP[best_key]
            ali_desc = f"MongoDB 独享型 {vcpu}C{mem_gb}G"
            source = "精确"
        else:
            source = "推断"
            if env == "prod":
                ali_spec = "mdb.shard.2x.xlarge.d"
                ali_desc = "MongoDB 独享型 4C8G"
            else:
                ali_spec = "mdb.shard.2x.large.d"
                ali_desc = "MongoDB 独享型 2C4G"

        hw_spec_desc = f"{vcpu}C{mem_gb}G ({source})" if vcpu else f"推断: {env}"
        confirm_tag = "" if source == "精确" else " [需确认]"

        results.append({
            "华为云资源": name,
            "华为云规格": hw_spec_desc,
            "环境": env,
            "阿里云产品": "云数据库 MongoDB 版",
            "阿里云规格": ali_spec,
            "阿里云规格描述": f"{ali_desc}{confirm_tag}",
            "磁盘(GB)": 40,
            "包月单价(元)": 0,
        })
    return results


def process_eip_bandwidth(wb) -> list:
    """处理弹性公网IP
    以 EIP Sheet (84条) 为主体, 从 VPC.带宽 Sheet 查找每个 EIP 的实际带宽值.
    每个 EIP 映射为阿里云 EIP, 按其关联带宽询价.
    """
    # 先从带宽 Sheet 建立 带宽名→Mbps 映射
    ws_bw = wb["虚拟私有云 VPC.带宽"]
    bw_map = {}  # name → Mbps
    for row in ws_bw.iter_rows(min_row=2, max_row=ws_bw.max_row, values_only=True):
        if not row[0]:
            continue
        bw_name = str(row[1] or "")
        bw_mbps = int(row[4]) if row[4] else 5
        bw_map[bw_name] = bw_mbps

    # 以 EIP Sheet 为主体
    ws_eip = wb["虚拟私有云 VPC.弹性公网IP"]
    results = []
    for row in ws_eip.iter_rows(min_row=2, max_row=ws_eip.max_row, values_only=True):
        if not row[0]:
            continue
        ip_addr = str(row[1] or "")
        bw_name = str(row[5] or "")
        project = str(row[6] or "")

        # 从带宽 Sheet 查找 Mbps
        mbps = bw_map.get(bw_name, 5)

        results.append({
            "华为云资源": f"EIP: {ip_addr}",
            "华为云规格": f"{mbps} Mbps (带宽: {bw_name})",
            "环境": infer_env(bw_name, project),
            "阿里云产品": "弹性公网IP (EIP)",
            "阿里云规格": f"EIP {mbps}Mbps 按固定带宽",
            "阿里云规格描述": f"BGP多线 {mbps}Mbps 按固定带宽计费",
            "带宽(Mbps)": mbps,
            "包月单价(元)": 0,
        })
    return results


def process_nat(wb) -> list:
    """处理 NAT 网关"""
    ws = wb["NAT网关 NAT.公网NAT网关"]
    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        name = str(row[1] or "")
        project = str(row[3] or "")
        results.append({
            "华为云资源": f"NAT: {name}",
            "华为云规格": "公网NAT网关",
            "环境": infer_env(name, project),
            "阿里云产品": "NAT 网关",
            "阿里云规格": "增强型小规格",
            "阿里云规格描述": "公网NAT网关 增强型小规格(连接数10万, 新建5000/s)",
            "包月单价(元)": 0,
        })
    return results


def process_other_services(wb) -> list:
    """处理其他服务 (仅产品映射, 不询价)"""
    MAPPING = {
        "弹性文件服务.SFS Turbo": ("NAS 极速型", "文件存储"),
        "对象存储服务 OBS.桶": ("OSS 对象存储", "对象存储"),
        "云搜索服务 CSS.集群": ("Elasticsearch", "搜索引擎"),
        "分布式消息服务 DMS.Kafka实例": ("消息队列 Kafka 版", "消息队列"),
        "云容器引擎 CCE.集群": ("ACK 容器服务", "容器"),
        "CDN.域名": ("CDN", "CDN"),
        "虚拟专用网络 VPN.经典型VPN网关": ("VPN 网关", "网络"),
        "云专线 DC.物理连接": ("高速通道", "网络"),
        "云备份 CBR.存储库": ("云备份 HBR", "备份"),
        "云防火墙 CFW.云防火墙实例": ("云防火墙", "安全"),
        "云解析服务 DNS.内网Zone": ("云解析 PrivateZone", "DNS"),
        "数据加密服务 DEW.密钥": ("KMS 密钥管理", "安全"),
        "企业主机安全 HSS.主机代理": ("云安全中心", "安全"),
        "SSL证书管理 SCM.证书": ("数字证书管理", "安全"),
        "消息通知服务 SMN.主题": ("消息服务 MNS", "消息"),
        "虚拟私有云 VPC.虚拟私有云": ("VPC 专有网络", "网络(免费)"),
        "虚拟私有云 VPC.安全组": ("安全组", "网络(免费)"),
        "云审计服务.追踪器": ("操作审计", "审计(免费)"),
        "云日志服务 LTS.日志流": ("SLS 日志服务", "日志"),
        "云连接 CC.云连接": ("云企业网 CEN", "网络"),
    }

    results = []
    for sheet_name, (ali_name, category) in MAPPING.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        count = ws.max_row - 1
        if count <= 0:
            continue
        # 收集条目
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            name = str(row[1] or "")
            results.append({
                "华为云产品": sheet_name.split(".")[0] if "." in sheet_name else sheet_name,
                "华为云资源": name,
                "阿里云产品": ali_name,
                "分类": category,
                "备注": "仅映射, 需根据实际用量报价" if "免费" not in category else "免费",
            })
    return results


# ============================================================
# 询价函数
# ============================================================

def batch_pricing(pricing_url: str, items: list, region_id: str) -> dict:
    """调用批量询价 API"""
    try:
        resp = requests.post(
            f"{pricing_url}/api/pricing/batch",
            json={"region_id": region_id, "period": 1, "items": items},
            timeout=120,
        )
        data = resp.json()
        return data.get("data", {})
    except Exception as e:
        print(f"  询价失败: {e}")
        return {}


def fill_ecs_prices(ecs_results: list, pricing_url: str, region_id: str):
    """ECS 批量询价"""
    # 按 (规格, 系统盘, 数据盘) 去重
    spec_set = {}
    for r in ecs_results:
        spec = r["阿里云规格"]
        sys_disk = r.get("阿里云系统盘", "ESSD PL0 80GB")
        data_disk_desc = r.get("阿里云数据盘", "")
        key = (spec, data_disk_desc)
        if key not in spec_set:
            # 解析数据盘
            data_disk_size = 0
            data_disk_pl = "PL1"
            if data_disk_desc:
                m = re.search(r"(\d+)\s*GB", data_disk_desc)
                if m:
                    data_disk_size = int(m.group(1))
                if "PL0" in data_disk_desc:
                    data_disk_pl = "PL0"

            item = {
                "product_code": "ecs",
                "instance_spec": spec,
                "disk_size": 80,  # 系统盘
                "count": 1,
            }
            if data_disk_size > 0:
                item["data_disk_size"] = data_disk_size
                item["data_disk_pl"] = data_disk_pl
            spec_set[key] = item

    items = list(spec_set.values())
    if not items:
        return

    print(f"  询价 ECS: {len(items)} 种配置...")
    result = batch_pricing(pricing_url, items, region_id)

    # 建立价格映射
    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    filled = 0
    for r in ecs_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]
            filled += 1

    # 对未填充的实例, 按 vCPU 估算参考价 (约 ¥50/vCPU/月)
    unfilled = [r for r in ecs_results if r["包月单价(元)"] == 0]
    for r in unfilled:
        desc = r.get("阿里云规格描述", "")
        m = re.search(r"(\d+)C", desc)
        if m:
            vcpu = int(m.group(1))
            r["包月单价(元)"] = round(vcpu * 50, 2)  # 参考估价
            r["阿里云规格描述"] += " [参考估价]"
            filled += 1

    print(f"    成功填充 {filled}/{len(ecs_results)} 台价格")


def fill_rds_prices(rds_results: list, pricing_url: str, region_id: str):
    """RDS 批量询价"""
    items = []
    spec_set = set()
    for r in rds_results:
        spec = r["阿里云规格"]
        if spec not in spec_set:
            spec_set.add(spec)
            items.append({
                "product_code": "rds",
                "instance_spec": spec,
                "engine": r.get("引擎", "MySQL"),
                "engine_version": r.get("引擎版本", "5.7"),
                "disk_size": r.get("存储(GB)", 200),
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 RDS: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in rds_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


def fill_redis_prices(redis_results: list, pricing_url: str, region_id: str):
    """Redis 批量询价"""
    items = []
    spec_set = set()
    for r in redis_results:
        spec = r["阿里云规格"]
        if spec not in spec_set:
            spec_set.add(spec)
            items.append({
                "product_code": "redis",
                "instance_spec": spec,
                "capacity": r.get("容量(MB)", 4096),
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 Redis: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in redis_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


def fill_eip_prices(eip_results: list, pricing_url: str, region_id: str):
    """EIP 批量询价 (按带宽档位去重)"""
    bw_set = set()
    items = []
    for r in eip_results:
        bw = r.get("带宽(Mbps)", 5)
        if bw not in bw_set:
            bw_set.add(bw)
            items.append({
                "product_code": "eip",
                "instance_spec": f"{bw}Mbps",
                "bandwidth": bw,
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 EIP: {len(items)} 种带宽...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            # key: "5Mbps"
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in eip_results:
        bw = r.get("带宽(Mbps)", 5)
        key = f"{bw}Mbps"
        if key in price_map:
            r["包月单价(元)"] = price_map[key]
        else:
            # fallback: 23元/Mbps (1-5), 80元/Mbps (6+)
            if bw <= 5:
                r["包月单价(元)"] = bw * 23
            else:
                r["包月单价(元)"] = 5 * 23 + (bw - 5) * 80


def fill_nat_prices(nat_results: list, pricing_url: str, region_id: str):
    """NAT 网关询价"""
    if not nat_results:
        return
    items = [{"product_code": "nat", "instance_spec": "Small", "count": len(nat_results)}]
    print(f"  询价 NAT: {len(nat_results)} 个...")
    result = batch_pricing(pricing_url, items, region_id)

    price = 220  # fallback
    for p in result.get("prices", []):
        if p.get("success"):
            price = p["trade_price"]

    for r in nat_results:
        r["包月单价(元)"] = price


def fill_gaussdb_prices(gaussdb_results: list, pricing_url: str, region_id: str):
    """GaussDB/PolarDB 批量询价"""
    items = []
    spec_set = set()
    for r in gaussdb_results:
        spec = r["阿里云规格"]
        if spec not in spec_set:
            spec_set.add(spec)
            items.append({
                "product_code": "polardb",
                "instance_spec": spec,
                "engine": r.get("引擎", "MySQL"),
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 PolarDB(GaussDB): {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in gaussdb_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


def fill_mongodb_prices(mongodb_results: list, pricing_url: str, region_id: str):
    """MongoDB 批量询价"""
    items = []
    spec_set = set()
    for r in mongodb_results:
        spec = r["阿里云规格"]
        if spec not in spec_set:
            spec_set.add(spec)
            items.append({
                "product_code": "mongodb",
                "instance_spec": spec,
                "disk_size": r.get("磁盘(GB)", 40),
                "count": 1,
            })
    if not items:
        return

    print(f"  询价 MongoDB: {len(items)} 种规格...")
    result = batch_pricing(pricing_url, items, region_id)

    price_map = {}
    for p in result.get("prices", []):
        if p.get("success"):
            price_map[p["instance_spec"]] = p["trade_price"]

    for r in mongodb_results:
        spec = r["阿里云规格"]
        if spec in price_map:
            r["包月单价(元)"] = price_map[spec]


# ============================================================
# Excel 输出
# ============================================================

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
MONEY_FMT = '#,##0.00'


def write_sheet(ws, title: str, headers: list, rows: list):
    """写入一个工作表"""
    ws.title = title
    # 表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # 数据
    for row_idx, row_data in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER
            if h in ("包月单价(元)", "月度小计(元)"):
                cell.number_format = MONEY_FMT

    # 自动列宽
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row_idx in range(2, min(len(rows) + 2, 20)):
            val = str(ws.cell(row=row_idx, column=col).value or "")
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 4


def generate_output(output_path: str, ecs_results, rds_results, redis_results,
                    gaussdb_results, mongodb_results,
                    elb_results, eip_results, nat_results, other_results,
                    data_disks, region_id):
    """生成输出 Excel"""
    wb_out = openpyxl.Workbook()

    # Sheet 1: 资源映射总览
    overview = []
    categories = [
        ("ECS 云服务器", len(ecs_results), "ECS 云服务器", "SMC 服务器迁移"),
        ("EVS 云硬盘(数据盘)", len(data_disks), "ESSD 云盘", "随ECS迁移"),
        ("RDS 数据库", len(rds_results), "RDS MySQL/PG", "DTS 数据迁移"),
        ("GaussDB/PolarDB", len(gaussdb_results), "PolarDB MySQL/PG", "DTS 数据迁移"),
        ("Redis 缓存", len(redis_results), "Redis/Tair 云原生版", "DTS/Redis-shake"),
        ("MongoDB", len(mongodb_results), "云数据库 MongoDB 版", "DTS 数据迁移"),
        ("ELB 负载均衡", len(elb_results), "ALB 应用负载均衡", "手动重建"),
        ("EIP 弹性公网IP", len(eip_results), "弹性公网IP (EIP)", "手动分配"),
        ("NAT 网关", len(nat_results), "NAT 网关", "手动重建"),
    ]
    # 统计其他服务
    other_cats = Counter()
    for o in other_results:
        key = (o.get("华为云产品", ""), o.get("阿里云产品", ""))
        other_cats[key] += 1
    for (hw, ali), cnt in other_cats.most_common():
        categories.append((hw, cnt, ali, "手动重建"))

    for i, (hw, cnt, ali, tool) in enumerate(categories, 1):
        overview.append({
            "序号": i,
            "华为云产品": hw,
            "数量": cnt,
            "阿里云产品": ali,
            "推荐迁移工具": tool,
        })

    ws1 = wb_out.active
    write_sheet(ws1, "资源映射总览",
                ["序号", "华为云产品", "数量", "阿里云产品", "推荐迁移工具"],
                overview)

    # Sheet 2: ECS 计算明细
    ws2 = wb_out.create_sheet()
    ecs_headers = ["华为云资源", "华为云规格", "环境", "操作系统",
                   "阿里云规格", "阿里云规格描述", "阿里云系统盘", "包月单价(元)"]
    write_sheet(ws2, "ECS计算明细", ecs_headers, ecs_results)
    # 合计行
    total_row = len(ecs_results) + 2
    ws2.cell(row=total_row, column=7, value="ECS 合计:").font = Font(bold=True)
    ws2.cell(row=total_row, column=8).value = \
        f"=SUM(H2:H{total_row - 1})"
    ws2.cell(row=total_row, column=8).number_format = MONEY_FMT
    ws2.cell(row=total_row, column=8).font = Font(bold=True)

    # Sheet 3: 数据库明细
    ws3 = wb_out.create_sheet()
    db_all = []
    for r in rds_results:
        db_all.append({
            "华为云资源": r["华为云资源"],
            "华为云规格": r["华为云规格"],
            "环境": r["环境"],
            "阿里云产品": r["阿里云产品"],
            "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"],
            "包月单价(元)": r["包月单价(元)"],
        })
    for r in gaussdb_results:
        db_all.append({
            "华为云资源": r["华为云资源"],
            "华为云规格": r["华为云规格"],
            "环境": r["环境"],
            "阿里云产品": r["阿里云产品"],
            "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"],
            "包月单价(元)": r["包月单价(元)"],
        })
    for r in redis_results:
        db_all.append({
            "华为云资源": r["华为云资源"],
            "华为云规格": r["华为云规格"],
            "环境": r["环境"],
            "阿里云产品": r["阿里云产品"],
            "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"],
            "包月单价(元)": r["包月单价(元)"],
        })
    for r in mongodb_results:
        db_all.append({
            "华为云资源": r["华为云资源"],
            "华为云规格": r["华为云规格"],
            "环境": r["环境"],
            "阿里云产品": r["阿里云产品"],
            "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"],
            "包月单价(元)": r["包月单价(元)"],
        })
    db_headers = ["华为云资源", "华为云规格", "环境", "阿里云产品",
                  "阿里云规格", "阿里云规格描述", "包月单价(元)"]
    write_sheet(ws3, "数据库明细", db_headers, db_all)
    total_row = len(db_all) + 2
    ws3.cell(row=total_row, column=6, value="数据库合计:").font = Font(bold=True)
    ws3.cell(row=total_row, column=7).value = f"=SUM(G2:G{total_row - 1})"
    ws3.cell(row=total_row, column=7).number_format = MONEY_FMT
    ws3.cell(row=total_row, column=7).font = Font(bold=True)

    # Sheet 4: 网络明细
    ws4 = wb_out.create_sheet()
    net_all = []
    for r in elb_results:
        net_all.append({
            "华为云资源": r["华为云资源"],
            "华为云规格": r["华为云规格"],
            "阿里云产品": r["阿里云产品"],
            "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"],
            "包月单价(元)": r["包月单价(元)"],
        })
    for r in eip_results:
        net_all.append({
            "华为云资源": r["华为云资源"],
            "华为云规格": r["华为云规格"],
            "阿里云产品": r["阿里云产品"],
            "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"],
            "包月单价(元)": r["包月单价(元)"],
        })
    for r in nat_results:
        net_all.append({
            "华为云资源": r["华为云资源"],
            "华为云规格": r["华为云规格"],
            "阿里云产品": r["阿里云产品"],
            "阿里云规格": r["阿里云规格"],
            "阿里云规格描述": r["阿里云规格描述"],
            "包月单价(元)": r["包月单价(元)"],
        })
    net_headers = ["华为云资源", "华为云规格", "阿里云产品",
                   "阿里云规格", "阿里云规格描述", "包月单价(元)"]
    write_sheet(ws4, "网络明细", net_headers, net_all)
    total_row = len(net_all) + 2
    ws4.cell(row=total_row, column=5, value="网络合计:").font = Font(bold=True)
    ws4.cell(row=total_row, column=6).value = f"=SUM(F2:F{total_row - 1})"
    ws4.cell(row=total_row, column=6).number_format = MONEY_FMT
    ws4.cell(row=total_row, column=6).font = Font(bold=True)

    # Sheet 5: 存储明细
    ws5 = wb_out.create_sheet()
    storage_rows = []
    for d in data_disks:
        price_per_gb = 0.35 if d["aliyun_pl"] == "PL0" else 0.50
        monthly = round(d["size_gb"] * price_per_gb, 2)
        storage_rows.append({
            "华为云资源": d["name"],
            "华为云规格": f"{d['hw_type']} {d['size_gb']}GB ({d['disk_attr']})",
            "阿里云产品": "ESSD 云盘",
            "阿里云规格": d["aliyun_desc"],
            "单价(元/GB/月)": price_per_gb,
            "包月单价(元)": monthly,
        })
    storage_headers = ["华为云资源", "华为云规格", "阿里云产品",
                       "阿里云规格", "单价(元/GB/月)", "包月单价(元)"]
    write_sheet(ws5, "存储明细", storage_headers, storage_rows)
    total_row = len(storage_rows) + 2
    ws5.cell(row=total_row, column=5, value="存储合计:").font = Font(bold=True)
    ws5.cell(row=total_row, column=6).value = f"=SUM(F2:F{total_row - 1})"
    ws5.cell(row=total_row, column=6).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=6).font = Font(bold=True)

    # Sheet 6: 其他服务
    ws6 = wb_out.create_sheet()
    other_headers = ["华为云产品", "华为云资源", "阿里云产品", "分类", "备注"]
    write_sheet(ws6, "其他服务", other_headers, other_results)

    # Sheet 7: 成本汇总
    ws7 = wb_out.create_sheet()
    ecs_total = sum(r.get("包月单价(元)", 0) for r in ecs_results)
    rds_total = sum(r.get("包月单价(元)", 0) for r in rds_results)
    gaussdb_total = sum(r.get("包月单价(元)", 0) for r in gaussdb_results)
    redis_total = sum(r.get("包月单价(元)", 0) for r in redis_results)
    mongodb_total = sum(r.get("包月单价(元)", 0) for r in mongodb_results)
    elb_total = sum(r.get("包月单价(元)", 0) for r in elb_results)
    eip_total = sum(r.get("包月单价(元)", 0) for r in eip_results)
    nat_total = sum(r.get("包月单价(元)", 0) for r in nat_results)
    storage_total = sum(r.get("包月单价(元)", 0) for r in storage_rows)
    db_total = rds_total + gaussdb_total + redis_total + mongodb_total
    network_total = elb_total + eip_total + nat_total
    monthly_total = ecs_total + db_total + network_total + storage_total

    summary = [
        {"分类": "ECS 计算", "数量": f"{len(ecs_results)} 台", "月度费用(元)": round(ecs_total, 2)},
        {"分类": "RDS 数据库", "数量": f"{len(rds_results)} 个", "月度费用(元)": round(rds_total, 2)},
        {"分类": "GaussDB/PolarDB", "数量": f"{len(gaussdb_results)} 个", "月度费用(元)": round(gaussdb_total, 2)},
        {"分类": "Redis/Tair 缓存", "数量": f"{len(redis_results)} 个", "月度费用(元)": round(redis_total, 2)},
        {"分类": "MongoDB", "数量": f"{len(mongodb_results)} 个", "月度费用(元)": round(mongodb_total, 2)},
        {"分类": "网络(ELB+EIP+NAT)", "数量": f"{len(elb_results)+len(eip_results)+len(nat_results)} 个",
         "月度费用(元)": round(network_total, 2)},
        {"分类": "存储(数据盘)", "数量": f"{len(storage_rows)} 块", "月度费用(元)": round(storage_total, 2)},
        {"分类": "", "数量": "", "月度费用(元)": ""},
        {"分类": "月度合计", "数量": "", "月度费用(元)": round(monthly_total, 2)},
        {"分类": "包年预估(85折)", "数量": "", "月度费用(元)": round(monthly_total * 12 * 0.85, 2)},
        {"分类": "3年预估(5折)", "数量": "", "月度费用(元)": round(monthly_total * 36 * 0.5, 2)},
    ]
    summary_headers = ["分类", "数量", "月度费用(元)"]
    write_sheet(ws7, "成本汇总", summary_headers, summary)
    # 加粗合计行
    for row_idx in range(8, 11):
        for col in range(1, 4):
            ws7.cell(row=row_idx, column=col).font = Font(bold=True)

    wb_out.save(output_path)
    return monthly_total, summary


# ============================================================
# main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="华为云资源列表 → 阿里云映射与报价")
    parser.add_argument("input", help="华为云资源列表 Excel 文件路径")
    parser.add_argument("--output", "-o", help="输出文件路径")
    parser.add_argument("--region", default="cn-shenzhen", help="阿里云目标区域 (默认 cn-shenzhen)")
    parser.add_argument("--pricing-url", default="http://localhost:5001", help="询价服务地址")
    args = parser.parse_args()

    input_path = args.input
    output_path = args.output or input_path.replace(".xlsx", "_阿里云映射报价.xlsx")
    region_id = args.region
    pricing_url = args.pricing_url

    print(f"📂 输入文件: {input_path}")
    print(f"🎯 目标区域: {region_id}")
    print()

    # 读取输入
    wb = openpyxl.load_workbook(input_path, data_only=True)
    print(f"📋 产品映射处理...")

    # 处理各类资源
    ecs_results = process_ecs(wb)
    print(f"  ECS: {len(ecs_results)} 台实例")

    sys_disks, data_disks = process_evs(wb)
    print(f"  EVS: {len(sys_disks)} 系统盘 + {len(data_disks)} 数据盘")

    rds_results = process_rds(wb)
    print(f"  RDS: {len(rds_results)} 个实例")

    gaussdb_results = process_gaussdb(wb)
    print(f"  GaussDB: {len(gaussdb_results)} 个实例")

    redis_results = process_redis(wb)
    print(f"  Redis: {len(redis_results)} 个实例")

    mongodb_results = process_mongodb(wb)
    print(f"  MongoDB: {len(mongodb_results)} 个实例")

    elb_results = process_elb(wb)
    print(f"  ELB: {len(elb_results)} 个负载均衡")

    eip_results = process_eip_bandwidth(wb)
    print(f"  EIP+带宽: {len(eip_results)} 条(按带宽去重)")

    nat_results = process_nat(wb)
    print(f"  NAT: {len(nat_results)} 个")

    other_results = process_other_services(wb)
    print(f"  其他服务: {len(other_results)} 项")

    # 检查询价服务
    print()
    print(f"💰 实时询价 (服务地址: {pricing_url})...")
    try:
        r = requests.get(f"{pricing_url}/health", timeout=5)
        print("  询价服务已连接 ✓")
    except Exception:
        print("  ⚠️ 询价服务未运行, 使用参考价格")

    # 批量询价
    fill_ecs_prices(ecs_results, pricing_url, region_id)
    fill_rds_prices(rds_results, pricing_url, region_id)
    fill_gaussdb_prices(gaussdb_results, pricing_url, region_id)
    fill_redis_prices(redis_results, pricing_url, region_id)
    fill_mongodb_prices(mongodb_results, pricing_url, region_id)
    fill_eip_prices(eip_results, pricing_url, region_id)
    fill_nat_prices(nat_results, pricing_url, region_id)

    # 生成输出
    print()
    print(f"📤 生成输出 Excel...")
    monthly_total, summary = generate_output(
        output_path, ecs_results, rds_results, redis_results,
        gaussdb_results, mongodb_results,
        elb_results, eip_results, nat_results, other_results,
        data_disks, region_id
    )

    print(f"\n输出文件已保存: {output_path}")
    print("=" * 50)
    print("  💰 月度成本汇总")
    print("=" * 50)
    for s in summary:
        cat = s["分类"]
        cost = s["月度费用(元)"]
        if not cat:
            print("─" * 50)
            continue
        if isinstance(cost, (int, float)):
            cnt = s.get("数量", "")
            print(f"  {cat:20s} ¥ {cost:>12,.2f}  {cnt}")
        else:
            print(f"  {cat:20s}")
    print("=" * 50)


if __name__ == "__main__":
    main()
