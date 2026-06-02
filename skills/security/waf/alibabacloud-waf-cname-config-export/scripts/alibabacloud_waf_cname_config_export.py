#!/usr/bin/env python3
"""
WAF 3.0 CNAME Domain Configuration Export Script

Batch export all CNAME-based domain configurations under WAF 3.0 instances to Excel.
Supports both Chinese mainland (cn-hangzhou) and non-Chinese mainland (ap-southeast-1) regions.
"""

import json
import subprocess
import time
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl==3.1.5")
    sys.exit(1)

REGION_CONFIG = {
    "cn-hangzhou": "Chinese Mainland",
    "ap-southeast-1": "Non-Chinese Mainland"
}


def run_cli(args, cli_profile=None, timeout=60):
    """Execute aliyun CLI command. Returns parsed JSON dict on success, or None on failure."""
    cmd = ["aliyun"] + args
    if cli_profile:
        cmd.extend(["--profile", cli_profile])
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    except FileNotFoundError:
        print("  [Error] `aliyun` CLI not found. Install via: curl -fsSL https://aliyuncli.alicdn.com/setup.sh | bash")
        sys.exit(1)
    except subprocess.TimeoutExpired:
        print(f"  [Error] CLI command timed out after {timeout}s: {' '.join(args[:3])}...")
        return None
    if result.returncode != 0:
        print(f"  [Error] CLI command failed: {result.stderr.strip()}")
        return None
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        print(f"  [Error] Failed to parse CLI output as JSON: {result.stdout[:200]}")
        return None


def discover_instances(regions=None, cli_profile=None):
    """Auto-discover all WAF instances under the account"""
    if regions is None:
        regions = list(REGION_CONFIG.keys())
    instances = []
    for region in regions:
        print(f"Querying {REGION_CONFIG[region]} instance...")
        data = run_cli(["waf-openapi", "describe-instance", "--region", region], cli_profile)
        if data and data.get("InstanceId"):
            instances.append({
                "region": region,
                "instance_id": data["InstanceId"],
                "region_label": REGION_CONFIG[region]
            })
            print(f"  [Found] {REGION_CONFIG[region]}: {data['InstanceId']}")
        else:
            print(f"  [None] {REGION_CONFIG.get(region, region)}")
    return instances


def query_domains(instance_id, region, cli_profile=None):
    """Paginate through all domains"""
    all_domains, page = [], 1
    while True:
        data = run_cli([
            "waf-openapi", "describe-domains",
            "--region", region, "--instance-id", instance_id,
            "--page-number", str(page), "--page-size", "50"
        ], cli_profile)
        if not data:
            break
        all_domains.extend(data.get("Domains", []))
        if len(all_domains) >= data.get("TotalCount", 0):
            break
        page += 1
    return all_domains


def query_domain_detail(instance_id, domain, region, cli_profile=None):
    """Query single domain detailed config"""
    return run_cli([
        "waf-openapi", "describe-domain-detail",
        "--region", region, "--instance-id", instance_id, "--domain", domain
    ], cli_profile) or {}


def export_to_excel(region_details, output_path):
    """Export domain configs to Excel, one sheet per region"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    headers = [
        "Domain", "CNAME", "Status", "HTTP Ports", "HTTPS Ports",
        "Backends", "Backup Backends", "Load Balancing", "TLS Version", "HTTP/2", "Cert ID",
        "SNI", "SNI Host", "Connect Timeout(s)", "Read Timeout(s)", "Write Timeout(s)",
        "Force HTTP Backend", "Resource Group"
    ]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    for region_label, details in region_details.items():
        if not details:
            continue
        ws = wb.create_sheet(title=region_label[:31])
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.border = hfont, hfill, border
            c.alignment = Alignment(horizontal="center", vertical="center")
        for ri, d in enumerate(details, 2):
            listen, redirect = d.get("Listen", {}), d.get("Redirect", {})
            row = [
                d.get("Domain", ""),
                d.get("Cname", ""),
                {1: "Normal", 2: "Abnormal", 3: "Configuring"}.get(d.get("Status"), str(d.get("Status", ""))),
                ", ".join(str(x) for x in listen.get("HttpPorts", [])),
                ", ".join(str(x) for x in listen.get("HttpsPorts", [])),
                ", ".join(str(x) for x in redirect.get("BackendList", [])),
                ", ".join(str(x) for x in redirect.get("BackupBackends", [])),
                redirect.get("Loadbalance", ""),
                listen.get("TLSVersion", ""),
                "Yes" if listen.get("Http2Enabled") else "No",
                listen.get("CertId", ""),
                "Yes" if redirect.get("SniEnabled") else "No",
                redirect.get("SniHost", ""),
                redirect.get("ConnectTimeout", ""),
                redirect.get("ReadTimeout", ""),
                redirect.get("WriteTimeout", ""),
                "Yes" if redirect.get("FocusHttpBackend") else "No",  # Note: API field is `FocusHttpBackend` (Force-HTTP-to-backend semantics)
                d.get("ResourceManagerResourceGroupId", "")
            ]
            for col, v in enumerate(row, 1):
                c = ws.cell(row=ri, column=col, value=v)
                c.border, c.alignment = border, Alignment(vertical="center", wrap_text=True)
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max(ml + 4, 10), 40)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    CLI_PROFILE = None   # Set to profile name if user specifies (e.g., "intl")
    REGIONS = None       # Always both regions
    OUTPUT_FILE = f"waf_cname_config_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print("=" * 60)
    print("WAF 3.0 CNAME Domain Configuration Export")
    print("=" * 60)
    print()

    print("[Step 1] Discovering WAF instances...")
    instances = discover_instances(REGIONS, CLI_PROFILE)
    if not instances:
        print("\n[Error] No WAF instances found. Exiting.")
        sys.exit(1)

    region_details = {}  # {"Chinese Mainland": [...], "Non-Chinese Mainland": [...]}
    for inst in instances:
        print(f"\n[Step 2] Querying [{inst['region_label']}] {inst['instance_id']} domains...")
        domains = query_domains(inst["instance_id"], inst["region"], CLI_PROFILE)
        print(f"  Found {len(domains)} domains")
        
        print(f"\n[Step 3] Querying detailed config for each domain...")
        details = []
        if not domains:
            # [MUST per SKILL.md] Probe DescribeDomainDetail once even with an empty
            # domain list, to verify API reachability/permissions. Error result is
            # tolerated (run_cli returns None -> {}); we still record "no domains".
            print(f"  [Probe] Empty domain list; calling describe-domain-detail once to verify API reachability...")
            query_domain_detail(inst["instance_id"], "none", inst["region"], CLI_PROFILE)
        for i, dm in enumerate(domains):
            dn = dm.get("Domain", "")
            print(f"  [{i+1}/{len(domains)}] {dn}")
            detail = query_domain_detail(inst["instance_id"], dn, inst["region"], CLI_PROFILE)
            detail["Status"] = dm.get("Status", "")
            details.append(detail)
            time.sleep(0.2)  # Rate limiting
        region_details[inst["region_label"]] = details

    print(f"\n[Step 4] Generating Excel export...")
    export_to_excel(region_details, OUTPUT_FILE)
    total = sum(len(v) for v in region_details.values())
    print(f"\n[Success] Export complete: {OUTPUT_FILE} ({total} domains)")
    print("=" * 60)
