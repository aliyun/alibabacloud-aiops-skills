#!/usr/bin/env python3
"""
Alibaba Cloud WAF 3.0 Full Configuration Backup Exporter

Auto-discovers WAF instances across regions (cn-hangzhou, ap-southeast-1),
queries every dimension with pagination and throttling, tolerates unavailable/empty
dimensions, and writes a multi-sheet Excel workbook.

Usage:
    python3 export_waf_config_backup.py [--profile PROFILE] [--output PATH]
"""

import argparse
import glob
import json
import os
import re
import subprocess
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# === Configuration ===

REGIONS = [
    {"id": "cn-hangzhou", "label": "CN", "desc": "Chinese Mainland"},
    {"id": "ap-southeast-1", "label": "Intl", "desc": "International"},
]

DEFENSE_SCENES = [
    {"scene": "waf_group", "rule_type": "defense"},
    {"scene": "custom_acl", "rule_type": "defense"},
    {"scene": "cc", "rule_type": "defense"},
    {"scene": "ip_blacklist", "rule_type": "defense"},
    {"scene": "whitelist", "rule_type": "whitelist"},
    {"scene": "region_block", "rule_type": "defense"},
    {"scene": "antiscan", "rule_type": "defense"},
    {"scene": "bot_manager", "rule_type": "defense"},
    {"scene": "dlp", "rule_type": "defense"},
    {"scene": "tamperproof", "rule_type": "defense"},
    {"scene": "custom_response", "rule_type": "defense"},
]

API_DELAY = 0.2  # 200ms between API calls

# === Retry Policy ===
RETRY_MAX = 3                     # Max attempts (initial + retries)
RETRY_BASE_DELAY = 1.0            # Base delay seconds; exponential: 1s -> 2s -> 4s
CLI_TIMEOUT_SEC = 60              # Single CLI subprocess timeout

# Retryable: transient server / network conditions
RETRYABLE_ERRORS = (
    "Throttling",
    "ServiceUnavailable",
    "InternalError",
    "RequestTimeout",
    "Unavailable",
)

# Non-retryable: client/auth/business errors -- fail fast
NON_RETRYABLE_ERRORS = (
    "InvalidInstance",
    "InstanceNotFound",
    "InvalidAccessKey",
    "Forbidden",
    "InvalidSecurityToken",
)

# === Observability ===
SKILL_NAME = "alibabacloud-waf-config-backup"
SKILL_SESSION_ID = uuid.uuid4().hex  # 32-char lowercase hex, generated once per invocation
USER_AGENT = "AlibabaCloud-Agent-Skills/alibabacloud-waf-config-backup/" + SKILL_SESSION_ID

# === i18n Labels ===

LABELS = {
    "en": {
        "title": "WAF 3.0 Full Configuration Backup",
        "backup_time": "Backup Time",
        "profile": "Profile",
        "region_summary": "Region Summary",
        "total": "Total",
        "sheet_summary": "Summary",
        "sheet_instance": "Instance Info",
        "sheet_guide": "Guide",
        "sheet_cname": "CNAME",
        "sheet_cloud": "Cloud Resource",
        "sheet_hybrid": "Hybrid Cloud",
        "sheet_objects": "Protected Objects",
        "sheet_groups": "Object Groups",
        "sheet_templates": "Templates",
        "sheet_bindings": "Bindings",
        "sheet_audit": "Audit View",
        "sheet_reverse": "Object Rules View",
        "sheet_rules": "Defense Rules",
        "sheet_members": "Member Accounts",
        "scene_labels": {
            "waf_group": "Basic Protection",
            "custom_acl": "Custom Rules",
            "cc": "CC Protection",
            "ip_blacklist": "IP Blacklist",
            "whitelist": "Whitelist",
            "region_block": "Region Block",
            "antiscan": "Anti-Scan",
            "bot_manager": "Bot Management",
            "dlp": "Data Leak Prevention",
            "tamperproof": "Anti-Tampering",
            "custom_response": "Custom Response",
        },
        "headers": {
            "summary": ["Region", "Instance", "CNAME Domains", "Cloud Resources",
                        "Hybrid Cloud", "Protected Objects", "Object Groups",
                        "Templates", "Rules", "Bindings"],
            "members": ["AccountId", "AccountName", "AccountStatus", "Description", "GmtCreate"],
            "instance": ["Region", "InstanceId", "Edition", "PayType", "Status",
                         "StartTime", "EndTime", "Raw JSON"],
            "cname": ["OwnerUid", "Domain", "Status", "HTTP Ports", "HTTPS Ports", "TLS Version",
                      "HTTP2", "IPv6", "Backends", "Loadbalance", "CertId", "CertName",
                      "ConnectTimeout", "ReadTimeout", "WriteTimeout", "Raw JSON"],
            "cloud": ["OwnerUid", "ResourceInstanceId", "ResourceProduct", "ResourceRegionId",
                      "ResourceDomain", "ResourceInstanceIp", "Raw JSON"],
            "hybrid": ["OwnerUid", "Domain", "Ports", "HTTPS", "Backends", "LoadBalance", "Raw JSON"],
            "objects": ["OwnerUid", "Resource", "Product", "Pattern", "OwnerUserId",
                        "GmtCreate", "GmtModified", "Raw JSON"],
            "groups": ["GroupName", "Description", "Members", "Raw JSON"],
            "templates": ["TemplateId", "TemplateName", "DefenseScene", "TemplateOrigin",
                          "TemplateStatus", "DefenseSubScene", "Raw JSON"],
            "rules": ["DefenseScene", "RuleId", "RuleName", "Purpose", "TemplateId",
                      "Status", "DefenseOrigin", "GmtModified", "Action",
                      "Conditions Summary", "Config (Raw JSON)"],
            "bindings": ["OwnerUid", "TemplateId", "TemplateName", "DefenseScene", "Resource",
                         "Product", "Pattern", "Raw JSON"],
        },
        "status_active": "Active",
        "status_inactive": "Inactive",
        "actions": {},
        "origins": {},
        "products": {},
        "patterns": {},
        "pay_types": {},
        "editions": {},
        "guide": [
            "WAF 3.0 Configuration Backup - Guide",
            "",
            "This workbook contains a full point-in-time backup of your WAF 3.0 configuration.",
            "",
            "═══ Core Concepts ═══",
            "",
            "WAF Workflow:  Domain/Instance → Protected Object → Defense Template → Rules",
            "",
            "• Protected Object: The unit WAF protects. Each CNAME domain or cloud-product",
            "  instance becomes a protected object. Rules apply TO protected objects.",
            "• Defense Template: A policy container that groups rules of ONE scene type.",
            "  E.g., one 'Basic Protection' template, one 'CC Protection' template.",
            "• Binding: Links a template to one or more protected objects. An object can",
            "  bind multiple templates (one per scene).",
            "• Defense Scene: Category of protection. 11 scenes available (see below).",
            "",
            "═══ Sheet Descriptions ═══",
            "",
            "  Summary         - Dashboard with backup metadata and counts per region",
            "  Instance Info   - WAF instance specs (edition, quotas) + raw JSON",
            "  Audit View      - Coverage matrix: which objects have which scene protection",
            "  Object Rules    - Reverse view: per-object, all effective rules listed",
            "  CNAME (XX)      - CNAME onboarding: ports, backends, TLS, cert",
            "  Cloud Resource  - Cloud-product onboarding (ALB/CLB/ECS)",
            "  Hybrid Cloud    - Hybrid-cloud onboarding",
            "  Protected Obj.  - Full standalone object inventory",
            "  Object Groups   - Object group -> member mappings",
            "  Templates       - Defense templates (id, type, origin, status)",
            "  Defense Rules   - All scene rules merged into one sheet per region; first column 'DefenseScene' is filterable; raw Config JSON column hidden",
            "  Bindings        - Template -> protected object mappings",
            "",
            "═══ Reconstruction Notes ═══",
            "",
            "  1. Unhide the Raw JSON columns (last col of most sheets) for rebuild data.",
            "  2. TLS certificate private keys are NOT exportable - only CertId/CertName.",
            "  3. Cloud-product instances (ALB/CLB/ECS) must exist before re-onboarding.",
            "  4. Re-creating rules: use the Config JSON from each rule row.",
        ],
    },
    "zh": {
        "title": "WAF 3.0 全量配置备份",
        "backup_time": "备份时间",
        "profile": "配置档",
        "region_summary": "区域概览",
        "total": "合计",
        "sheet_summary": "概览",
        "sheet_instance": "实例信息",
        "sheet_guide": "说明",
        "sheet_cname": "CNAME接入",
        "sheet_cloud": "云产品接入",
        "sheet_hybrid": "混合云接入",
        "sheet_objects": "防护对象",
        "sheet_groups": "对象组",
        "sheet_templates": "防护模板",
        "sheet_bindings": "绑定关系",
        "sheet_audit": "审计视图",
        "sheet_reverse": "对象规则汇总",
        "sheet_rules": "防护规则",
        "sheet_members": "成员账号",
        "scene_labels": {
            "waf_group": "基础防护",
            "custom_acl": "自定义规则",
            "cc": "CC防护",
            "ip_blacklist": "IP黑名单",
            "whitelist": "白名单",
            "region_block": "区域封禁",
            "antiscan": "扫描防护",
            "bot_manager": "Bot管理",
            "dlp": "数据防泄漏",
            "tamperproof": "网页防篡改",
            "custom_response": "自定义响应",
        },
        "headers": {
            "summary": ["区域", "实例", "CNAME域名", "云产品资源",
                        "混合云", "防护对象", "对象组",
                        "模板", "规则", "绑定"],
            "members": ["账号ID", "账号名称", "状态", "描述", "加入时间"],
            "instance": ["区域", "实例ID", "版本", "付费类型", "状态",
                         "开通时间", "到期时间", "原始JSON"],
            "cname": ["归属账号", "域名", "状态", "HTTP端口", "HTTPS端口", "TLS版本",
                      "HTTP2", "IPv6", "回源地址", "负载均衡", "证书ID", "证书名称",
                      "连接超时", "读超时", "写超时", "原始JSON"],
            "cloud": ["归属账号", "资源实例ID", "资源产品", "资源地域",
                      "资源域名", "资源IP", "原始JSON"],
            "hybrid": ["归属账号", "域名", "端口", "HTTPS", "后端", "负载均衡", "原始JSON"],
            "objects": ["归属账号", "资源名称", "产品", "匹配模式", "所属账号",
                        "创建时间", "修改时间", "原始JSON"],
            "groups": ["组名称", "描述", "成员", "原始JSON"],
            "templates": ["模板ID", "模板名称", "防护场景", "来源",
                          "状态", "子场景", "原始JSON"],
            "rules": ["防护场景", "规则ID", "规则名称", "规则用途", "模板ID",
                      "状态", "来源", "修改时间", "动作",
                      "条件摘要", "配置(原始JSON)"],
            "bindings": ["归属账号", "模板ID", "模板名称", "防护场景", "资源",
                         "产品", "匹配模式", "原始JSON"],
        },
        "status_active": "已启用",
        "status_inactive": "已停用",
        "actions": {
            "block": "拦截",
            "monitor": "观察",
            "captcha": "滑块验证",
            "strict_captcha": "严格滑块",
            "js": "JS验证",
            "allow": "放行",
            "bypass": "跳过检测",
            "deny": "拒绝",
            "continue": "继续执行",
        },
        "origins": {
            "custom": "自定义",
            "system": "系统内置",
            "scene": "场景内置",
        },
        "products": {
            "waf": "WAF域名",
            "alb": "ALB实例",
            "clb": "CLB实例",
            "ecs": "ECS实例",
            "mse": "MSE实例",
            "fc": "FC函数",
        },
        "patterns": {
            "domain": "域名",
            "instance": "实例",
        },
        "pay_types": {
            "POSTPAY": "按量付费",
            "PREPAY": "包年包月",
        },
        "editions": {
            "default_version": "基础版",
            "basic": "基础版",
            "pro": "专业版",
            "enterprise": "企业版",
            "ultimate": "旗舰版",
        },
        "guide": [
            "WAF 3.0 全量配置备份 - 使用说明",
            "",
            "本工作簿包含 WAF 3.0 全量配置的时间点备份。",
            "",
            "═══ 核心概念 ═══",
            "",
            "WAF 工作流：域名/实例 → 防护对象 → 防护模板 → 规则",
            "",
            "• 防护对象：WAF 保护的最小单元。每个CNAME域名或云产品实例接入后",
            "  自动生成一个防护对象。规则通过模板绑定到防护对象上。",
            "• 防护模板：策略容器，将同一场景的规则组织在一起。",
            "  例如：一个\'基础防护\'模板、一个\'CC防护\'模板。",
            "• 绑定关系：将模板关联到一个或多个防护对象。一个对象可以绑定",
            "  多个模板（每个场景最多一个）。",
            "• 防护场景：保护类型分类，共 11 个场景（见下方说明）。",
            "",
            "═══ Sheet 说明 ═══",
            "",
            "  概览            - 备份元数据及各区域统计",
            "  实例信息        - WAF 实例规格（版本、配额）",
            "  审计视图        - 覆盖矩阵：哪些对象已启用哪些防护场景",
            "  对象规则汇总    - 反向视图：每个对象的所有生效规则列表",
            "  CNAME接入(XX)  - CNAME 接入：端口、回源、TLS、证书",
            "  云产品接入      - 云产品接入（ALB/CLB/ECS）",
            "  混合云接入      - 混合云接入",
            "  防护对象        - 防护对象全量清单",
            "  对象组          - 对象组 -> 成员映射",
            "  防护模板        - 防护模板（ID、类型、来源、状态）",
            "  防护规则 (XX)   - 各区域所有场景规则合并为一个 sheet，首列\u201c防护场景\u201d可筛选；原始 Config JSON 列已隐藏",
            "  绑定关系        - 模板 -> 防护对象绑定",
            "",
            "═══ 恢复说明 ═══",
            "",
            "  1. 取消隐藏“原始JSON”列（各sheet最后一列）可获取重建配置的权威数据源。",
            "  2. TLS 证书私钥不可导出，仅记录 CertId/CertName 引用。",
            "  3. 云产品实例（ALB/CLB/ECS）需先存在才能重新接入。",
            "  4. 重建规则时使用各规则行的 Config JSON。",
        ],
    },
}
PAGE_SIZE = 50

# === System Rule Purpose Dictionary ===
# Maps known system rule name patterns to human-readable descriptions
RULE_PURPOSE_DICT = {
    "en": {
        "WAF sema rule": "Semantic analysis engine - detects SQLi/XSS/RCE via code semantics",
        "WAF group rule": "Rule-group engine - signature-based detection (OWASP Top 10)",
        "WAF compliance rule": "Compliance engine - HTTP protocol compliance checks",
        "scan_tool": "Blocks known scanner/vulnerability-assessment tools",
        "scantools": "Blocks known scanner/vulnerability-assessment tools",
        "scan_path": "Blocks probing of sensitive directories/files",
        "dirscan": "Blocks probing of sensitive directories/files",
        "scan_frequency": "Rate-limits high-frequency scanning behavior",
        "highfreq": "Rate-limits high-frequency scanning behavior",
    },
    "zh": {
        "WAF sema rule": "语义分析引擎 - 基于代码语义检测SQLi/XSS/RCE等攻击",
        "WAF group rule": "规则组引擎 - 基于签名的攻击检测（OWASP Top 10）",
        "WAF compliance rule": "合规引擎 - HTTP协议合规性检查",
        "scan_tool": "拦截已知扫描器/漏洞评估工具",
        "scantools": "拦截已知扫描器/漏洞评估工具",
        "scan_path": "阻止对敏感目录/文件的探测",
        "dirscan": "阻止对敏感目录/文件的探测",
        "scan_frequency": "限制高频扫描行为",
        "highfreq": "限制高频扫描行为",
    },
}

# Bot management rule name pattern -> purpose
BOT_RULE_PURPOSE = {
    "en": {
        "suspicious_": "Suspicious behavior detection",
        "malicious_": "Malicious bot blocking",
        "normal_": "Known benign bot allowlist",
        "search_engine": "Search engine crawler management",
        "feed_reader": "Feed reader bot management",
        "monitoring": "Monitoring service management",
    },
    "zh": {
        "suspicious_": "可疑行为检测",
        "malicious_": "恶意Bot拦截",
        "normal_": "已知良性Bot白名单",
        "search_engine": "搜索引擎爬虫管理",
        "feed_reader": "Feed阅读器Bot管理",
        "monitoring": "监控服务管理",
    },
}

# Scene description for guide
SCENE_DESCRIPTIONS = {
    "en": {
        "waf_group": "Web Application Firewall core engine: protects against SQL injection, XSS, command injection, and other OWASP Top 10 attacks",
        "custom_acl": "User-defined access control rules: match by IP/URL/Header/Cookie/etc. to block/allow/challenge requests",
        "cc": "HTTP flood (CC) protection: rate-limits requests per IP/Session to prevent application-layer DDoS",
        "ip_blacklist": "IP reputation blacklist: blocks known malicious IPs or specified IP ranges",
        "whitelist": "Trusted traffic bypass: skips WAF detection for specified conditions (internal IPs, health checks)",
        "region_block": "Geo-blocking: blocks or allows traffic based on source country/region",
        "antiscan": "Anti-scanning: detects and blocks automated vulnerability scanning and directory traversal",
        "bot_manager": "Bot management: identifies and controls automated access (crawlers, scrapers, brute-force tools)",
        "dlp": "Data leak prevention: masks or blocks sensitive data (phone/ID/bank card numbers) in responses",
        "tamperproof": "Anti-tampering: caches and locks specified pages to prevent content modification",
        "custom_response": "Custom block pages: defines response pages shown when requests are blocked",
    },
    "zh": {
        "waf_group": "Web应用防火墙核心引擎：防护SQL注入、XSS跨站脚本、命令注入等OWASP Top 10攻击",
        "custom_acl": "用户自定义访问控制规则：按IP/URL/Header/Cookie等条件匹配，执行拦截/放行/验证等动作",
        "cc": "HTTP Flood（CC）防护：按IP/Session限制请求频率，防止应用层DDoS攻击",
        "ip_blacklist": "IP信誉黑名单：拦截已知恶意IP或指定IP段",
        "whitelist": "可信流量放行：对符合条件的请求（内部IP、健康检查等）跳过WAF检测",
        "region_block": "地域封禁：基于请求来源国家/地区进行拦截或放行",
        "antiscan": "扫描防护：检测并阻止自动化漏洞扫描和目录遍历行为",
        "bot_manager": "Bot管理：识别和管控自动化访问（爬虫、数据抓取、暴力破解工具等）",
        "dlp": "数据防泄漏：对响应中的敏感信息（手机号/身份证/银行卡号）进行脱敏或拦截",
        "tamperproof": "网页防篡改：缓存并锁定指定页面，防止内容被恶意篡改",
        "custom_response": "自定义响应页面：定义请求被拦截时显示的自定义错误页面",
    },
}


# === CLI Helper ===

def run_cli(cmd, profile=None):
    """Run aliyun CLI command with retry policy and return parsed JSON or None.

    Retry policy:
      - Max attempts: RETRY_MAX (default 3)
      - Backoff: exponential (1s -> 2s -> 4s)
      - Retryable: throttling / 5xx / network timeout / JSON decode flakiness
      - Non-retryable: invalid instance / auth errors -> fail fast

    Observability:
      - Appends --user-agent to every call for tracing.
    """
    if profile:
        cmd += f" --profile {profile}"
    cmd += f' --user-agent "{USER_AGENT}"'

    last_err = ""
    for attempt in range(1, RETRY_MAX + 1):
        try:
            result = subprocess.run(
                cmd, shell=True, capture_output=True, text=True,
                timeout=CLI_TIMEOUT_SEC,
            )
            if result.returncode == 0:
                if not result.stdout.strip():
                    return None
                try:
                    return json.loads(result.stdout)
                except json.JSONDecodeError as e:
                    last_err = f"JSONDecodeError: {e}"
                    # Treat as transient (CLI sometimes prints non-JSON warnings)
            else:
                stderr = result.stderr.strip()
                last_err = stderr
                # Fail fast on non-retryable errors
                if any(k in stderr for k in NON_RETRYABLE_ERRORS):
                    return None
                # Only retry on known transient errors
                if not any(k in stderr for k in RETRYABLE_ERRORS):
                    return None
        except subprocess.TimeoutExpired:
            last_err = f"TimeoutExpired ({CLI_TIMEOUT_SEC}s)"
            # Timeouts are always retryable

        if attempt < RETRY_MAX:
            delay = RETRY_BASE_DELAY * (2 ** (attempt - 1))
            print(f"  \u26a0 Retryable error (attempt {attempt}/{RETRY_MAX}): {last_err[:120]}; retry in {delay:.0f}s...")
            time.sleep(delay)

    print(f"  \u2717 Gave up after {RETRY_MAX} attempts: {last_err[:160]}")
    return None


def paginate(base_cmd, items_key, profile=None, page_size=PAGE_SIZE):
    """Paginate through a list API and return all items."""
    all_items = []
    page = 1
    while True:
        cmd = f"{base_cmd} --page-size {page_size} --page-number {page}"
        data = run_cli(cmd, profile)
        time.sleep(API_DELAY)
        if not data:
            break
        items = data.get(items_key, [])
        if isinstance(items, list):
            all_items.extend(items)
        total = data.get("TotalCount", 0)
        if page * page_size >= total:
            break
        page += 1
    return all_items


# === Account Topology (Single UID / Delegated Multi-UID) ===

def fetch_caller_identity(profile):
    """Fetch the caller's AliUid via sts get-caller-identity."""
    cmd = "aliyun sts get-caller-identity"
    data = run_cli(cmd, profile)
    if isinstance(data, dict):
        uid = data.get("AccountId") or data.get("UserId")
        if uid:
            return str(uid).strip()
    return None


def fetch_delegated_status(region_id, instance_id, profile):
    """Check whether the caller is the WAF delegated administrator.

    Returns True only when the API explicitly confirms delegated-admin status;
    any error or negative response is treated as single-UID.
    """
    cmd = (f"aliyun waf-openapi describe-account-delegated-status "
           f"--region {region_id} --instance-id {instance_id}")
    data = run_cli(cmd, profile)
    time.sleep(API_DELAY)
    if not isinstance(data, dict):
        return False
    val = data.get("DelegatedStatus")
    if val is None:
        val = data.get("Status")
    if val is None:
        val = data.get("IsDelegated")
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.lower() in ("enabled", "true", "yes", "1")
    return False


def fetch_member_accounts(region_id, instance_id, profile):
    """Fetch the list of member accounts under multi-account management."""
    cmd = (f"aliyun waf-openapi describe-member-accounts "
           f"--region {region_id} --instance-id {instance_id}")
    data = run_cli(cmd, profile)
    time.sleep(API_DELAY)
    if not isinstance(data, dict):
        return []
    accounts = data.get("AccountInfos") or data.get("AccountInfo") or []
    return accounts if isinstance(accounts, list) else []


def fetch_resource_owner_uid(region_id, instance_id, resource_names, profile):
    """Batch query ResourceName -> OwnerUid mapping (max 100 per call)."""
    if not resource_names:
        return {}
    mapping = {}
    batch_size = 100
    unique_names = list(dict.fromkeys(n for n in resource_names if n))
    for i in range(0, len(unique_names), batch_size):
        batch = unique_names[i:i + batch_size]
        names_arg = ",".join(batch)
        cmd = (f"aliyun waf-openapi describe-defense-resource-owner-uid "
               f"--region {region_id} --instance-id {instance_id} "
               f"--resource-names \"{names_arg}\"")
        data = run_cli(cmd, profile)
        time.sleep(API_DELAY)
        if not isinstance(data, dict):
            continue
        for info in data.get("OwnerInfos", []) or []:
            name = info.get("ResourceName")
            uid = info.get("OwnerUserId")
            if name and uid:
                mapping[name] = str(uid)
    return mapping


class OwnerUidResolver:
    """Resolve OwnerUid for any row given account topology.

    - Single-UID scenario: every row returns the caller's UID.
    - Delegated multi-UID: lookup by ResourceName, fall back to caller's UID.

    Note: Only member-scoped resources (CNAME/Cloud/Hybrid/Objects/Bindings)
    use this resolver. Admin-scoped sheets (Instance/Groups/Templates/Rules)
    have no per-row OwnerUid column.
    """

    def __init__(self, caller_uid, is_delegated, region_owner_maps=None):
        self.caller_uid = str(caller_uid or "")
        self.is_delegated = bool(is_delegated)
        # region_label -> {ResourceName: OwnerUid}
        self.region_owner_maps = region_owner_maps or {}

    def for_resource(self, region_label, resource_name):
        if not self.is_delegated:
            return self.caller_uid
        if not resource_name:
            return self.caller_uid
        return self.region_owner_maps.get(region_label, {}).get(
            resource_name, self.caller_uid)

    def for_caller(self):
        return self.caller_uid


# === Data Collection ===

def discover_instance(region_id, profile):
    """Discover WAF instance in a region."""
    print(f"  Discovering instance in {region_id}...")
    cmd = f"aliyun waf-openapi describe-instance --region {region_id}"
    data = run_cli(cmd, profile)
    time.sleep(API_DELAY)
    if data and data.get("InstanceId"):
        print(f"    ✓ Found: {data['InstanceId']}")
        return data
    return None


def fetch_domains(region_id, instance_id, profile):
    """Fetch CNAME domain list."""
    print(f"  Fetching CNAME domains...")
    base_cmd = f"aliyun waf-openapi describe-domains --region {region_id} --instance-id {instance_id}"
    domains = paginate(base_cmd, "Domains", profile)
    print(f"    ✓ {len(domains)} domains")
    return domains


def fetch_domain_detail(region_id, instance_id, domain, profile):
    """Fetch single domain detail."""
    cmd = f"aliyun waf-openapi describe-domain-detail --region {region_id} --instance-id {instance_id} --domain {domain}"
    data = run_cli(cmd, profile)
    time.sleep(API_DELAY)
    return data


def fetch_cloud_resources(region_id, instance_id, profile):
    """Fetch cloud-product onboarding resources."""
    print(f"  Fetching cloud resources...")
    cmd = f"aliyun waf-openapi describe-cloud-resources --region {region_id} --instance-id {instance_id}"
    data = run_cli(cmd, profile)
    time.sleep(API_DELAY)
    if data:
        resources = data.get("CloudResources", [])
        print(f"    ✓ {len(resources)} cloud resources")
        return resources, data
    print(f"    - No cloud resources")
    return [], None


def fetch_hybrid_cloud_resources(region_id, instance_id, profile):
    """Fetch hybrid-cloud onboarding resources."""
    print(f"  Fetching hybrid cloud resources...")
    cmd = f"aliyun waf-openapi describe-hybrid-cloud-resources --region {region_id} --instance-id {instance_id}"
    data = run_cli(cmd, profile)
    time.sleep(API_DELAY)
    if data:
        resources = data.get("HybridCloudResources", data.get("Resources", []))
        if isinstance(resources, list):
            print(f"    ✓ {len(resources)} hybrid cloud resources")
            return resources, data
    print(f"    - No hybrid cloud resources")
    return [], None


def fetch_defense_resources(region_id, instance_id, profile, template_id=None):
    """Fetch protected objects (full inventory or per-template bindings)."""
    if template_id:
        base_cmd = (f"aliyun waf-openapi describe-defense-resources --region {region_id} "
                    f"--instance-id {instance_id} --query '{{\"TemplateId\":{template_id}}}'")
    else:
        base_cmd = (f"aliyun waf-openapi describe-defense-resources --region {region_id} "
                    f"--instance-id {instance_id}")
    resources = paginate(base_cmd, "Resources", profile)
    return resources


def fetch_defense_resource_groups(region_id, instance_id, profile):
    """Fetch protected-object groups."""
    print(f"  Fetching object groups...")
    base_cmd = (f"aliyun waf-openapi describe-defense-resource-groups --region {region_id} "
                f"--instance-id {instance_id}")
    groups = paginate(base_cmd, "Groups", profile)
    print(f"    ✓ {len(groups)} object groups")
    return groups


def fetch_defense_templates(region_id, instance_id, scene, profile):
    """Fetch defense templates for a scene."""
    base_cmd = (f"aliyun waf-openapi describe-defense-templates --region {region_id} "
                f"--instance-id {instance_id} --defense-scene {scene}")
    templates = paginate(base_cmd, "Templates", profile)
    return templates


def fetch_defense_rules(region_id, instance_id, scene, rule_type, template_id, profile):
    """Fetch defense rules for a template."""
    query_args = json.dumps({"DefenseScene": scene, "TemplateId": template_id})
    base_cmd = (f"aliyun waf-openapi describe-defense-rules --region {region_id} "
                f"--instance-id {instance_id} --rule-type {rule_type} "
                f"--query '{query_args}'")
    rules = paginate(base_cmd, "Rules", profile)
    return rules


# === Excel Writing Helpers ===

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


# === Formatting Helpers ===

def fmt_epoch_ms(val):
    """Convert epoch milliseconds to human-readable datetime string."""
    if not val or not isinstance(val, (int, float)):
        return val
    try:
        dt = datetime.fromtimestamp(val / 1000, tz=timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
    except (ValueError, OSError):
        return val


def fmt_status(val, labels=None):
    """Convert numeric status to human-readable label."""
    L = labels or LABELS["en"]
    if val == 1:
        return L["status_active"]
    elif val == 0:
        return L["status_inactive"]
    return val


def fmt_action(val, labels=None):
    """Translate action value using labels mapping."""
    if not val:
        return val
    L = labels or LABELS["en"]
    return L.get("actions", {}).get(str(val), val)


def fmt_origin(val, labels=None):
    """Translate defense origin value."""
    if not val:
        return val
    L = labels or LABELS["en"]
    return L.get("origins", {}).get(str(val), val)


def fmt_product(val, labels=None):
    """Translate product type."""
    if not val:
        return val
    L = labels or LABELS["en"]
    return L.get("products", {}).get(str(val), val)


def fmt_pattern(val, labels=None):
    """Translate pattern type."""
    if not val:
        return val
    L = labels or LABELS["en"]
    return L.get("patterns", {}).get(str(val), val)


def fmt_pay_type(val, labels=None):
    """Translate pay type."""
    if not val:
        return val
    L = labels or LABELS["en"]
    return L.get("pay_types", {}).get(str(val), val)


def fmt_edition(val, labels=None):
    """Translate edition."""
    if not val:
        return val
    L = labels or LABELS["en"]
    return L.get("editions", {}).get(str(val), val)


def fmt_list(val):
    """Format a list as comma-separated string without brackets."""
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    if isinstance(val, str) and val.startswith("["):
        try:
            items = json.loads(val.replace("'", '"'))
            if isinstance(items, list):
                return ", ".join(str(v) for v in items)
        except (json.JSONDecodeError, ValueError):
            pass
    return val


def fmt_bool(val):
    """Format boolean values."""
    if val is True or val == "True":
        return "Yes"
    elif val is False or val == "False":
        return "No"
    return val


def get_rule_purpose(rule_name, scene_name, lang):
    """Get a human-readable purpose description for a rule."""
    if not rule_name:
        return ""
    # Check exact match in RULE_PURPOSE_DICT
    purpose_dict = RULE_PURPOSE_DICT.get(lang, RULE_PURPOSE_DICT["en"])
    if rule_name in purpose_dict:
        return purpose_dict[rule_name]
    # Check prefix match for bot rules
    if scene_name == "bot_manager":
        bot_dict = BOT_RULE_PURPOSE.get(lang, BOT_RULE_PURPOSE["en"])
        for prefix, desc in bot_dict.items():
            if rule_name.startswith(prefix) or prefix in rule_name:
                return desc
    # Check partial matches in general dict
    for key, desc in purpose_dict.items():
        if key in rule_name:
            return desc
    return ""


def parse_rule_config(config, scene_name):
    """Parse rule Config to extract action and conditions summary."""
    if not config:
        return "", ""
    try:
        if isinstance(config, str):
            cfg = json.loads(config)
        elif isinstance(config, dict):
            cfg = config
        else:
            return "", ""
    except (json.JSONDecodeError, ValueError):
        return "", ""

    # Extract action
    action = cfg.get("action", cfg.get("Action", ""))

    # Extract conditions summary based on scene
    conditions_parts = []

    # Condition field name mapping for readability
    FIELD_NAMES = {
        "IP": "IP地址", "URLPath": "URL路径", "URL": "URL",
        "Referer": "Referer头", "UserAgent": "User-Agent",
        "Cookie": "Cookie", "Header": "HTTP头",
        "Method": "HTTP方法", "PostBody": "请求体",
        "Params": "请求参数", "ContentType": "Content-Type",
    }
    OP_NAMES = {
        "contain": "包含", "not-contain": "不包含",
        "eq": "等于", "ne": "不等于",
        "prefix-match": "前缀匹配", "suffix-match": "后缀匹配",
        "regex-match": "正则匹配", "len-gt": "长度大于",
        "len-lt": "长度小于", "len-eq": "长度等于",
        "belong": "属于", "not-belong": "不属于",
    }

    if scene_name in ("custom_acl", "whitelist", "cc"):
        # Custom rules / whitelist / CC typically have conditions array
        conds = cfg.get("conditions", cfg.get("Conditions", []))
        if isinstance(conds, list):
            for c in conds[:5]:  # Show first 5 conditions max
                key = c.get("key", c.get("Key", ""))
                op = c.get("opValue", c.get("opCode", c.get("operate", c.get("OpCode", ""))))
                vals = c.get("values", c.get("Values", ""))
                if isinstance(vals, list):
                    shown_vals = ", ".join(str(v) for v in vals[:3])
                    if len(vals) > 3:
                        shown_vals += f" (+{len(vals)-3}more)"
                else:
                    shown_vals = str(vals)
                # Use friendly names
                friendly_key = FIELD_NAMES.get(key, key)
                friendly_op = OP_NAMES.get(op, op)
                conditions_parts.append(f"{friendly_key} {friendly_op} {shown_vals}")
        # CC specific: rate limit info
        if scene_name == "cc":
            rate_limit = cfg.get("rateLimit", {})
            if isinstance(rate_limit, dict):
                threshold = rate_limit.get("threshold", "")
                interval = rate_limit.get("interval", "")
                if threshold and interval:
                    conditions_parts.append(f"频率限制: {threshold}次/{interval}秒")
            elif cfg.get("threshold"):
                conditions_parts.append(f"阈值: {cfg['threshold']}")

    elif scene_name == "ip_blacklist":
        # IP blacklist has remoteAddr list
        ips = cfg.get("remoteAddr", cfg.get("RemoteAddr", []))
        if isinstance(ips, list):
            shown = ips[:5]
            conditions_parts.append(f"IP: {', '.join(shown)}")
            if len(ips) > 5:
                conditions_parts.append(f"(+{len(ips)-5}个)")

    elif scene_name == "region_block":
        regions = cfg.get("regions", cfg.get("Regions", []))
        if isinstance(regions, list):
            if len(regions) <= 3:
                conditions_parts.append(f"地区: {', '.join(str(r) for r in regions)}")
            else:
                conditions_parts.append(f"地区: {len(regions)}个地区被封禁")

    elif scene_name == "bot_manager":
        # Bot management rules
        bot_behavior = cfg.get("botBehavior", "")
        if bot_behavior:
            behavior_names = {"suspicious": "可疑行为", "malicious": "恶意行为", "normal": "正常行为"}
            conditions_parts.append(f"行为={behavior_names.get(bot_behavior, bot_behavior)}")
        group = cfg.get("group", "")
        if group:
            conditions_parts.append(f"分组={group}")
        target = cfg.get("target", "")
        if target:
            conditions_parts.append(f"目标={target}")
        intel_type = cfg.get("intelligenceType", "")
        if intel_type:
            conditions_parts.append(f"情报类型={intel_type}")

    elif scene_name in ("waf_group", "antiscan"):
        # Intelligence-type rules
        intel_type = cfg.get("intelligenceType", "")
        if intel_type:
            conditions_parts.append(f"类型={intel_type}")
        # Rate limit / threshold
        rate = cfg.get("rateLimit", cfg.get("threshold", ""))
        if rate:
            conditions_parts.append(f"阈值={rate}")

    elif scene_name == "dlp":
        # Data leak prevention
        content_type = cfg.get("contentType", "")
        if content_type:
            dlp_types = {"phone": "手机号", "idcard": "身份证号", "bankcard": "银行卡号",
                         "email": "邮箱", "sensitive_word": "敏感词"}
            conditions_parts.append(f"内容类型={dlp_types.get(content_type, content_type)}")

    elif scene_name == "tamperproof":
        uri = cfg.get("uri", cfg.get("Uri", ""))
        if uri:
            conditions_parts.append(f"防护URI={uri}")

    elif scene_name == "custom_response":
        status = cfg.get("responseCode", cfg.get("status", ""))
        if status:
            conditions_parts.append(f"HTTP状态码={status}")

    return str(action), "; ".join(conditions_parts)


def write_header(ws, headers, row=1):
    """Write styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def write_empty_placeholder(ws):
    """Mark a created sheet as 'queried but no data returned'."""
    ws.cell(row=2, column=1, value="(API 已查询，无数据 / API queried; no data)").font = Font(italic=True, color="808080")


def auto_width(ws, max_width=60):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_len = min(len(str(cell.value)), max_width)
                max_len = max(max_len, cell_len)
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)
    # Freeze header row
    ws.freeze_panes = "A2"


# === Sheet Writers ===

def write_summary(wb, backup_meta, region_data, L, lang="en"):
    """Write Summary dashboard sheet."""
    ws = wb.active
    ws.title = L["sheet_summary"]
    ws.cell(row=1, column=1, value=L["title"]).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{L['backup_time']}: {backup_meta['time_display']}")
    ws.cell(row=3, column=1, value=f"{L['profile']}: {backup_meta['profile']}")

    # Account topology line: caller UID + scenario
    caller_uid = backup_meta.get("caller_uid") or "unknown"
    is_delegated = backup_meta.get("is_delegated", False)
    member_count = backup_meta.get("member_count", 0)
    if lang == "zh":
        scenario = (f"委派管理员（管理 {member_count} 个成员账号）"
                    if is_delegated else "单 UID")
        topo_line = f"执行账号 UID: {caller_uid}（场景：{scenario}）"
    else:
        scenario = (f"Delegated Administrator (managing {member_count} member accounts)"
                    if is_delegated else "Single UID")
        topo_line = f"Caller UID: {caller_uid} (Scenario: {scenario})"
    ws.cell(row=4, column=1, value=topo_line).font = Font(bold=True, color="4472C4")

    # Diff baseline line (row 5)
    diff_baseline = backup_meta.get("diff_baseline")
    diff_disabled = backup_meta.get("diff_disabled", False)
    if lang == "zh":
        if diff_baseline:
            baseline_line = f"比对基线: {diff_baseline}"
        elif diff_disabled:
            baseline_line = "比对基线: 已通过 --no-diff 禁用"
        else:
            baseline_line = "比对基线: 无（基线模式 — 未发现历史备份）"
    else:
        if diff_baseline:
            baseline_line = f"Diff baseline: {diff_baseline}"
        elif diff_disabled:
            baseline_line = "Diff baseline: disabled via --no-diff"
        else:
            baseline_line = "Diff baseline: none (baseline mode — no previous backup found)"
    ws.cell(row=5, column=1, value=baseline_line).font = Font(italic=True, color="666666")

    ws.cell(row=6, column=1, value=L["region_summary"]).font = Font(bold=True, size=12)

    write_header(ws, L["headers"]["summary"], row=7)

    row = 8
    for rd in region_data:
        ws.cell(row=row, column=1, value=rd.get("region_label", ""))
        ws.cell(row=row, column=2, value=rd.get("instance_id", "N/A"))
        ws.cell(row=row, column=3, value=rd.get("domain_count", 0))
        ws.cell(row=row, column=4, value=rd.get("cloud_count", 0))
        ws.cell(row=row, column=5, value=rd.get("hybrid_count", 0))
        ws.cell(row=row, column=6, value=rd.get("object_count", 0))
        ws.cell(row=row, column=7, value=rd.get("group_count", 0))
        ws.cell(row=row, column=8, value=rd.get("template_count", 0))
        ws.cell(row=row, column=9, value=rd.get("rule_count", 0))
        ws.cell(row=row, column=10, value=rd.get("binding_count", 0))
        row += 1
    if len(region_data) > 1:
        ws.cell(row=row, column=1, value=L["total"]).font = Font(bold=True)
        count_keys = ["domain_count", "cloud_count", "hybrid_count",
                      "object_count", "group_count", "template_count",
                      "rule_count", "binding_count"]
        for col_offset, key in enumerate(count_keys):
            total = sum(rd.get(key, 0) for rd in region_data)
            ws.cell(row=row, column=col_offset + 3, value=total).font = Font(bold=True)
    auto_width(ws)


def write_member_accounts_sheet(wb, member_accounts, L):
    """Write member accounts sheet (delegated administrator scenario only)."""
    if not member_accounts:
        return
    ws = wb.create_sheet(L["sheet_members"])
    write_header(ws, L["headers"]["members"])
    for row_idx, acct in enumerate(member_accounts, 2):
        ws.cell(row=row_idx, column=1, value=acct.get("AccountId", ""))
        ws.cell(row=row_idx, column=2, value=acct.get("AccountName", ""))
        ws.cell(row=row_idx, column=3, value=acct.get("AccountStatus", acct.get("Status", "")))
        ws.cell(row=row_idx, column=4, value=acct.get("Description", ""))
        ws.cell(row=row_idx, column=5, value=fmt_epoch_ms(acct.get("GmtCreate", "")))
    auto_width(ws)


def write_instance_info(wb, instances, L, resolver=None):
    """Write Instance Info sheet.

    Instance is always admin-scoped; no per-row OwnerUid column.
    The `resolver` parameter is kept for signature compatibility but unused.
    """
    if not instances:
        return
    ws = wb.create_sheet(L["sheet_instance"])
    write_header(ws, L["headers"]["instance"])
    for row_idx, inst in enumerate(instances, 2):
        data = inst["data"]
        ws.cell(row=row_idx, column=1, value=inst["region"])
        ws.cell(row=row_idx, column=2, value=data.get("InstanceId", ""))
        ws.cell(row=row_idx, column=3, value=fmt_edition(data.get("Edition", ""), L))
        ws.cell(row=row_idx, column=4, value=fmt_pay_type(data.get("PayType", ""), L))
        ws.cell(row=row_idx, column=5, value=fmt_status(data.get("Status", ""), L))
        ws.cell(row=row_idx, column=6, value=fmt_epoch_ms(data.get("StartTime", "")))
        ws.cell(row=row_idx, column=7, value=fmt_epoch_ms(data.get("EndTime", "")))
        ws.cell(row=row_idx, column=8, value=json.dumps(data, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(8)].hidden = True


def write_guide(wb, L, lang):
    """Write Guide sheet with embedded documentation."""
    ws = wb.create_sheet(L["sheet_guide"])
    row = 1
    for line in L["guide"]:
        ws.cell(row=row, column=1, value=line)
        row += 1

    # Add scene descriptions
    row += 1
    scene_desc = SCENE_DESCRIPTIONS.get(lang, SCENE_DESCRIPTIONS["en"])
    scene_title = "═══ 防护场景详解 ═══" if lang == "zh" else "═══ Defense Scene Details ═══"
    ws.cell(row=row, column=1, value=scene_title)
    row += 1
    ws.cell(row=row, column=1, value="")
    row += 1
    for ds in DEFENSE_SCENES:
        scene_name = ds["scene"]
        label = L["scene_labels"].get(scene_name, scene_name)
        desc = scene_desc.get(scene_name, "")
        ws.cell(row=row, column=1, value=f"  {label}")
        ws.cell(row=row, column=2, value=desc)
        row += 1

    # Add action definitions
    row += 1
    action_title = "═══ 动作定义 ═══" if lang == "zh" else "═══ Action Definitions ═══"
    ws.cell(row=row, column=1, value=action_title)
    row += 1
    ws.cell(row=row, column=1, value="")
    row += 1
    if lang == "zh":
        actions_desc = [
            ("拦截 (block)", "直接拒绝请求，返回拦截页面"),
            ("观察 (monitor)", "仅记录日志，不实际拦截（用于规则调试期）"),
            ("滑块验证 (captcha)", "要求用户完成滑块验证后放行"),
            ("JS验证 (js)", "通过JavaScript计算验证客户端是否为真实浏览器"),
            ("放行 (allow/continue)", "允许请求通过，不做处理"),
        ]
    else:
        actions_desc = [
            ("block", "Reject request and return block page"),
            ("monitor", "Log only, do not block (used during rule testing)"),
            ("captcha", "Require user to complete slider captcha"),
            ("js", "JavaScript challenge to verify real browser"),
            ("allow/continue", "Allow request to pass without action"),
        ]
    for act_name, act_desc in actions_desc:
        ws.cell(row=row, column=1, value=f"  {act_name}")
        ws.cell(row=row, column=2, value=act_desc)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80


def write_cname_sheet(wb, region_label, domains_detail, L, resolver=None):
    """Write CNAME domain sheet."""
    sheet_name = f"{L['sheet_cname']} ({region_label})"
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["cname"])
    if not domains_detail:
        write_empty_placeholder(ws)
        auto_width(ws)
        if resolver and not resolver.is_delegated:
            ws.column_dimensions["A"].hidden = True
        return
    fallback_uid = resolver.for_caller() if resolver else ""
    for row_idx, dd in enumerate(domains_detail, 2):
        listen = dd.get("Listen", {}) if isinstance(dd.get("Listen"), dict) else {}
        redirect = dd.get("Redirect", {}) if isinstance(dd.get("Redirect"), dict) else {}
        domain_name = dd.get("Domain", "")
        owner_uid = resolver.for_resource(region_label, domain_name) if resolver else fallback_uid
        ws.cell(row=row_idx, column=1, value=owner_uid)
        ws.cell(row=row_idx, column=2, value=domain_name)
        ws.cell(row=row_idx, column=3, value=fmt_status(dd.get("Status", ""), L))
        ws.cell(row=row_idx, column=4, value=fmt_list(listen.get("HttpPorts", [])))
        ws.cell(row=row_idx, column=5, value=fmt_list(listen.get("HttpsPorts", [])))
        ws.cell(row=row_idx, column=6, value=listen.get("TLSVersion", ""))
        ws.cell(row=row_idx, column=7, value=fmt_bool(listen.get("Http2Enabled", "")))
        ws.cell(row=row_idx, column=8, value=fmt_bool(listen.get("IPv6Enabled", "")))
        ws.cell(row=row_idx, column=9, value=fmt_list(redirect.get("BackendList", [])))
        ws.cell(row=row_idx, column=10, value=redirect.get("Loadbalance", ""))
        ws.cell(row=row_idx, column=11, value=listen.get("CertId", ""))
        ws.cell(row=row_idx, column=12, value=listen.get("CertName", ""))
        ws.cell(row=row_idx, column=13, value=redirect.get("ConnectTimeout", ""))
        ws.cell(row=row_idx, column=14, value=redirect.get("ReadTimeout", ""))
        ws.cell(row=row_idx, column=15, value=redirect.get("WriteTimeout", ""))
        ws.cell(row=row_idx, column=16, value=json.dumps(dd, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(16)].hidden = True
    # Hide OwnerUid column (col A) in single-UID scenario (value-redundant)
    if resolver and not resolver.is_delegated:
        ws.column_dimensions["A"].hidden = True


def write_cloud_resource_sheet(wb, region_label, resources, L, resolver=None):
    """Write Cloud Resource sheet."""
    sheet_name = f"{L['sheet_cloud']} ({region_label})"
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["cloud"])
    if not resources:
        write_empty_placeholder(ws)
        auto_width(ws)
        if resolver and not resolver.is_delegated:
            ws.column_dimensions["A"].hidden = True
        return
    fallback_uid = resolver.for_caller() if resolver else ""
    for row_idx, r in enumerate(resources, 2):
        rid = r.get("ResourceInstanceId", "")
        owner_uid = resolver.for_resource(region_label, rid) if resolver else fallback_uid
        ws.cell(row=row_idx, column=1, value=owner_uid)
        ws.cell(row=row_idx, column=2, value=rid)
        ws.cell(row=row_idx, column=3, value=r.get("ResourceProduct", ""))
        ws.cell(row=row_idx, column=4, value=r.get("ResourceRegionId", ""))
        ws.cell(row=row_idx, column=5, value=r.get("ResourceDomain", ""))
        ws.cell(row=row_idx, column=6, value=r.get("ResourceInstanceIp", ""))
        ws.cell(row=row_idx, column=7, value=json.dumps(r, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(7)].hidden = True
    # Hide OwnerUid column (col A) in single-UID scenario
    if resolver and not resolver.is_delegated:
        ws.column_dimensions["A"].hidden = True


def write_hybrid_cloud_sheet(wb, region_label, resources, L, resolver=None):
    """Write Hybrid Cloud sheet."""
    sheet_name = f"{L['sheet_hybrid']} ({region_label})"
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["hybrid"])
    if not resources:
        write_empty_placeholder(ws)
        auto_width(ws)
        if resolver and not resolver.is_delegated:
            ws.column_dimensions["A"].hidden = True
        return
    fallback_uid = resolver.for_caller() if resolver else ""
    for row_idx, r in enumerate(resources, 2):
        rname = r.get("Domain", r.get("ResourceName", ""))
        owner_uid = resolver.for_resource(region_label, rname) if resolver else fallback_uid
        ws.cell(row=row_idx, column=1, value=owner_uid)
        ws.cell(row=row_idx, column=2, value=rname)
        ws.cell(row=row_idx, column=3, value=fmt_list(r.get("Ports", [])))
        ws.cell(row=row_idx, column=4, value=fmt_bool(r.get("HttpsEnabled", "")))
        ws.cell(row=row_idx, column=5, value=fmt_list(r.get("Backends", [])))
        ws.cell(row=row_idx, column=6, value=r.get("LoadBalance", ""))
        ws.cell(row=row_idx, column=7, value=json.dumps(r, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(7)].hidden = True
    # Hide OwnerUid column (col A) in single-UID scenario
    if resolver and not resolver.is_delegated:
        ws.column_dimensions["A"].hidden = True


def write_protected_objects_sheet(wb, region_label, objects, L, resolver=None):
    """Write Protected Objects sheet."""
    sheet_name = f"{L['sheet_objects']} ({region_label})"
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["objects"])
    if not objects:
        write_empty_placeholder(ws)
        auto_width(ws)
        if resolver and not resolver.is_delegated:
            ws.column_dimensions["A"].hidden = True
        return
    fallback_uid = resolver.for_caller() if resolver else ""
    for row_idx, obj in enumerate(objects, 2):
        rname = obj.get("Resource", "")
        # Prefer the API's own OwnerUserId field; fall back to resolver lookup.
        owner_uid = obj.get("OwnerUserId") or (
            resolver.for_resource(region_label, rname) if resolver else fallback_uid)
        ws.cell(row=row_idx, column=1, value=owner_uid)
        ws.cell(row=row_idx, column=2, value=rname)
        ws.cell(row=row_idx, column=3, value=fmt_product(obj.get("Product", ""), L))
        ws.cell(row=row_idx, column=4, value=fmt_pattern(obj.get("Pattern", ""), L))
        ws.cell(row=row_idx, column=5, value=obj.get("OwnerUserId", ""))
        ws.cell(row=row_idx, column=6, value=fmt_epoch_ms(obj.get("GmtCreate", "")))
        ws.cell(row=row_idx, column=7, value=fmt_epoch_ms(obj.get("GmtModified", "")))
        ws.cell(row=row_idx, column=8, value=json.dumps(obj, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(8)].hidden = True
    # Hide OwnerUid column (col A) in single-UID scenario
    if resolver and not resolver.is_delegated:
        ws.column_dimensions["A"].hidden = True


def write_object_groups_sheet(wb, region_label, groups, L, resolver=None):
    """Write Object Groups sheet.

    Object groups are admin-scoped; no per-row OwnerUid column.
    The `resolver` parameter is kept for signature compatibility but unused.
    """
    sheet_name = f"{L['sheet_groups']} ({region_label})"
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["groups"])
    if not groups:
        write_empty_placeholder(ws)
        auto_width(ws)
        return
    for row_idx, g in enumerate(groups, 2):
        ws.cell(row=row_idx, column=1, value=g.get("GroupName", ""))
        ws.cell(row=row_idx, column=2, value=g.get("Description", ""))
        members = g.get("ResourceList", g.get("Members", []))
        ws.cell(row=row_idx, column=3, value=fmt_list(members))
        ws.cell(row=row_idx, column=4, value=json.dumps(g, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(4)].hidden = True


def write_templates_sheet(wb, region_label, templates, L, resolver=None):
    """Write Templates sheet.

    Templates are admin-scoped; no per-row OwnerUid column.
    The `resolver` parameter is kept for signature compatibility but unused.
    """
    sheet_name = f"{L['sheet_templates']} ({region_label})"
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["templates"])
    if not templates:
        write_empty_placeholder(ws)
        auto_width(ws)
        return
    for row_idx, t in enumerate(templates, 2):
        tid = t.get("TemplateId", "")
        ws.cell(row=row_idx, column=1, value=tid)
        ws.cell(row=row_idx, column=2, value=t.get("TemplateName", ""))
        ws.cell(row=row_idx, column=3, value=L["scene_labels"].get(t.get("DefenseScene", ""), t.get("DefenseScene", "")))
        ws.cell(row=row_idx, column=4, value=fmt_origin(t.get("TemplateOrigin", ""), L))
        ws.cell(row=row_idx, column=5, value=fmt_status(t.get("TemplateStatus", ""), L))
        ws.cell(row=row_idx, column=6, value=t.get("DefenseSubScene", ""))
        ws.cell(row=row_idx, column=7, value=json.dumps(t, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(7)].hidden = True


def write_rules_sheet(wb, region_label, scene_rules_map, L, lang, resolver=None):
    """Write a single merged Defense Rules sheet for the region.

    Rules belong to admin-scoped templates; no per-row OwnerUid column.
    The first column is the scene label (filterable). Empty regions still
    produce a sheet with a placeholder row so the audit trail is preserved.
    The `resolver` parameter is kept for signature compatibility but unused.
    """
    sheet_name = f"{L['sheet_rules']} ({region_label})"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["rules"])

    # Flatten all scenes into one ordered list (follow DEFENSE_SCENES order)
    flat_rows = []
    for ds in DEFENSE_SCENES:
        scene_name = ds["scene"]
        scene_label = L["scene_labels"].get(scene_name, scene_name)
        for r in scene_rules_map.get(scene_name, []) or []:
            flat_rows.append((scene_name, scene_label, r))

    if not flat_rows:
        write_empty_placeholder(ws)
        auto_width(ws)
        ws.freeze_panes = "A2"
        return

    for row_idx, (scene_name, scene_label, r) in enumerate(flat_rows, 2):
        tid = r.get("TemplateId", "")
        ws.cell(row=row_idx, column=1, value=scene_label)
        ws.cell(row=row_idx, column=2, value=r.get("RuleId", ""))
        ws.cell(row=row_idx, column=3, value=r.get("RuleName", ""))
        ws.cell(row=row_idx, column=4, value=get_rule_purpose(r.get("RuleName", ""), scene_name, lang))
        ws.cell(row=row_idx, column=5, value=tid)
        ws.cell(row=row_idx, column=6, value=fmt_status(r.get("Status", ""), L))
        ws.cell(row=row_idx, column=7, value=fmt_origin(r.get("DefenseOrigin", ""), L))
        ws.cell(row=row_idx, column=8, value=fmt_epoch_ms(r.get("GmtModified", "")))
        config = r.get("Config", "")
        config_str = config
        if isinstance(config, dict):
            config_str = json.dumps(config, ensure_ascii=False)
        action, conditions = parse_rule_config(config, scene_name)
        ws.cell(row=row_idx, column=9, value=fmt_action(action, L))
        ws.cell(row=row_idx, column=10, value=conditions)
        ws.cell(row=row_idx, column=11, value=config_str)

    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(11)].hidden = True
    # Enable autofilter on the header row and freeze first row for navigation
    last_col = get_column_letter(len(L["headers"]["rules"]))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    ws.freeze_panes = "A2"


def write_bindings_sheet(wb, region_label, bindings, template_map, L, resolver=None):
    """Write Bindings sheet with template name cross-reference."""
    sheet_name = f"{L['sheet_bindings']} ({region_label})"
    ws = wb.create_sheet(sheet_name)
    write_header(ws, L["headers"]["bindings"])
    if not bindings:
        write_empty_placeholder(ws)
        auto_width(ws)
        if resolver and not resolver.is_delegated:
            ws.column_dimensions["A"].hidden = True
        return
    fallback_uid = resolver.for_caller() if resolver else ""
    tmap = template_map or {}
    for row_idx, b in enumerate(bindings, 2):
        tid = b.get("TemplateId", "")
        rname = b.get("Resource", "")
        owner_uid = resolver.for_resource(region_label, rname) if resolver else fallback_uid
        tinfo = tmap.get(tid, {})
        ws.cell(row=row_idx, column=1, value=owner_uid)
        ws.cell(row=row_idx, column=2, value=tid)
        ws.cell(row=row_idx, column=3, value=tinfo.get("name", ""))
        ws.cell(row=row_idx, column=4, value=L["scene_labels"].get(tinfo.get("scene", ""), tinfo.get("scene", "")))
        ws.cell(row=row_idx, column=5, value=rname)
        ws.cell(row=row_idx, column=6, value=fmt_product(b.get("Product", ""), L))
        ws.cell(row=row_idx, column=7, value=fmt_pattern(b.get("Pattern", ""), L))
        ws.cell(row=row_idx, column=8, value=json.dumps(b, ensure_ascii=False))
    auto_width(ws)
    # Hide raw JSON column
    ws.column_dimensions[get_column_letter(8)].hidden = True
    # Hide OwnerUid column (col A) in single-UID scenario
    if resolver and not resolver.is_delegated:
        ws.column_dimensions["A"].hidden = True


def write_audit_view(wb, region_data_list, L, lang):
    """Write Audit View sheet - coverage matrix of objects x scenes."""
    ws = wb.create_sheet(L["sheet_audit"])
    scene_labels = [L["scene_labels"][ds["scene"]] for ds in DEFENSE_SCENES]
    headers = ["Region", "Protected Object", "Product"] + scene_labels + ["Bound Templates", "Total Rules"]
    if lang == "zh":
        headers = ["区域", "防护对象", "产品"] + scene_labels + ["已绑模板数", "生效规则数"]
    write_header(ws, headers)

    row = 2
    for rdata in region_data_list:
        rl = rdata["region_label"]
        objects = rdata["objects"]
        bindings = rdata["all_bindings"]
        template_map = rdata["template_map"]
        scene_rules_map = rdata["scene_rules_map"]

        # Build: resource -> set of bound scenes
        resource_scenes = {}  # resource_name -> {scene1, scene2, ...}
        resource_templates = {}  # resource_name -> count of templates
        for b in bindings:
            res = b.get("Resource", "")
            tid = b.get("TemplateId", "")
            tinfo = template_map.get(tid, {})
            scene = tinfo.get("scene", "")
            if res:
                resource_scenes.setdefault(res, set()).add(scene)
                resource_templates.setdefault(res, set()).add(tid)

        # Build: resource -> rule count
        resource_rules = {}  # resource_name -> total rule count
        for scene_name, rules in scene_rules_map.items():
            for r in rules:
                tid = r.get("TemplateId", "")
                # Find which resources this template is bound to
                for b in bindings:
                    if b.get("TemplateId") == tid:
                        res = b.get("Resource", "")
                        resource_rules[res] = resource_rules.get(res, 0) + 1

        for obj in objects:
            res_name = obj.get("Resource", "")
            ws.cell(row=row, column=1, value=rl)
            ws.cell(row=row, column=2, value=res_name)
            ws.cell(row=row, column=3, value=fmt_product(obj.get("Product", ""), L))
            bound_scenes = resource_scenes.get(res_name, set())
            for col_idx, ds in enumerate(DEFENSE_SCENES):
                has_coverage = ds["scene"] in bound_scenes
                cell = ws.cell(row=row, column=4 + col_idx, value="\u2713" if has_coverage else "\u2717")
                if not has_coverage:
                    cell.font = Font(color="CC0000")
                else:
                    cell.font = Font(color="006600")
            ws.cell(row=row, column=4 + len(DEFENSE_SCENES), value=len(resource_templates.get(res_name, set())))
            ws.cell(row=row, column=5 + len(DEFENSE_SCENES), value=resource_rules.get(res_name, 0))
            row += 1

    auto_width(ws)


def write_reverse_view(wb, region_data_list, L, lang):
    """Write reverse view: per protected object, list all effective rules."""
    ws = wb.create_sheet(L["sheet_reverse"])
    if lang == "zh":
        headers = ["区域", "防护对象", "产品", "防护场景", "模板名称",
                   "规则ID", "规则名称", "规则用途", "动作", "状态"]
    else:
        headers = ["Region", "Protected Object", "Product", "Defense Scene", "Template Name",
                   "RuleId", "Rule Name", "Purpose", "Action", "Status"]
    write_header(ws, headers)

    row = 2
    for rdata in region_data_list:
        rl = rdata["region_label"]
        objects = rdata["objects"]
        bindings = rdata["all_bindings"]
        template_map = rdata["template_map"]
        scene_rules_map = rdata["scene_rules_map"]

        # Build: template_id -> [rules]
        tid_rules = {}
        for scene_name, rules in scene_rules_map.items():
            for r in rules:
                tid = r.get("TemplateId", "")
                tid_rules.setdefault(tid, []).append((scene_name, r))

        # Build: resource -> list of bound template_ids
        resource_tids = {}
        for b in bindings:
            res = b.get("Resource", "")
            tid = b.get("TemplateId", "")
            if res and tid:
                resource_tids.setdefault(res, set()).add(tid)

        for obj in objects:
            res_name = obj.get("Resource", "")
            product = fmt_product(obj.get("Product", ""), L)
            bound_tids = resource_tids.get(res_name, set())

            if not bound_tids:
                # Object has no bindings
                ws.cell(row=row, column=1, value=rl)
                ws.cell(row=row, column=2, value=res_name)
                ws.cell(row=row, column=3, value=product)
                no_rule = "无绑定规则" if lang == "zh" else "No rules bound"
                ws.cell(row=row, column=4, value=no_rule)
                row += 1
                continue

            for tid in sorted(bound_tids):
                tinfo = template_map.get(tid, {})
                tname = tinfo.get("name", "")
                rules_for_tid = tid_rules.get(tid, [])
                if not rules_for_tid:
                    ws.cell(row=row, column=1, value=rl)
                    ws.cell(row=row, column=2, value=res_name)
                    ws.cell(row=row, column=3, value=product)
                    ws.cell(row=row, column=4, value=L["scene_labels"].get(tinfo.get("scene", ""), tinfo.get("scene", "")))
                    ws.cell(row=row, column=5, value=tname)
                    no_rule = "模板无规则" if lang == "zh" else "Template has no rules"
                    ws.cell(row=row, column=6, value=no_rule)
                    row += 1
                else:
                    for scene_name, r in rules_for_tid:
                        ws.cell(row=row, column=1, value=rl)
                        ws.cell(row=row, column=2, value=res_name)
                        ws.cell(row=row, column=3, value=product)
                        ws.cell(row=row, column=4, value=L["scene_labels"].get(scene_name, scene_name))
                        ws.cell(row=row, column=5, value=tname)
                        ws.cell(row=row, column=6, value=r.get("RuleId", ""))
                        ws.cell(row=row, column=7, value=r.get("RuleName", ""))
                        ws.cell(row=row, column=8, value=get_rule_purpose(r.get("RuleName", ""), scene_name, lang))
                        config = r.get("Config", "")
                        action, _ = parse_rule_config(config, scene_name)
                        ws.cell(row=row, column=9, value=fmt_action(action, L))
                        ws.cell(row=row, column=10, value=fmt_status(r.get("Status", ""), L))
                        row += 1

    auto_width(ws)


# === Main ===

def collect_data(profile):
    """Collect all WAF data from API (region-independent of output language).

    Also probes account topology:
      - Single UID  : caller is a normal account.
      - Delegated   : caller is the WAF delegated administrator managing
                      multiple member accounts.
    """
    # === Phase 0: Probe caller identity ===
    print("\nProbing account topology...")
    caller_uid = fetch_caller_identity(profile)
    print(f"  Caller UID: {caller_uid or 'unknown'}")

    is_delegated = False
    member_accounts = []
    probe_done = False

    instances = []
    region_data_list = []  # [{region_label, instance_id, domains_detail, ...}]
    region_summary = []
    found_any = False

    for region in REGIONS:
        region_id = region["id"]
        region_label = region["label"]
        separator = '\u2500' * 40
        print(f"\n{separator}")
        print(f"Region: {region['desc']} ({region_id})")
        print(separator)

        inst_data = discover_instance(region_id, profile)
        if not inst_data:
            print(f"  \u2717 No WAF instance in {region_id}")
            region_summary.append({"region_label": region_label, "instance_id": "N/A"})
            continue

        found_any = True
        instance_id = inst_data["InstanceId"]
        instances.append({"region": region_label, "data": inst_data})

        # Probe delegated-admin status once, on the first available instance
        if not probe_done:
            probe_done = True
            is_delegated = fetch_delegated_status(region_id, instance_id, profile)
            if is_delegated:
                member_accounts = fetch_member_accounts(region_id, instance_id, profile)
                print(f"  \u2713 Delegated administrator detected"
                      f" ({len(member_accounts)} member accounts)")
            else:
                print(f"  \u2713 Single-UID scenario")

        rd = {
            "region_label": region_label,
            "instance_id": instance_id,
            "domain_count": 0, "cloud_count": 0, "hybrid_count": 0,
            "object_count": 0, "group_count": 0,
            "template_count": 0, "rule_count": 0, "binding_count": 0,
        }

        # CNAME domains
        domains = fetch_domains(region_id, instance_id, profile)
        domains_detail = []
        if domains:
            print(f"  Fetching domain details...")
            for d in domains:
                domain_name = d.get("Domain", "")
                if domain_name:
                    detail = fetch_domain_detail(region_id, instance_id, domain_name, profile)
                    if detail:
                        domains_detail.append(detail)
            print(f"    \u2713 {len(domains_detail)} domain details fetched")
        rd["domain_count"] = len(domains_detail)

        # Cloud resources
        cloud_resources, _ = fetch_cloud_resources(region_id, instance_id, profile)
        rd["cloud_count"] = len(cloud_resources)

        # Hybrid cloud
        hybrid_resources, _ = fetch_hybrid_cloud_resources(region_id, instance_id, profile)
        rd["hybrid_count"] = len(hybrid_resources)

        # Protected objects
        print(f"  Fetching protected objects...")
        objects = fetch_defense_resources(region_id, instance_id, profile)
        print(f"    \u2713 {len(objects)} protected objects")
        rd["object_count"] = len(objects)

        # Object groups
        groups = fetch_defense_resource_groups(region_id, instance_id, profile)
        rd["group_count"] = len(groups)

        # Defense templates & rules
        print(f"  Fetching defense templates & rules...")
        all_templates = []
        all_bindings = []
        template_map = {}
        scene_rules_map = {}  # scene_name -> [rules]

        for ds in DEFENSE_SCENES:
            templates = fetch_defense_templates(region_id, instance_id, ds["scene"], profile)
            all_templates.extend(templates)
            for t in templates:
                tid = t.get("TemplateId")
                if tid:
                    template_map[tid] = {
                        "name": t.get("TemplateName", ""),
                        "scene": ds["scene"],
                    }

            scene_rules = []
            for t in templates:
                tid = t.get("TemplateId")
                if not tid:
                    continue
                rules = fetch_defense_rules(
                    region_id, instance_id, ds["scene"], ds["rule_type"], tid, profile
                )
                for r in rules:
                    r["TemplateId"] = tid
                    r["DefenseScene"] = ds["scene"]
                scene_rules.extend(rules)

                bindings = fetch_defense_resources(region_id, instance_id, profile, template_id=tid)
                for b in bindings:
                    b["TemplateId"] = tid
                all_bindings.extend(bindings)

            scene_rules_map[ds["scene"]] = scene_rules

        all_rules_count = sum(len(v) for v in scene_rules_map.values())
        print(f"    \u2713 {len(all_templates)} templates, {all_rules_count} rules, {len(all_bindings)} bindings")
        rd["template_count"] = len(all_templates)
        rd["rule_count"] = all_rules_count
        rd["binding_count"] = len(all_bindings)

        region_data_list.append({
            "region_label": region_label,
            "region_id": region_id,
            "instance_id": instance_id,
            "domains_detail": domains_detail,
            "cloud_resources": cloud_resources,
            "hybrid_resources": hybrid_resources,
            "objects": objects,
            "groups": groups,
            "all_templates": all_templates,
            "all_bindings": all_bindings,
            "template_map": template_map,
            "scene_rules_map": scene_rules_map,
        })
        region_summary.append(rd)

    if not found_any:
        print("\n\u2717 ERROR: No WAF instance found in any region.")
        sys.exit(1)

    # === Phase 2: Build ResourceName -> OwnerUid map (delegated only) ===
    # Only member-scoped sheets (CNAME/Cloud/Hybrid/Objects/Bindings) need this.
    region_owner_maps = {}

    if is_delegated:
        print("\nResolving resource ownership for delegated administrator...")
        for rdata in region_data_list:
            rl = rdata["region_label"]
            region_id_lookup = rdata["region_id"]
            instance_id_lookup = rdata["instance_id"]

            # Collect all ResourceNames seen in this region
            resource_names = set()
            for obj in rdata["objects"]:
                if obj.get("Resource"):
                    resource_names.add(obj["Resource"])
            for b in rdata["all_bindings"]:
                if b.get("Resource"):
                    resource_names.add(b["Resource"])

            owner_map = {}
            if resource_names:
                owner_map = fetch_resource_owner_uid(
                    region_id_lookup, instance_id_lookup,
                    list(resource_names), profile)
                print(f"  [{rl}] {len(owner_map)}/{len(resource_names)} resources resolved")
            region_owner_maps[rl] = owner_map

    resolver = OwnerUidResolver(
        caller_uid, is_delegated,
        region_owner_maps=region_owner_maps,
    )

    return {
        "instances": instances,
        "region_data_list": region_data_list,
        "region_summary": region_summary,
        "account_topology": {
            "caller_uid": caller_uid,
            "is_delegated": is_delegated,
            "member_accounts": member_accounts,
        },
        "resolver": resolver,
    }


def auto_discover_prev(caller_uid, lang, output_dir, exclude_path):
    """Scan output_dir for the most recent matching backup as diff baseline.

    Pattern: waf_config_backup_<caller_uid>_<YYYYMMDD_HHMMSS>(_<lang>)?.xlsx
    Files matching the OTHER lang are filtered out (e.g. lang='zh' skips '_en.xlsx').
    The current run's output (exclude_path) is filtered out to avoid self-diff.
    Returns absolute path of the most recent candidate, or None.
    """
    if not caller_uid or caller_uid == "unknown":
        return None
    output_dir = output_dir or "."
    pattern = os.path.join(output_dir, f"waf_config_backup_{caller_uid}_*.xlsx")
    name_re = re.compile(
        r"^waf_config_backup_(\d+|unknown)_(\d{8}_\d{6})(?:_(zh|en))?\.xlsx$"
    )
    candidates = []
    exclude_abs = os.path.abspath(exclude_path) if exclude_path else None
    for fp in glob.glob(pattern):
        if exclude_abs and os.path.abspath(fp) == exclude_abs:
            continue
        m = name_re.match(os.path.basename(fp))
        if not m:
            continue
        file_lang = m.group(3)
        if file_lang and file_lang != lang:
            continue
        try:
            candidates.append((os.path.getmtime(fp), fp))
        except OSError:
            pass
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return os.path.abspath(candidates[0][1])


def write_diff_sheet(wb, prev_path, region_data_list, L, lang):
    """Compare with a previous backup Excel and write diff sheet."""
    try:
        from openpyxl import load_workbook as load_wb
        prev_wb = load_wb(prev_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  \u26a0 Cannot load previous backup for diff: {e}")
        return

    sheet_title = "\u53d8\u66f4\u5bf9\u6bd4" if lang == "zh" else "Changes (Diff)"
    ws = wb.create_sheet(sheet_title)
    if lang == "zh":
        headers = ["\u53d8\u66f4\u7c7b\u578b", "\u533a\u57df", "\u573a\u666f", "\u89c4\u5219ID", "\u89c4\u5219\u540d\u79f0", "\u8be6\u60c5"]
    else:
        headers = ["Change Type", "Region", "Scene", "RuleId", "Rule Name", "Details"]
    write_header(ws, headers)

    # Collect current rule details per region (name, action, status, conditions)
    current_rules = {}  # (region, scene, rule_id) -> dict
    for rdata in region_data_list:
        rl = rdata["region_label"]
        for scene_name, rules in rdata["scene_rules_map"].items():
            for r in rules:
                rid = str(r.get("RuleId", ""))
                config = r.get("Config", "")
                action_raw, conditions = parse_rule_config(config, scene_name)
                current_rules[(rl, scene_name, rid)] = {
                    "name": r.get("RuleName", "") or "",
                    "status": fmt_status(r.get("Status", ""), L) or "",
                    "action": fmt_action(action_raw, L) if action_raw else "",
                    "conditions": conditions or "",
                }

    # Collect previous rule IDs from the prev workbook (supports both layouts)
    #   - Legacy layout: one sheet per scene; first column is RuleId; scene parsed from sheet name.
    #   - New layout: single merged "Defense Rules" sheet per region; first column is DefenseScene label.
    prev_rules = {}  # (region, scene_key, rule_id) -> dict (same shape as current_rules)
    
    def _scene_key_from_label(label):
        for ds in DEFENSE_SCENES:
            for l_key in ("en", "zh"):
                if LABELS[l_key]["scene_labels"].get(ds["scene"]) == label:
                    return ds["scene"]
        return ""
    
    for sheet_name in prev_wb.sheetnames:
        sheet = prev_wb[sheet_name]
        first_cell = sheet.cell(row=1, column=1).value
    
        # Parse region from sheet name suffix "... (region)"
        region = ""
        if "(" in sheet_name and ")" in sheet_name:
            region = sheet_name.split("(")[-1].rstrip(")")
    
        if first_cell in ("DefenseScene", "\u9632\u62a4\u573a\u666f"):
            # New merged-rules layout: col0=scene label, col1=RuleId, col2=RuleName,
            # col5=Status, col8=Action, col9=Conditions Summary
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                scene_label = row[0]
                rid = str(row[1]) if len(row) > 1 and row[1] is not None else ""
                rname = row[2] if len(row) > 2 else ""
                if not rid or rid.startswith("("):  # skip placeholder rows
                    continue
                scene_key = _scene_key_from_label(scene_label)
                prev_rules[(region, scene_key, rid)] = {
                    "name": (rname or ""),
                    "status": (row[5] if len(row) > 5 and row[5] is not None else "") or "",
                    "action": (row[8] if len(row) > 8 and row[8] is not None else "") or "",
                    "conditions": (row[9] if len(row) > 9 and row[9] is not None else "") or "",
                }
        elif first_cell in ("RuleId", "\u89c4\u5219ID"):
            # Legacy per-scene layout: col0=RuleId, col1=RuleName; scene from sheet name.
            # Legacy backups lack structured Action/Status/Conditions columns,
            # so only rule add/remove diffs are produced for legacy baselines.
            scene_part = sheet_name.split(" (")[0] if " (" in sheet_name else sheet_name
            scene_key = _scene_key_from_label(scene_part)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    rid = str(row[0])
                    rname = row[1] if len(row) > 1 else ""
                    prev_rules[(region, scene_key, rid)] = {
                        "name": (rname or ""),
                        "status": "",
                        "action": "",
                        "conditions": "",
                    }

    prev_wb.close()

    # Compute diff
    row_idx = 2
    added_label = "\u65b0\u589e" if lang == "zh" else "Added"
    removed_label = "\u5220\u9664" if lang == "zh" else "Removed"
    modified_label = "\u4fee\u6539" if lang == "zh" else "Modified"

    # New rules
    for key, info in sorted(current_rules.items()):
        if key not in prev_rules:
            rl, scene, rid = key
            scene_label = L["scene_labels"].get(scene, scene)
            ws.cell(row=row_idx, column=1, value=added_label)
            ws.cell(row=row_idx, column=2, value=rl)
            ws.cell(row=row_idx, column=3, value=scene_label)
            ws.cell(row=row_idx, column=4, value=rid)
            ws.cell(row=row_idx, column=5, value=info["name"])
            detail = "\u65b0\u589e\u89c4\u5219" if lang == "zh" else "New rule added"
            ws.cell(row=row_idx, column=6, value=detail)
            ws.cell(row=row_idx, column=1).font = Font(color="006600")
            row_idx += 1

    # Removed rules
    for key, info in sorted(prev_rules.items()):
        if key not in current_rules:
            rl, scene, rid = key
            scene_label = L["scene_labels"].get(scene, scene)
            ws.cell(row=row_idx, column=1, value=removed_label)
            ws.cell(row=row_idx, column=2, value=rl)
            ws.cell(row=row_idx, column=3, value=scene_label)
            ws.cell(row=row_idx, column=4, value=rid)
            ws.cell(row=row_idx, column=5, value=info["name"])
            detail = "\u89c4\u5219\u5df2\u79fb\u9664" if lang == "zh" else "Rule removed"
            ws.cell(row=row_idx, column=6, value=detail)
            ws.cell(row=row_idx, column=1).font = Font(color="CC0000")
            row_idx += 1

    # Modified rules (same RuleId but content changed: name / status / action / conditions)
    field_labels_zh = {"name": "\u89c4\u5219\u540d\u79f0", "status": "\u72b6\u6001",
                       "action": "\u52a8\u4f5c", "conditions": "\u6761\u4ef6\u6458\u8981"}
    field_labels_en = {"name": "Name", "status": "Status",
                       "action": "Action", "conditions": "Conditions"}
    field_labels = field_labels_zh if lang == "zh" else field_labels_en
    for key in sorted(set(current_rules.keys()) & set(prev_rules.keys())):
        cur = current_rules[key]
        prv = prev_rules[key]
        diffs = []
        for fld in ("name", "status", "action", "conditions"):
            cur_v = (cur.get(fld) or "").strip()
            prv_v = (prv.get(fld) or "").strip()
            # Legacy baselines have empty action/status/conditions; skip those
            # to avoid noisy false-positives like "" -> "block" on every rule.
            if fld in ("status", "action", "conditions") and not prv_v:
                continue
            if cur_v != prv_v:
                diffs.append(f"{field_labels[fld]}: {prv_v!r} \u2192 {cur_v!r}")
        if diffs:
            rl, scene, rid = key
            scene_label = L["scene_labels"].get(scene, scene)
            ws.cell(row=row_idx, column=1, value=modified_label)
            ws.cell(row=row_idx, column=2, value=rl)
            ws.cell(row=row_idx, column=3, value=scene_label)
            ws.cell(row=row_idx, column=4, value=rid)
            ws.cell(row=row_idx, column=5, value=cur["name"])
            ws.cell(row=row_idx, column=6, value="; ".join(diffs))
            ws.cell(row=row_idx, column=1).font = Font(color="CC8800")
            row_idx += 1

    if row_idx == 2:
        no_change = "\u65e0\u53d8\u66f4 - \u89c4\u5219\u914d\u7f6e\u4e0e\u4e0a\u6b21\u5907\u4efd\u4e00\u81f4" if lang == "zh" else "No changes - rules match previous backup"
        ws.cell(row=2, column=1, value="\u2705")
        ws.cell(row=2, column=2, value=no_change)

    auto_width(ws)


def generate_workbook(collected, profile, lang, prev_path=None, diff_disabled=False):
    """Generate an Excel workbook from collected data in the specified language."""
    L = LABELS[lang]
    wb = Workbook()

    instances = collected["instances"]
    region_data_list = collected["region_data_list"]
    region_summary = collected["region_summary"]
    topology = collected.get("account_topology", {})
    resolver = collected.get("resolver")

    # Write per-region data sheets
    for rdata in region_data_list:
        rl = rdata["region_label"]
        write_cname_sheet(wb, rl, rdata["domains_detail"], L, resolver=resolver)
        write_cloud_resource_sheet(wb, rl, rdata["cloud_resources"], L, resolver=resolver)
        write_hybrid_cloud_sheet(wb, rl, rdata["hybrid_resources"], L, resolver=resolver)
        write_protected_objects_sheet(wb, rl, rdata["objects"], L, resolver=resolver)
        write_object_groups_sheet(wb, rl, rdata["groups"], L, resolver=resolver)

        # Single merged Defense Rules sheet per region (all scenes combined)
        write_rules_sheet(wb, rl, rdata["scene_rules_map"], L, lang, resolver=resolver)

        write_templates_sheet(wb, rl, rdata["all_templates"], L, resolver=resolver)
        write_bindings_sheet(wb, rl, rdata["all_bindings"], rdata["template_map"], L, resolver=resolver)

    # Write audit/reverse view sheets
    write_audit_view(wb, region_data_list, L, lang)
    write_reverse_view(wb, region_data_list, L, lang)

    # Write diff sheet if previous backup provided
    if prev_path:
        write_diff_sheet(wb, prev_path, region_data_list, L, lang)

    # Write summary, instance info, member accounts (delegated only), guide
    time_display = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    member_accounts = topology.get("member_accounts", []) or []
    backup_meta = {
        "time_display": time_display,
        "profile": profile or "default",
        "caller_uid": topology.get("caller_uid"),
        "is_delegated": topology.get("is_delegated", False),
        "member_count": len(member_accounts),
        "diff_baseline": os.path.basename(prev_path) if prev_path else None,
        "diff_disabled": diff_disabled,
    }
    write_summary(wb, backup_meta, region_summary, L, lang=lang)
    write_instance_info(wb, instances, L, resolver=resolver)
    if topology.get("is_delegated") and member_accounts:
        write_member_accounts_sheet(wb, member_accounts, L)
    write_guide(wb, L, lang)

    # Reorder sheets: Summary, Instance, Members, Guide, Audit, Reverse first
    front_sheets = [L["sheet_summary"], L["sheet_instance"], L["sheet_members"],
                    L["sheet_guide"], L["sheet_audit"], L["sheet_reverse"]]
    existing_front = [s for s in front_sheets if s in wb.sheetnames]
    other_sheets = [s for s in wb.sheetnames if s not in front_sheets]
    desired_order = existing_front + other_sheets
    for i, name in enumerate(desired_order):
        current_idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - current_idx)

    return wb


def main():
    parser = argparse.ArgumentParser(description="WAF 3.0 Full Configuration Backup")
    parser.add_argument("--profile", help="Aliyun CLI profile name", default=None)
    parser.add_argument("--output", help="Output Excel file path (without extension for --lang all)", default=None)
    parser.add_argument("--lang", choices=["zh", "en", "all"], default="all",
                        help="Output language: zh (Chinese), en (English), all (both)")
    parser.add_argument("--prev",
                        help="Explicit path to a previous backup Excel for diff comparison "
                             "(overrides auto-discovery).",
                        default=None)
    parser.add_argument("--no-diff", action="store_true",
                        help="Disable diff comparison entirely "
                             "(skip auto-discovery and ignore --prev).")
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sep = '=' * 60

    print(sep)
    print(f"  WAF 3.0 Full Configuration Backup")
    print(f"  Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Profile: {args.profile or 'default'}")
    print(f"  Language: {args.lang}")
    print(sep)

    # Phase 1: Collect data (only once)
    collected = collect_data(args.profile)

    # Build filename UID prefix from caller identity
    topology = collected.get("account_topology", {})
    caller_uid = topology.get("caller_uid") or "unknown"
    is_delegated = topology.get("is_delegated", False)

    # Phase 2: Generate Excel(s)
    langs = ["zh", "en"] if args.lang == "all" else [args.lang]
    output_files = []

    for lang in langs:
        if args.output:
            if args.lang == "all":
                base = args.output.replace(".xlsx", "")
                output_path = f"{base}_{lang}.xlsx"
            else:
                output_path = args.output if args.output.endswith(".xlsx") else f"{args.output}.xlsx"
        else:
            if args.lang == "all":
                output_path = f"waf_config_backup_{caller_uid}_{timestamp}_{lang}.xlsx"
            else:
                output_path = f"waf_config_backup_{caller_uid}_{timestamp}.xlsx"

        # Resolve diff baseline for this lang
        if args.no_diff:
            prev_for_lang = None
        elif args.prev:
            prev_for_lang = args.prev
        else:
            output_dir = os.path.dirname(os.path.abspath(output_path)) or "."
            prev_for_lang = auto_discover_prev(caller_uid, lang, output_dir, output_path)

        if args.no_diff:
            print(f"  [{lang.upper()}] Diff disabled via --no-diff")
        elif prev_for_lang:
            tag = "explicit" if args.prev else "auto"
            print(f"  [{lang.upper()}] Diff vs {os.path.basename(prev_for_lang)} ({tag})")
        else:
            print(f"  [{lang.upper()}] Baseline mode: no previous backup found")

        wb = generate_workbook(collected, args.profile, lang,
                               prev_path=prev_for_lang,
                               diff_disabled=args.no_diff)
        wb.save(output_path)
        output_files.append(output_path)
        print(f"  \u2713 [{lang.upper()}] {output_path}")

    print(f"\n{sep}")
    print("\u2713 Backup complete!")
    scenario_msg = (f"Delegated Administrator (UID {caller_uid}, "
                    f"{len(topology.get('member_accounts', []) or [])} member accounts)"
                    if is_delegated else f"Single UID ({caller_uid})")
    print(f"  Scenario: {scenario_msg}")
    total_domains = sum(r.get("domain_count", 0) for r in collected["region_summary"])
    total_rules = sum(r.get("rule_count", 0) for r in collected["region_summary"])
    total_templates = sum(r.get("template_count", 0) for r in collected["region_summary"])
    print(f"  Instances: {len(collected['instances'])} | Domains: {total_domains} | "
          f"Rules: {total_rules} | Templates: {total_templates}")
    for f in output_files:
        print(f"  Output: {f}")
    print(sep)


if __name__ == "__main__":
    main()
