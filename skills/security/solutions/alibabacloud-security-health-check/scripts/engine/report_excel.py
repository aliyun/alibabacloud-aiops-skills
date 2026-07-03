#!/usr/bin/env python3
"""
report_excel.py - Generate a 10-column remediation checklist Excel file
Columns: Product / Check Item / Category / Current Status / Severity / Effort / Steps / Doc Link / Owner / Target Date / Implementation Status
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PRODUCT_LABEL = {
    "waf": "WAF 3.0",
    "sas": "Security Center (SAS)",
    "cfw": "Cloud Firewall (CFW)",
    "ddos": "DDoS Protection",
}
VERDICT_LABEL = {
    "full": "Passed", "partial": "Partial",
    "fail": "Failed", "error": "Collection Error",
}
SEVERITY_LABEL = {"high": "High", "medium": "Medium", "low": "Low"}
EFFORT_LABEL = {"low": "Low (<=1 day)", "medium": "Medium (1-3 days)", "high": "High (>3 days)"}

SEVERITY_RANK = {"high": 3, "medium": 2, "low": 1}
EFFORT_RANK   = {"low": 3, "medium": 2, "high": 1}

# Colors (Excel fill, hex without #)
COLORS = {
    "header_bg": "1F4E78",
    "header_fg": "FFFFFF",
    "full":      "C6EFCE",
    "partial":   "FFEB9C",
    "fail":      "FFC7CE",
    "error":     "D9D9D9",
    "section_bg": "D9E1F2",
}

HEADERS = [
    ("Product",           12),
    ("Check Item",        30),
    ("Category",          12),
    ("Current Status",    18),
    ("Severity",          10),
    ("Effort",            14),
    ("Remediation Steps", 50),
    ("Doc Link",          18),
    ("Owner",             14),
    ("Target Date",       14),
    ("Implementation Status", 14),
]


def main():
    parser = argparse.ArgumentParser(description="Generate remediation checklist Excel")
    parser.add_argument("--scores", default="output/scores.json")
    parser.add_argument("--output", default="output/remediation.xlsx")
    parser.add_argument("--customer", default="Example Customer ABC")
    args = parser.parse_args()

    scores = json.loads(Path(args.scores).read_text(encoding="utf-8"))

    wb = Workbook()

    # ====== Sheet 1: Remediation Checklist ======
    ws = wb.active
    ws.title = "Remediation Checklist"

    # Title
    ws["A1"] = f"Cloud Security Product Configuration Health Check - Remediation Checklist (Customer: {args.customer})"
    ws.merge_cells("A1:K1")
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(bold=True, color=COLORS["header_fg"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 32

    # Sort: fail first, then by severity & weight
    def sort_key(r):
        verd = 0 if r["verdict"] == "fail" else (1 if r["verdict"] == "partial" else 2)
        sev = -SEVERITY_RANK.get(r["severity"], 1)
        return (verd, sev, -r["weight"])

    rows = sorted(scores["all_results"], key=sort_key)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_idx = 3
    for r in rows:
        rem = r.get("remediation", {})
        actual = r["actual_value"]
        if isinstance(actual, float):
            actual = f"{actual:.2f}"
        else:
            actual = str(actual)

        data = [
            PRODUCT_LABEL.get(r["product"], r["product"]),
            r["name"],
            r["category"],
            actual,
            SEVERITY_LABEL.get(r["severity"], "—"),
            EFFORT_LABEL.get(rem.get("effort", "medium"), "Medium"),
            (rem.get("steps") or "").strip(),
            rem.get("doc", ""),
            "",  # Owner - to be filled by customer
            "",  # Target Date
            VERDICT_LABEL.get(r["verdict"], "Pending"),
        ]
        for col_idx, val in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col_idx == 8 and val:  # Doc Link
                cell.value = "View Doc"
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            if col_idx == 11:  # Implementation Status -> color by verdict
                cell.fill = PatternFill("solid", fgColor=COLORS.get(r["verdict"], "FFFFFF"))
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust row height for remediation steps (rough estimate by line count)
        line_count = max(1, (data[6] or "").count("\n") + 1)
        ws.row_dimensions[row_idx].height = max(45, line_count * 16)
        row_idx += 1

    # Freeze header
    ws.freeze_panes = "A3"

    # ====== Sheet 2: Score Summary ======
    ws2 = wb.create_sheet("Score Summary")
    ws2["A1"] = f"Score Summary - {args.customer}"
    ws2.merge_cells("A1:E1")
    ws2["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    headers2 = ["Product", "Score", "Passed", "Partial", "Failed"]
    for i, h in enumerate(headers2, start=1):
        c = ws2.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color=COLORS["header_fg"])
        c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(i)].width = 16

    row = 3
    overall_total = 0
    for pid, p in scores["products"].items():
        checks = [x for x in scores["all_results"] if x["product"] == pid]
        full = sum(1 for x in checks if x["verdict"] == "full")
        part = sum(1 for x in checks if x["verdict"] == "partial")
        fail = sum(1 for x in checks if x["verdict"] == "fail")
        for i, v in enumerate([p["name"], p["score"], full, part, fail], start=1):
            cell = ws2.cell(row=row, column=i, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        overall_total += p["score"]
        row += 1

    ws2.cell(row=row, column=1, value="Overall Average").font = Font(bold=True)
    avg = round(overall_total / max(1, len(scores["products"])), 1)
    ws2.cell(row=row, column=2, value=avg).font = Font(bold=True, color="1F4E78")

    # ====== Save ======
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[Excel] Generated: {out_path}")
    print(f"        Remediation checklist: {len(rows)} rows, Score summary: {row - 2} rows")


if __name__ == "__main__":
    main()
