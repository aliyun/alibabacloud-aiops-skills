# -*- coding: utf-8 -*-
"""
Quick BI 小Q解读：Excel 文件解析 / 仪表板快照 + 数据解读流式输出。

支持三种数据来源：
  1. Excel 文件      --excel-file '/path/to/data.xlsx'
  2. 仪表板/门户 URL --dashboard-url 'https://bi.aliyun.com/dashboard/view/pc.htm?pageId=xxx'
                                    'https://bi.aliyun.com/product/view.htm?productId=xxx&menuId=yyy'
  3. 直接传 worksId  --works-id 'pageId'   （等价于已知 pageId 时跳过 URL 解析）

三种参数互斥，必须指定其中一个。
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

from common.utils import (
    read_config,
    require_user_id,
    request_openapi,
    request_openapi_stream,
    parse_sse_event,
    set_workspace_dir,
)
from common.config_loader import check_known_error_code
from common.messages import msg, set_locale
from dashboard.quickbi_openapi import (
    extract_page_id,
    extract_dataportal_ids,
    get_dataportal_page_id,
    is_dataportal_url,
)

SNAPSHOT_URI = "/openapi/v2/snapshot/calling/shot"
INTERPRETATION_URI = "/openapi/v2/smartq/dataInterpretationStream"

POLL_INTERVAL_SECONDS = 3
MAX_POLL_COUNT = 60
MAX_MARKDOWN_CHARS = 100000
MAX_EXCEL_FILE_SIZE = 5 * 1024 * 1024  # 5MB

ACC_PROMPT = """【关键要求】：若输入数据中未包含用户提问的相关内容，则直接回复“不存在”，禁止随意捏造数据。若报告中包含相关数据，且用户问题有明确指向，则须在报告开头设置“用户问题解答”章节，针对问题作出直接、明确的回答，并予以标注。"""


# ---------------------------------------------------------------------------
# Excel 文件解析
# ---------------------------------------------------------------------------

def _parse_xls_to_markdown(file_path: str) -> Optional[str]:
    """解析 .xls 文件（旧版 Excel 97-2003 格式），需要 xlrd 库。"""
    try:
        import xlrd
    except ImportError:
        print(msg('insight_excel_no_xlrd'), flush=True)
        return None

    try:
        wb = xlrd.open_workbook(file_path)
    except Exception as e:
        print(msg('insight_excel_open_failed', exc=e), flush=True)
        return None

    md_parts: List[str] = []

    for sheet in wb.sheets():
        if sheet.nrows == 0:
            continue

        rows: List[List[str]] = []
        for row_idx in range(sheet.nrows):
            rows.append([str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)])

        headers = rows[0]
        col_count = len(headers)

        lines: List[str] = []
        lines.append(f"## {sheet.name}\n")
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join([" --- "] * col_count) + "|")

        for data_row in rows[1:]:
            padded = data_row + [""] * (col_count - len(data_row))
            lines.append("| " + " | ".join(padded[:col_count]) + " |")

        md_parts.append("\n".join(lines))

    if not md_parts:
        print(msg('insight_excel_no_data'), flush=True)
        return None

    markdown_text = "\n\n".join(md_parts)
    print(msg('insight_excel_parsed', sheets=wb.nsheets, chars=len(markdown_text)), flush=True)
    return markdown_text


def _parse_xlsx_to_markdown(file_path: str) -> Optional[str]:
    """解析 .xlsx 文件（Office Open XML 格式），需要 openpyxl 库。"""
    try:
        import openpyxl
    except ImportError:
        print(msg('insight_excel_no_openpyxl'), flush=True)
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(msg('insight_excel_open_failed', exc=e), flush=True)
        return None

    md_parts: List[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            continue

        headers = rows[0]
        col_count = len(headers)

        lines: List[str] = []
        lines.append(f"## {sheet_name}\n")
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join([" --- "] * col_count) + "|")

        for data_row in rows[1:]:
            padded = data_row + [""] * (col_count - len(data_row))
            lines.append("| " + " | ".join(padded[:col_count]) + " |")

        md_parts.append("\n".join(lines))

    wb.close()

    if not md_parts:
        print(msg('insight_excel_no_data'), flush=True)
        return None

    markdown_text = "\n\n".join(md_parts)
    print(msg('insight_excel_parsed', sheets=len(wb.sheetnames), chars=len(markdown_text)), flush=True)
    return markdown_text


def parse_excel_to_markdown(file_path: str) -> Optional[str]:
    """
    将 Excel 文件解析为 Markdown 表格文本。

    自动识别 .xls / .xlsx 格式，支持多 Sheet。
    返回合并后的 Markdown 文本；解析失败时返回 None。
    """
    if not os.path.isfile(file_path):
        print(msg('insight_excel_not_found', path=file_path), flush=True)
        return None

    # 文件大小校验
    file_size = os.path.getsize(file_path)
    if file_size > MAX_EXCEL_FILE_SIZE:
        print(msg('insight_excel_size_exceeded', name=os.path.basename(file_path), size=f"{file_size / 1024 / 1024:.1f}"), flush=True)
        print("[STOP] Do NOT preprocess, filter, split, or compress the file locally. Show the above message to the user AS-IS and terminate.", flush=True)
        sys.exit(1)

    print(msg('insight_excel_parsing', path=file_path), flush=True)

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        return _parse_xls_to_markdown(file_path)
    elif ext == ".xlsx":
        return _parse_xlsx_to_markdown(file_path)
    else:
        print(msg('insight_excel_unsupported', ext=ext), flush=True)
        return None


# ---------------------------------------------------------------------------
# 仪表板快照
# ---------------------------------------------------------------------------

def call_snapshot(
    works_id: str,
    user_id: str,
    *,
    config: Optional[dict] = None,
) -> dict:
    """
    调用 POST /openapi/v2/snapshot/calling/shot 拉取仪表板快照。

    OpenAPI 统一响应结构 {success, data: SnapshotResultModel, code, message}，
    本函数在拦截业务错误码后返回 `data` 子对象（SnapshotResultModel）：
      {status: "PROCESSING"|"SUCCESS"|"FAILED", errorInfo: {...}, resultMarkdownText: "..."}
    """
    payload: Dict[str, Any] = {
        "worksId": works_id,
        "worksType": "dashboard",
        "userId": user_id,
        "targetType": "excel",
    }
    resp = request_openapi(
        "POST",
        SNAPSHOT_URI,
        json_body=payload,
        timeout=60,
        config=config,
    )
    envelope = resp.json()
    # 拦截业务错误码（席位/额度/权限/试用过期等）
    if isinstance(envelope, dict) and envelope.get("success") is False:
        check_known_error_code(envelope)
    # 解包外层 envelope，返回 SnapshotResultModel（各业务字段在 data 子对象下）
    if isinstance(envelope, dict) and isinstance(envelope.get("data"), dict):
        return envelope["data"]
    # 兼容：某些异常路径下服务端可能直接返回业务体
    return envelope if isinstance(envelope, dict) else {}


def resolve_works_id_from_url(
    url: str,
    *,
    config: Optional[dict] = None,
) -> Optional[str]:
    """
    从仪表板 URL 或数据门户 URL 解析出 worksId（= pageId）。

    支持两种格式（复用 dashboard 模块的解析逻辑，保证与仪表板生成技能路由一致）：
      1. 仪表板 URL：https://.../dashboard/view/pc.htm?pageId=xxx
      2. 数据门户 URL：https://.../product/view.htm?productId=xxx&menuId=yyy

    Returns:
        解析成功返回 pageId 字符串；失败返回 None 并打印提示（调用方需终止流程）
    """
    config = config or read_config()
    print(msg('insight_url_parsing', url=url), flush=True)

    try:
        if is_dataportal_url(url):
            # 数据门户 URL：先拿 productId/menuId，再查菜单树拿 pageId
            portal_ids = extract_dataportal_ids(url)
            portal_result = get_dataportal_page_id(
                host=config["server_domain"],
                access_id=config["api_key"],
                access_key=config["api_secret"],
                dataportal_id=portal_ids["productId"],
                menu_id=portal_ids["menuId"],
            )
            if not portal_result.get("success"):
                print(msg('insight_url_parse_failed',
                          url=url,
                          error=portal_result.get("error_message", "unknown")), flush=True)
                return None
            return portal_result["page_id"]
        else:
            # 仪表板 URL：直接提取 pageId
            return extract_page_id(url)
    except ValueError as e:
        print(msg('insight_url_parse_failed', url=url, error=str(e)), flush=True)
        return None
    except Exception as e:
        print(msg('insight_url_parse_failed', url=url, error=str(e)), flush=True)
        return None


def poll_snapshot(
    works_id: str,
    user_id: str,
    *,
    config: Optional[dict] = None,
) -> Optional[str]:
    """
    轮询快照接口直到 status 为 success 或 failed。

    返回 resultMarkdownText（成功时）或 None（失败 / 超时时）。
    """
    for attempt in range(1, MAX_POLL_COUNT + 1):
        result = call_snapshot(works_id, user_id, config=config)

        # 服务端状态枚举返回大写（SUCCESS / PROCESSING / FAILED），此处做大小写不敏感处理以增强健壮性
        status_raw = result.get("status", "")
        status = str(status_raw).strip().lower()
        print(msg('insight_snapshot_poll', attempt=attempt, status=status_raw), flush=True)

        if status == "success":
            markdown_text = result.get("resultMarkdownText", "")
            if not markdown_text:
                print(msg('insight_snapshot_empty'), flush=True)
                return None
            print(msg('insight_snapshot_ok', chars=len(markdown_text)), flush=True)
            return markdown_text

        if status == "failed":
            error_info = result.get("errorInfo", {})
            print(msg('insight_snapshot_failed', error=json.dumps(error_info, ensure_ascii=False)), flush=True)
            print(msg('insight_snapshot_failed_contact'), flush=True)
            return None

        if status == "processing":
            time.sleep(POLL_INTERVAL_SECONDS)
            continue

        print(msg('insight_snapshot_unknown', status=status_raw, response=json.dumps(result, ensure_ascii=False)), flush=True)
        return None

    print(msg('insight_snapshot_timeout', seconds=MAX_POLL_COUNT * POLL_INTERVAL_SECONDS), flush=True)
    return None


# ---------------------------------------------------------------------------
# 数据解读（SSE 流式）
# ---------------------------------------------------------------------------

def run_interpretation_stream(
    string_data: str,
    user_question: str,
    *,
    locale: str,
    config: Optional[dict] = None,
):
    """
    调用 POST /openapi/v2/smartq/dataInterpretationStream 进行数据解读。

    实时解析 SSE 事件并输出推理过程和解读结果。
    """
    config = config or read_config()
    oapi_user_id = require_user_id(config)

    payload: Dict[str, Any] = {
        "stringData": base64.b64encode(string_data.encode("utf-8")).decode("utf-8"),
        "userQuestion": ACC_PROMPT + user_question,
        "oapiUserId": oapi_user_id,
        "locale": locale,
        "runningBySkill": True,
    }

    print(f"\n{'=' * 60}", flush=True)
    print(msg('insight_interpretation_question', question=user_question), flush=True)
    print(msg('insight_dashboard_locale', locale=locale), flush=True)
    print(f"{'=' * 60}\n", flush=True)

    reasoning_buf: List[str] = []
    text_buf: List[str] = []

    try:
        for raw_event in request_openapi_stream(
            INTERPRETATION_URI, json_body=payload, config=config, timeout=600
        ):
            event_data = parse_sse_event(raw_event)
            if not event_data:
                continue

            event_type = event_data.get("type", "")
            data = event_data.get("data", "")

            if event_type in ("heartbeat", "trace", "locale", "message"):
                continue

            elif event_type == "reasoning":
                reasoning_buf.append(str(data))

            elif event_type in ("text", "summary"):
                text_buf.append(str(data))

            elif event_type == "error":
                print(msg('insight_interpretation_error', data=data), flush=True)
                check_known_error_code(str(data))

            elif event_type == "finish":
                break

            # 其他未知事件静默丢弃，避免污染输出

    except Exception as e:
        print(msg('insight_interpretation_failed', exc=e), flush=True)
        check_known_error_code(str(e))

    if reasoning_buf:
        reasoning_text = "".join(reasoning_buf)
        print(f"{msg('insight_reasoning')}\n{reasoning_text}\n", flush=True)

    if text_buf:
        interpretation_text = "".join(text_buf)
        print(f"{msg('insight_result')}\n{interpretation_text}", flush=True)
    else:
        print(msg('insight_no_result'), flush=True)

    print(msg('insight_done'), flush=True)
    return "".join(text_buf)


# ---------------------------------------------------------------------------
# 完整流程
# ---------------------------------------------------------------------------

def run_insights(
    question: str,
    works_id: Optional[str] = None,
    *,
    excel_file: Optional[str] = None,
    dashboard_url: Optional[str] = None,
    locale: str,
    config: Optional[dict] = None,
):
    """
    执行小Q解读的完整流程。

    数据来源优先级（三者互斥，取首个非空值）：
      excel_file > dashboard_url > works_id
    全部为空时返回错误。
    """
    config = config or read_config()

    # 统一调用 require_user_id，确保试用注册流程正常执行
    user_id = require_user_id(config)

    print(f"{'=' * 60}", flush=True)
    print(msg('insight_question', question=question), flush=True)

    # 优先从 Excel 文件获取数据
    if excel_file:
        print(msg('insight_source_excel', path=excel_file), flush=True)
        print(f"{'=' * 60}\n", flush=True)
        markdown_text = parse_excel_to_markdown(excel_file)
    elif dashboard_url:
        # 仪表板 / 门户 URL 路径：先解析出 pageId，再走快照轮询
        print(msg('insight_source_dashboard_url', url=dashboard_url), flush=True)
        print(f"{'=' * 60}\n", flush=True)
        resolved_works_id = resolve_works_id_from_url(dashboard_url, config=config)
        if not resolved_works_id:
            print(msg('insight_no_data'), flush=True)
            sys.exit(1)
        print(msg('insight_url_resolved', page_id=resolved_works_id), flush=True)
        print(msg('insight_snapshot_start'), flush=True)
        markdown_text = poll_snapshot(resolved_works_id, user_id, config=config)
    elif works_id:
        print(msg('insight_source_snapshot', works_id=works_id), flush=True)
        print(f"{'=' * 60}\n", flush=True)
        print(msg('insight_snapshot_start'), flush=True)
        markdown_text = poll_snapshot(works_id, user_id, config=config)
    else:
        print(f"{'=' * 60}\n", flush=True)
        print(msg('insight_no_source'), flush=True)
        sys.exit(1)

    if not markdown_text:
        print(f"\n{'=' * 60}", flush=True)
        print(msg('insight_no_data'), flush=True)
        print(f"{'=' * 60}", flush=True)
        sys.exit(1)

    # 数据超限检查：超限时报错终止，要求 Agent 先做数据过滤 / 拆分
    if len(markdown_text) > MAX_MARKDOWN_CHARS:
        print(msg('insight_data_too_large',
                  chars=len(markdown_text),
                  limit=MAX_MARKDOWN_CHARS), flush=True)
        sys.exit(1)

    # 调用数据解读流式接口
    result = run_interpretation_stream(
        string_data=markdown_text,
        user_question=question,
        locale=locale,
        config=config
    )

    print(f"\n{'=' * 60}", flush=True)
    print(msg('insight_workflow_done'), flush=True)
    print(f"{'=' * 60}", flush=True)

    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Quick BI 小Q解读")
    parser.add_argument("question", help="用户的解读问题")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--excel-file", dest="excel_file",
                       help="Excel 文件路径（.xls / .xlsx）")
    group.add_argument("--dashboard-url", dest="dashboard_url",
                       help="仪表板 URL 或数据门户 URL（自动解析出 pageId）")
    group.add_argument("--works-id", dest="works_id",
                       help="已知 pageId 时可直接传入（跳过 URL 解析）")
    parser.add_argument("--locale", required=True, choices=["zh_CN", "en_US"],
                       help="语言环境: zh_CN(简体中文) 或 en_US(英文)")
    parser.add_argument("--workspace-dir", default=None, help="用户工作目录路径")
    args = parser.parse_args()

    if args.workspace_dir:
        set_workspace_dir(args.workspace_dir)

    set_locale(args.locale)
    run_insights(
        args.question,
        works_id=args.works_id,
        excel_file=args.excel_file,
        dashboard_url=args.dashboard_url,
        locale=args.locale,
    )


if __name__ == "__main__":
    main()
